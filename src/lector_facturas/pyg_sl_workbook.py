from __future__ import annotations

from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:  # pragma: no cover
    psycopg = None
    dict_row = None

from lector_facturas.fx_rates import EcbFxService, FxRateAuditRow


COMPANY_CODE = "SL"
COMPANY_NAME = "ARTESTA STORE, S.L."
REPORTING_CURRENCY = "EUR"
DISPLAY_TIMEZONE = ZoneInfo("Europe/Madrid")
MONTH_NAMES_ES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
DEFAULT_SHOPIFY_MARKETS = ["ES", "FR", "DE", "IT", "AT", "BE", "NL", "PT", "XX"]
DEFAULT_SERVICE_LINES = ["HANNUN", "QHANDS", "Ltd", "Inc"]
DEFAULT_PAYMENT_FEE_LINES = ["SHOPIFY", "PAYPAL"]
DEFAULT_MARKETING_REGIONS = ["EU", "UK", "US"]
ADMINISTRATION_DETAIL_LINES = {"HANNUN": ["administration", "office", "services"]}
MONEY_FORMAT = '#,##0.00;[Red](#,##0.00);-'
PERCENT_FORMAT = '0.0%;[Red](0.0%);-'
RATIO_FORMAT = '0.00;[Red](0.00);-'
ROW_HEIGHT = 12
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
KPI_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBSECTION_FILL = PatternFill("solid", fgColor="EAF3F8")
TOTAL_FILL = PatternFill("solid", fgColor="D9D9D9")
SUBTOTAL_FILL = PatternFill("solid", fgColor="F3F6F8")
PERCENT_ROW_FILL = PatternFill("solid", fgColor="FAFAFA")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
TITLE_FONT = Font(size=12, bold=True, color="FFFFFF")
INFO_FONT = Font(size=9, italic=True, color="666666")
THIN_TOP_BORDER = Border(top=Side(style="thin", color="A6A6A6"))
MEDIUM_TOP_BORDER = Border(top=Side(style="medium", color="7F7F7F"))


@dataclass(frozen=True)
class StageRow:
    yyyymm: str
    entity: str
    line_item: str
    detail: str
    amount_net: Decimal
    currency: str
    source: str
    invoice_number: str = ""
    drive_url: str = ""


@dataclass(frozen=True)
class ExpenseRow:
    yyyymm: str
    entity: str
    category: str
    subcategory: str
    supplier_code: str
    detail: str
    amount_net: Decimal
    currency: str
    source: str
    invoice_number: str = ""
    drive_url: str = ""


@dataclass(frozen=True)
class PaymentFeeRow:
    yyyymm: str
    entity: str
    supplier_code: str
    amount_net: Decimal
    currency: str
    source: str


@dataclass(frozen=True)
class ProviderCatalogRow:
    supplier_code: str
    supplier_name: str
    current_folder: str
    destination_path: str
    notes: str


@dataclass(frozen=True)
class PygSlDataBundle:
    year: int
    generated_at: datetime
    shopify_rows: tuple[StageRow, ...]
    marketplace_rows: tuple[StageRow, ...]
    rappel_rows: tuple[StageRow, ...]
    supplies_rows: tuple[StageRow, ...]
    service_rows: tuple[StageRow, ...]
    expense_rows: tuple[ExpenseRow, ...]
    payment_fee_rows: tuple[PaymentFeeRow, ...]
    provider_catalog_rows: tuple[ProviderCatalogRow, ...]
    shopify_markets: tuple[str, ...]
    fx_rate_rows: tuple[FxRateAuditRow, ...] = field(default_factory=tuple)
    otros_ingresos_by_period: dict[str, Decimal] = field(default_factory=dict)
    diferencias_divisas_by_period: dict[str, Decimal] = field(default_factory=dict)
    # scope → {yyyymm → gross_amount}: royalties desglosados por región (uk, us)
    royalties_by_scope: dict[str, dict[str, Decimal]] = field(default_factory=dict)


def _normalize_company_name(value: str) -> str:
    return "".join(ch for ch in (value or "").upper() if ch.isalnum())


def default_output_path(root: Path, year: int) -> Path:
    return root / "output" / "spreadsheet" / f"pyg_sl_{year}.xlsx"


def month_keys(year: int) -> list[str]:
    return [f"{year}{month:02d}" for month in range(1, 13)]


def collect_pyg_sl_data(*, year: int, database_url: str | None) -> PygSlDataBundle:
    if not database_url or psycopg is None:
        return PygSlDataBundle(year, datetime.now(UTC), (), (), (), (), (), (), (), (), tuple(DEFAULT_SHOPIFY_MARKETS))
    with psycopg.connect(database_url, row_factory=dict_row) as conn:
        shared_services = conn.execute(
            """
            SELECT
                period_yyyymm,
                company_code,
                supplier_code,
                billed_company_name,
                division_invoice,
                currency_code,
                net_amount AS amount_net,
                invoice_number,
                drive_url
            FROM invoices.documents
            WHERE period_yyyymm LIKE %(period)s
              AND company_code IN ('LTD', 'INC')
              AND supplier_code = 'SHAREDSERVICESSL'
              AND status = 'classified'
            ORDER BY period_yyyymm, company_code, invoice_number
            """,
            {"period": f"{year}%"},
        ).fetchall()
        sales = conn.execute(
            """
            SELECT
                order_month_yyyymm,
                shipping_country_code,
                payment_currency,
                COALESCE(is_rever_tag, 0) AS is_rever_tag,
                COALESCE(is_hannun_tag, 0) AS is_hannun_tag,
                COALESCE(is_mirakl_tag, 0) AS is_mirakl_tag,
                SUM(net) AS amount_net
            FROM finance.ventas_pyg
            WHERE order_month_yyyymm LIKE %(period)s
            GROUP BY order_month_yyyymm, shipping_country_code, payment_currency, COALESCE(is_rever_tag, 0), COALESCE(is_hannun_tag, 0), COALESCE(is_mirakl_tag, 0)
            ORDER BY order_month_yyyymm, shipping_country_code, payment_currency
            """,
            {"period": f"{year}%"},
        ).fetchall()
        docs = conn.execute(
            """
            SELECT period_yyyymm, supplier_code, billed_company_name, division_invoice, document_type, currency_code, net_amount AS amount_net, invoice_number, drive_url, billing_period_end, invoice_date, parser_name
            FROM invoices.documents
            WHERE company_code = %(company)s
              AND status = 'classified'
              AND (
                period_yyyymm LIKE %(period)s
                OR (
                    supplier_code = 'RAILWAY'
                    AND billing_period_end >= %(year_start)s
                    AND billing_period_end < %(year_end)s
                )
              )
            ORDER BY period_yyyymm, supplier_code, division_invoice, invoice_number
            """,
            {
                "company": COMPANY_CODE,
                "period": f"{year}%",
                "year_start": f"{year}-01-01T00:00:00+00:00",
                "year_end": f"{year + 1}-01-01T00:00:00+00:00",
            },
        ).fetchall()
        suppliers = conn.execute(
            """
            SELECT supplier_code, supplier_name, current_folder, destination_path, notes
            FROM invoices.suppliers
            WHERE company_code = %(company)s AND is_active = TRUE
            ORDER BY destination_path, supplier_code
            """,
            {"company": COMPANY_CODE},
        ).fetchall()
        royalties = conn.execute(
            """
            SELECT period_yyyymm, gross_amount
            FROM invoices.artist_royalties_monthly_summary
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s AND summary_scope = 'total'
            ORDER BY period_yyyymm
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        royalties_scope = conn.execute(
            """
            SELECT period_yyyymm, summary_scope, gross_amount
            FROM invoices.artist_royalties_monthly_summary
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s AND summary_scope IN ('uk', 'us', 'eu')
            ORDER BY period_yyyymm, summary_scope
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        payroll = conn.execute(
            """
            SELECT period_yyyymm, currency_code, total_company_cost_amount
            FROM invoices.payroll_documents
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s
            ORDER BY period_yyyymm
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        otros_gastos = conn.execute(
            """
            SELECT period_yyyymm, amount_eur, notes
            FROM invoices.otros_gastos
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s
            ORDER BY period_yyyymm
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        otros_ingresos = conn.execute(
            """
            SELECT period_yyyymm, amount_eur
            FROM invoices.otros_ingresos
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s
            ORDER BY period_yyyymm
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        diferencias_divisas = conn.execute(
            """
            SELECT period_yyyymm, amount_eur
            FROM invoices.diferencias_divisas
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s
            ORDER BY period_yyyymm
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        payment_fees = conn.execute(
            """
            SELECT period_yyyymm, platform, currency_code, SUM(total_cost_amount) AS amount_net
            FROM invoices.payment_fee_monthly_summary
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s
            GROUP BY period_yyyymm, platform, currency_code
            ORDER BY period_yyyymm, platform
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
    supplier_map = {str(row["supplier_code"]): row for row in suppliers}
    shopify_rows: list[StageRow] = []
    marketplace_rows: list[StageRow] = []
    rappel_rows: list[StageRow] = []
    supplies_rows: list[StageRow] = []
    service_rows: list[StageRow] = []
    expense_rows: list[ExpenseRow] = []
    for row in sales:
        yyyymm = str(row["order_month_yyyymm"])
        shipping_country_code = str(row["shipping_country_code"] or "").upper()
        currency = str(row["payment_currency"] or "EUR")
        amount_net = _decimal(row["amount_net"])
        if shipping_country_code in {"GB", "US"}:
            continue
        if int(row["is_hannun_tag"] or 0) == 1:
            # HANNUN marketplace revenue comes from outgoing income invoices in invoices.documents,
            # not from ventas_pyg, to keep the PyG aligned with booked revenue.
            continue
        if int(row["is_rever_tag"] or 0) == 1:
            supplies_rows.append(StageRow(yyyymm, COMPANY_CODE, "REVER", _normalize_shopify_market(shipping_country_code), -amount_net, currency, "finance.ventas_pyg"))
            continue
        shopify_rows.append(StageRow(yyyymm, COMPANY_CODE, _normalize_shopify_market(shipping_country_code), shipping_country_code or "XX", amount_net, currency, "finance.ventas_pyg"))
    for row in _filter_periodified_documents(docs):
        supplier_code = str(row["supplier_code"])
        amount_net = _decimal(row["amount_net"])
        billed_company_name = str(row["billed_company_name"] or "")
        is_outgoing = _normalize_company_name(billed_company_name) != _normalize_company_name(COMPANY_NAME)
        yyyymm = str(row["period_yyyymm"])
        currency = str(row["currency_code"] or "EUR")
        division_invoice = str(row["division_invoice"] or "")
        document_type = str(row["document_type"] or "")
        invoice_number = str(row["invoice_number"] or "")
        drive_url = str(row["drive_url"] or "")
        invoice_date = row.get("invoice_date")
        billing_period_end = row.get("billing_period_end")
        if supplier_code == "RAILWAY" and invoice_date:
            effective_yyyymm = _period_from_timestamp(invoice_date)
        elif (
            supplier_code == "HANNUN"
            and division_invoice.lower() in {"administration", "office", "services"}
            and billing_period_end
        ):
            effective_yyyymm = _period_from_timestamp(billing_period_end)
        else:
            effective_yyyymm = yyyymm
        if is_outgoing:
            if supplier_code in {"TOASTY", "CHOOSE"}:
                marketplace_rows.append(StageRow(yyyymm, COMPANY_CODE, supplier_code, division_invoice or supplier_code.lower(), amount_net, currency, "documents", invoice_number, drive_url))
            elif supplier_code == "LIVITUM":
                rappel_rows.append(StageRow(yyyymm, COMPANY_CODE, "LIVITUM", division_invoice or "rappels", amount_net, currency, "documents", invoice_number, drive_url))
            elif supplier_code == "QHANDS":
                service_rows.append(StageRow(yyyymm, COMPANY_CODE, "QHANDS", division_invoice or "renting_cnc", amount_net, currency, "documents", invoice_number, drive_url))
            elif supplier_code == "HANNUN" and division_invoice.lower() == "orders":
                marketplace_rows.append(StageRow(yyyymm, COMPANY_CODE, "HANNUN", division_invoice or "orders", amount_net, currency, "documents", invoice_number, drive_url))
            elif supplier_code == "HANNUN":
                service_rows.append(StageRow(yyyymm, COMPANY_CODE, "HANNUN", division_invoice or "services", amount_net, currency, "documents", invoice_number, drive_url))
            continue
        if supplier_code == "REVER" and (document_type == "supplied_note" or division_invoice == "suplidos"):
            supplies_rows.append(StageRow(yyyymm, COMPANY_CODE, "REVER", "suplidos", -amount_net, currency, "documents", invoice_number, drive_url))
            continue
        supplier_meta = supplier_map.get(supplier_code)
        if supplier_meta and str(supplier_meta["destination_path"]).startswith("expenses/"):
            _, category, subcategory = str(supplier_meta["destination_path"]).split("/", 2)
            expense_rows.append(ExpenseRow(effective_yyyymm, COMPANY_CODE, category, subcategory, supplier_code, division_invoice.lower(), amount_net, currency, "documents", invoice_number, drive_url))
    for row in royalties:
        expense_rows.append(ExpenseRow(str(row["period_yyyymm"]), COMPANY_CODE, "cogs", "royalties", "ROYALTIES", "", _decimal(row["gross_amount"]), "EUR", "artist_royalties_monthly_summary"))
    for row in payroll:
        expense_rows.append(ExpenseRow(str(row["period_yyyymm"]), COMPANY_CODE, "opex", "staff", "PAYROLL", "", _decimal(row["total_company_cost_amount"]), str(row["currency_code"] or "EUR"), "payroll_documents"))
    for row in otros_gastos:
        expense_rows.append(ExpenseRow(str(row["period_yyyymm"]), COMPANY_CODE, "opex", "otros_gastos", "OTROSGASTOS", str(row["notes"] or ""), _decimal(row["amount_eur"]), "EUR", "otros_gastos"))
    for row in shared_services:
        service_line = "Ltd" if str(row["company_code"]).upper() == "LTD" else "Inc"
        service_rows.append(
            StageRow(
                str(row["period_yyyymm"]),
                COMPANY_CODE,
                service_line,
                str(row["division_invoice"] or "shared_services").lower(),
                _decimal(row["amount_net"]),
                str(row["currency_code"] or "EUR"),
                "documents:shared_services",
                str(row["invoice_number"] or ""),
                str(row["drive_url"] or ""),
            )
        )
    payment_fee_rows = tuple(
        PaymentFeeRow(
            str(row["period_yyyymm"]),
            COMPANY_CODE,
            str(row["platform"]).upper(),
            abs(_decimal(row["amount_net"])),
            str(row["currency_code"] or "EUR"),
            "payment_fee_monthly_summary",
        )
        for row in payment_fees
    )
    provider_rows = tuple(ProviderCatalogRow(str(row["supplier_code"]), str(row["supplier_name"]), str(row["current_folder"]), str(row["destination_path"]), str(row["notes"] or "")) for row in suppliers)
    marketplace_rows = _dedupe_stage_rows(marketplace_rows)
    supplies_rows = _dedupe_stage_rows(supplies_rows)
    service_rows = _dedupe_stage_rows(service_rows)
    otros_ingresos_by_period = {str(row["period_yyyymm"]): _decimal(row["amount_eur"]) for row in otros_ingresos}
    diferencias_divisas_by_period = {str(row["period_yyyymm"]): _decimal(row["amount_eur"]) for row in diferencias_divisas}
    royalties_by_scope: dict[str, dict[str, Decimal]] = {}
    for row in royalties_scope:
        scope = str(row["summary_scope"])
        royalties_by_scope.setdefault(scope, {})[str(row["period_yyyymm"])] = _decimal(row["gross_amount"])
    return PygSlDataBundle(year, datetime.now(UTC), tuple(shopify_rows), tuple(marketplace_rows), tuple(rappel_rows), tuple(supplies_rows), tuple(service_rows), tuple(expense_rows), payment_fee_rows, provider_rows, tuple(DEFAULT_SHOPIFY_MARKETS), otros_ingresos_by_period=otros_ingresos_by_period, diferencias_divisas_by_period=diferencias_divisas_by_period, royalties_by_scope=royalties_by_scope)


def build_pyg_sl_workbook(bundle: PygSlDataBundle, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    fx_service = EcbFxService()
    shopify_sheet_rows, shopify_fx_rows = _stage_rows_with_fx(bundle.shopify_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    marketplace_sheet_rows, marketplace_fx_rows = _stage_rows_with_fx(bundle.marketplace_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    rappel_sheet_rows, rappel_fx_rows = _stage_rows_with_fx(bundle.rappel_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    supplies_sheet_rows, supplies_fx_rows = _stage_rows_with_fx(bundle.supplies_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    service_sheet_rows, service_fx_rows = _stage_rows_with_fx(bundle.service_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    expense_sheet_rows, expense_fx_rows = _expense_rows_with_fx(bundle.expense_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    payment_fee_sheet_rows, payment_fee_fx_rows = _payment_fee_rows_with_fx(bundle.payment_fee_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    fx_rate_rows = _dedupe_fx_rows(
        bundle.fx_rate_rows
        + shopify_fx_rows
        + marketplace_fx_rows
        + rappel_fx_rows
        + supplies_fx_rows
        + service_fx_rows
        + expense_fx_rows
        + payment_fee_fx_rows
    )
    generated_at_utc = bundle.generated_at.astimezone(UTC) if bundle.generated_at.tzinfo else bundle.generated_at.replace(tzinfo=UTC)
    _sheet(wb, "params", ["key", "value"], [
        ["entity", COMPANY_CODE],
        ["year", bundle.year],
        ["generated_at_utc", generated_at_utc.replace(microsecond=0).isoformat().replace("+00:00", "Z")],
        ["pct_staff_uk", 0.15],
        ["pct_staff_us", 0.10],
        ["pct_admin_uk", 0.15],
        ["pct_admin_us", 0.10],
    ])
    _sheet(wb, "i-shopify-sl", ["yyyymm", "entity", "line_item", "detail", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source", "invoice_number", "drive_url"], shopify_sheet_rows)
    _sheet(wb, "i-marketplaces-sl", ["yyyymm", "entity", "line_item", "detail", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source", "invoice_number", "drive_url"], marketplace_sheet_rows)
    _sheet(wb, "i-rappels-sl", ["yyyymm", "entity", "line_item", "detail", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source", "invoice_number", "drive_url"], rappel_sheet_rows)
    _sheet(wb, "i-supplies-sl", ["yyyymm", "entity", "line_item", "detail", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source", "invoice_number", "drive_url"], supplies_sheet_rows)
    _sheet(wb, "i-services-sl", ["yyyymm", "entity", "line_item", "detail", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source", "invoice_number", "drive_url"], service_sheet_rows)
    _sheet(wb, "g-expenses-sl", ["yyyymm", "entity", "category", "subcategory", "supplier_code", "detail", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source", "invoice_number", "drive_url"], expense_sheet_rows)
    _sheet(wb, "g-payment-fees-sl", ["yyyymm", "entity", "supplier_code", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source"], payment_fee_sheet_rows)
    fx_rate_rows = _ensure_monthly_fx_rows(
        fx_rate_rows,
        year=bundle.year,
        fx_service=fx_service,
        currencies=("GBP", "USD"),
    )
    _sheet(wb, "fx-rates", ["yyyymm", "rate_date", "currency_original", "reporting_currency", "reference_rate", "fx_rate", "source"], [[r.yyyymm, r.rate_date, r.currency_original, r.reporting_currency, float(r.reference_rate), float(r.fx_rate), r.source] for r in fx_rate_rows])
    _sheet(wb, "catalog-sl", ["supplier_code", "supplier_name", "current_folder", "destination_path", "notes"], [[r.supplier_code, r.supplier_name, r.current_folder, r.destination_path, r.notes] for r in bundle.provider_catalog_rows])
    # Royalties desglosados por scope (uk, us) para shared services
    royalties_scope_rows: list[list[Any]] = []
    for scope, by_period in bundle.royalties_by_scope.items():
        for yyyymm, amount in sorted(by_period.items()):
            royalties_scope_rows.append([yyyymm, scope, float(amount)])
    _sheet(wb, "i-royalties-scope-sl", ["yyyymm", "scope", "gross_amount"], royalties_scope_rows)
    pos = _main_sheet(wb, bundle)
    _count_sheet_sl(wb, bundle)
    _add_back_links(wb)
    _shared_services_sheet(wb, bundle, pos)
    wb.save(output_path)
    return output_path


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(header) + 2)
        if header in {"amount_original", "amount_reporting"}:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).number_format = MONEY_FORMAT
        elif header in {"fx_rate", "reference_rate"}:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).number_format = "0.00000000"
    ws.freeze_panes = "A2"


def _add_back_links(wb: Workbook) -> None:
    detail_sheets = (
        "i-shopify-sl",
        "i-marketplaces-sl",
        "i-rappels-sl",
        "i-supplies-sl",
        "i-services-sl",
        "i-royalties-scope-sl",
        "g-expenses-sl",
        "g-payment-fees-sl",
        "fx-rates",
        "catalog-sl",
        "params",
    )
    for title in detail_sheets:
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        ws.insert_rows(1)
        ws["A1"] = "<- Volver a P&G-SL"
        ws["A1"]._hyperlink = Hyperlink(ref="A1", location="'P&G-SL'!A1", display="<- Volver a P&G-SL")
        ws["A1"].font = Font(bold=True, color="1F1F1F")
        ws["A1"].fill = SUBTOTAL_FILL
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = ROW_HEIGHT
        _apply_invoice_links(ws)
        ws.freeze_panes = "A3"


def _apply_invoice_links(ws) -> None:
    headers = [cell.value for cell in ws[2]]
    if "invoice_number" not in headers or "drive_url" not in headers:
        return
    invoice_col = headers.index("invoice_number") + 1
    drive_url_col = headers.index("drive_url") + 1
    for row_idx in range(3, ws.max_row + 1):
        invoice_cell = ws.cell(row=row_idx, column=invoice_col)
        drive_url_cell = ws.cell(row=row_idx, column=drive_url_col)
        invoice_cell.hyperlink = None
        if invoice_cell.value and drive_url_cell.value:
            invoice_cell.hyperlink = str(drive_url_cell.value)
            invoice_cell.style = "Hyperlink"
    ws.column_dimensions[get_column_letter(drive_url_col)].hidden = True


def _stage_rows_with_fx(
    rows: tuple[StageRow, ...],
    *,
    reporting_currency: str,
    fx_service: EcbFxService,
) -> tuple[list[list[Any]], tuple[FxRateAuditRow, ...]]:
    rendered: list[list[Any]] = []
    fx_rows: list[FxRateAuditRow] = []
    for row in rows:
        converted, audit = fx_service.convert(
            amount=row.amount_net,
            source_currency=row.currency,
            reporting_currency=reporting_currency,
            yyyymm=row.yyyymm,
        )
        rendered.append(
            [
                row.yyyymm,
                row.entity,
                row.line_item,
                row.detail,
                float(converted.amount_original),
                converted.currency_original,
                converted.reporting_currency,
                float(converted.fx_rate),
                float(converted.amount_reporting),
                row.source,
                row.invoice_number,
                row.drive_url,
            ]
        )
        fx_rows.append(audit)
    return rendered, tuple(fx_rows)


def _expense_rows_with_fx(
    rows: tuple[ExpenseRow, ...],
    *,
    reporting_currency: str,
    fx_service: EcbFxService,
) -> tuple[list[list[Any]], tuple[FxRateAuditRow, ...]]:
    rendered: list[list[Any]] = []
    fx_rows: list[FxRateAuditRow] = []
    for row in rows:
        converted, audit = fx_service.convert(
            amount=row.amount_net,
            source_currency=row.currency,
            reporting_currency=reporting_currency,
            yyyymm=row.yyyymm,
        )
        rendered.append(
            [
                row.yyyymm,
                row.entity,
                row.category,
                row.subcategory,
                row.supplier_code,
                row.detail,
                float(converted.amount_original),
                converted.currency_original,
                converted.reporting_currency,
                float(converted.fx_rate),
                float(converted.amount_reporting),
                row.source,
                row.invoice_number,
                row.drive_url,
            ]
        )
        fx_rows.append(audit)
    return rendered, tuple(fx_rows)


def _payment_fee_rows_with_fx(
    rows: tuple[PaymentFeeRow, ...],
    *,
    reporting_currency: str,
    fx_service: EcbFxService,
) -> tuple[list[list[Any]], tuple[FxRateAuditRow, ...]]:
    rendered: list[list[Any]] = []
    fx_rows: list[FxRateAuditRow] = []
    for row in rows:
        converted, audit = fx_service.convert(
            amount=row.amount_net,
            source_currency=row.currency,
            reporting_currency=reporting_currency,
            yyyymm=row.yyyymm,
        )
        rendered.append(
            [
                row.yyyymm,
                row.entity,
                row.supplier_code,
                float(converted.amount_original),
                converted.currency_original,
                converted.reporting_currency,
                float(converted.fx_rate),
                float(converted.amount_reporting),
                row.source,
            ]
        )
        fx_rows.append(audit)
    return rendered, tuple(fx_rows)


def _dedupe_fx_rows(rows: tuple[FxRateAuditRow, ...]) -> tuple[FxRateAuditRow, ...]:
    unique: dict[tuple[str, str, str], FxRateAuditRow] = {}
    for row in rows:
        unique[(row.yyyymm, row.currency_original, row.reporting_currency)] = row
    return tuple(sorted(unique.values(), key=lambda item: (item.yyyymm, item.currency_original, item.reporting_currency)))


def _ensure_monthly_fx_rows(
    rows: tuple[FxRateAuditRow, ...],
    *,
    year: int,
    fx_service: EcbFxService,
    currencies: tuple[str, ...],
) -> tuple[FxRateAuditRow, ...]:
    unique: dict[tuple[str, str, str], FxRateAuditRow] = {
        (row.yyyymm, row.currency_original, row.reporting_currency): row
        for row in rows
    }
    for yyyymm in month_keys(year):
        for currency in currencies:
            key = (yyyymm, currency, REPORTING_CURRENCY)
            if key in unique:
                continue
            _, audit = fx_service.convert(
                amount=Decimal("1"),
                source_currency=currency,
                reporting_currency=REPORTING_CURRENCY,
                yyyymm=yyyymm,
            )
            unique[key] = audit
    return tuple(sorted(unique.values(), key=lambda item: (item.yyyymm, item.currency_original, item.reporting_currency)))


def _main_sheet(wb: Workbook, bundle: PygSlDataBundle) -> dict[str, int]:
    ws = wb.create_sheet("P&G-SL", 0)
    for idx, yyyymm in enumerate(month_keys(bundle.year), start=4):
        ws.cell(row=1, column=idx, value=yyyymm)
        ws.cell(row=2, column=idx, value=MONTH_NAMES_ES[idx - 4])
    ws["P1"] = "TOTAL"
    ws["P2"] = "Total"
    ws["Q1"] = ""
    ws["Q2"] = ""
    ws["A2"] = f"P&G {COMPANY_CODE} {bundle.year}"
    ws.merge_cells("A2:C2")
    generated_at_local = bundle.generated_at.astimezone(DISPLAY_TIMEZONE) if bundle.generated_at.tzinfo else bundle.generated_at.replace(tzinfo=UTC).astimezone(DISPLAY_TIMEZONE)
    ws["A3"] = f"Actualizado: {generated_at_local.strftime('%d/%m/%Y %H:%M')} h"
    ws.merge_cells("A3:C3")
    for row in (1, 2):
        for cell in ws[row]:
            cell.fill = HEADER_FILL
            cell.font = WHITE_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = TITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(size=9, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].hidden = True
    ws.row_dimensions[2].height = ROW_HEIGHT
    ws.row_dimensions[3].height = ROW_HEIGHT
    groups = _provider_groups(bundle.provider_catalog_rows)
    pos: dict[str, int] = {}
    row = 4
    pos["turnover"] = row; ws[f"A{row}"] = "Turnover"; ws[f"A{row}"].font = BOLD; row += 1
    pos["product_sales"] = row; ws[f"B{row}"] = "Product sales"; ws[f"B{row}"].font = BOLD; row += 1
    pos["shopify_header"] = row; ws[f"C{row}"] = "Shopify"; ws[f"C{row}"].font = BOLD; row += 1
    shopify_rows = list(range(row, row + len(bundle.shopify_markets)))
    for idx, market in enumerate(bundle.shopify_markets):
        ws[f"C{row + idx}"] = market
    row += len(bundle.shopify_markets)
    pos["marketplaces_header"] = row; ws[f"C{row}"] = "Marketplaces"; ws[f"C{row}"].font = BOLD; row += 1
    marketplace_codes = ["HANNUN", "TOASTY", "CHOOSE"]
    marketplace_rows = list(range(row, row + len(marketplace_codes)))
    for idx, code in enumerate(marketplace_codes):
        ws[f"C{row + idx}"] = code
    row += len(marketplace_codes)
    pos["rappels_header"] = row; ws[f"C{row}"] = "Rappels"; ws[f"C{row}"].font = BOLD; row += 1
    pos["rappels_detail"] = row; ws[f"C{row}"] = "LIVITUM"; row += 1
    pos["supplies_header"] = row; ws[f"C{row}"] = "Supplies"; ws[f"C{row}"].font = BOLD; row += 1
    pos["supplies_detail"] = row; ws[f"C{row}"] = "REVER"; row += 2
    pos["services_header"] = row; ws[f"B{row}"] = "Services"; ws[f"B{row}"].font = BOLD; row += 1
    service_codes = ["HANNUN", "QHANDS", "Ltd", "Inc"]
    service_rows = list(range(row, row + len(service_codes)))
    for idx, code in enumerate(service_codes):
        ws[f"C{row + idx}"] = code
    row += len(service_codes)
    pos["otros_ingresos"] = row; ws[f"B{row}"] = "Uncategorized income"; ws[f"B{row}"].font = BOLD; row += 2

    pos["expenses"] = row; ws[f"A{row}"] = "Expenses"; ws[f"A{row}"].font = BOLD; row += 1
    pos["cogs"] = row; ws[f"B{row}"] = "COGS"; ws[f"B{row}"].font = BOLD; row += 1
    pos["manufacturing_header"] = row; ws[f"C{row}"] = "Manufacturing (% sales)"; ws[f"C{row}"].font = BOLD; row += 1
    manufacturing_rows = list(range(row, row + len(groups["manufacturing"])))
    for idx, supplier in enumerate(groups["manufacturing"]):
        ws[f"C{row + idx}"] = supplier
    row += len(manufacturing_rows)
    pos["manufacturing_pct"] = row; ws[f"C{row}"] = "% Manufacturing / sales"; ws[f"C{row}"].font = INFO_FONT; row += 1

    pos["logistics_header"] = row; ws[f"C{row}"] = "Logistics (% sales)"; ws[f"C{row}"].font = BOLD; row += 1
    logistics_rows = list(range(row, row + len(groups["logistics"])))
    for idx, supplier in enumerate(groups["logistics"]):
        ws[f"C{row + idx}"] = supplier
    row += len(logistics_rows)
    pos["logistics_pct"] = row; ws[f"C{row}"] = "% Logistics / sales"; ws[f"C{row}"].font = INFO_FONT; row += 1

    pos["royalties_header"] = row; ws[f"C{row}"] = "Royalties (% sales)"; ws[f"C{row}"].font = BOLD; row += 1
    pos["royalties_detail"] = row; ws[f"C{row}"] = "ROYALTIES"; row += 1
    pos["royalties_eu"] = row; ws[f"C{row}"] = "eu"; row += 1
    pos["royalties_uk"] = row; ws[f"C{row}"] = "uk"; row += 1
    pos["royalties_us"] = row; ws[f"C{row}"] = "us"; row += 1
    pos["royalties_pct"] = row; ws[f"C{row}"] = "% Royalties / sales"; ws[f"C{row}"].font = INFO_FONT; row += 1

    pos["payment_fees_header"] = row; ws[f"C{row}"] = "Payment fees (% sales)"; ws[f"C{row}"].font = BOLD; row += 1
    payment_fee_codes = ["SHOPIFY", "PAYPAL"]
    payment_fee_rows = list(range(row, row + len(payment_fee_codes)))
    for idx, code in enumerate(payment_fee_codes):
        ws[f"C{row + idx}"] = code
    row += len(payment_fee_rows)
    pos["payment_fees_pct"] = row; ws[f"C{row}"] = "% Payment fees / sales"; ws[f"C{row}"].font = INFO_FONT; row += 2

    pos["gross_margin"] = row; ws[f"A{row}"] = "GROSS MARGIN (SALES-MANUFACTURING)"; ws[f"A{row}"].font = BOLD; row += 1
    pos["gross_margin_pct"] = row; ws[f"A{row}"] = "% gross margin"; ws[f"A{row}"].font = INFO_FONT; row += 1
    pos["contributive_margin"] = row; ws[f"A{row}"] = "Contributive margin (product sales-COGS)"; ws[f"A{row}"].font = BOLD; row += 1
    pos["contributive_margin_pct"] = row; ws[f"A{row}"] = "% contributive margin"; ws[f"A{row}"].font = INFO_FONT; row += 2

    pos["opex"] = row; ws[f"B{row}"] = "Opex"; ws[f"B{row}"].font = BOLD; row += 1
    pos["marketing_header"] = row; ws[f"C{row}"] = "Marketing"; ws[f"C{row}"].font = BOLD; row += 1
    pos["marketing_meta"] = row; ws[f"C{row}"] = "METAADS"; ws[f"C{row}"].font = BOLD; row += 1
    marketing_meta_detail_rows = list(range(row, row + len(DEFAULT_MARKETING_REGIONS)))
    for idx, code in enumerate(DEFAULT_MARKETING_REGIONS):
        ws[f"C{row + idx}"] = code
    row += len(marketing_meta_detail_rows)
    pos["marketing_google"] = row; ws[f"C{row}"] = "GOOGLEADS"; ws[f"C{row}"].font = BOLD; row += 1
    marketing_google_detail_rows = list(range(row, row + len(DEFAULT_MARKETING_REGIONS)))
    for idx, code in enumerate(DEFAULT_MARKETING_REGIONS):
        ws[f"C{row + idx}"] = code
    row += len(marketing_google_detail_rows)
    marketing_rows = [pos["marketing_meta"], pos["marketing_google"]]
    pos["marketing_pct"] = row; ws[f"C{row}"] = "Sales / mkt EU"; ws[f"C{row}"].font = BOLD; row += 1
    pos["staff_header"] = row; ws[f"C{row}"] = "Staff"; ws[f"C{row}"].font = BOLD; row += 1
    staff_rows = [row, row + 1]
    ws[f"C{row}"] = "PAYROLL"; ws[f"C{row + 1}"] = "DOSCONSULTING"; row += 2
    pos["administration_header"] = row; ws[f"C{row}"] = "Administration"; ws[f"C{row}"].font = BOLD; row += 1
    administration_rows: list[int] = []
    administration_detail_rows: dict[int, list[int]] = {}
    for supplier in groups["administration"]:
        ws[f"C{row}"] = supplier
        administration_rows.append(row)
        row += 1
        detail_rows_for_supplier: list[int] = []
        for detail in ADMINISTRATION_DETAIL_LINES.get(supplier, []):
            ws[f"C{row}"] = detail
            detail_rows_for_supplier.append(row)
            row += 1
        if detail_rows_for_supplier:
            administration_detail_rows[administration_rows[-1]] = detail_rows_for_supplier
    pos["technology_header"] = row; ws[f"C{row}"] = "Technology"; ws[f"C{row}"].font = BOLD; row += 1
    technology_rows = list(range(row, row + len(groups["technology"])))
    for idx, supplier in enumerate(groups["technology"]):
        ws[f"C{row + idx}"] = supplier
    row += len(technology_rows)
    pos["otros_gastos"] = row; ws[f"C{row}"] = "Uncategorized Expenses"; ws[f"C{row}"].font = BOLD; row += 1
    pos["diferencias_divisas"] = row; ws[f"B{row}"] = "Currency Adjustment"; ws[f"B{row}"].font = BOLD; row += 2
    pos["profit"] = row; ws[f"A{row}"] = "PROFIT"; ws[f"A{row}"].font = BOLD; row += 1
    pos["profit_pct"] = row; ws[f"A{row}"] = "% profit / product sales"; ws[f"A{row}"].font = INFO_FONT

    for section_row in (pos["turnover"], pos["expenses"]):
        for col in range(1, 18):
            ws.cell(row=section_row, column=col).fill = SECTION_FILL
    for kpi_row in (pos["gross_margin"], pos["gross_margin_pct"], pos["contributive_margin"], pos["contributive_margin_pct"], pos["profit"], pos["profit_pct"]):
        for col in range(1, 18):
            ws.cell(row=kpi_row, column=col).fill = KPI_FILL

    for row_key in (
        "product_sales", "services_header", "cogs", "opex",
        "shopify_header", "marketplaces_header", "rappels_header", "supplies_header",
        "manufacturing_header", "logistics_header", "royalties_header", "payment_fees_header",
        "marketing_header", "staff_header", "administration_header", "technology_header",
        "otros_gastos", "otros_ingresos", "diferencias_divisas",
    ):
        for col in range(1, 18):
            ws.cell(row=pos[row_key], column=col).fill = SUBSECTION_FILL

    _fill_month_formulas(
        ws,
        pos=pos,
        shopify_rows=shopify_rows,
        marketplace_rows=marketplace_rows,
        service_rows=service_rows,
        manufacturing_rows=manufacturing_rows,
        logistics_rows=logistics_rows,
        payment_fee_rows=payment_fee_rows,
        marketing_rows=marketing_rows,
        marketing_meta_detail_rows=marketing_meta_detail_rows,
        marketing_google_detail_rows=marketing_google_detail_rows,
        staff_rows=staff_rows,
        administration_rows=administration_rows,
        administration_detail_rows=administration_detail_rows,
        technology_rows=technology_rows,
    )
    for idx, yyyymm in enumerate(month_keys(bundle.year), start=4):
        amount = bundle.otros_ingresos_by_period.get(yyyymm)
        if amount is not None:
            ws.cell(row=pos["otros_ingresos"], column=idx).value = float(amount)
        amount_dd = bundle.diferencias_divisas_by_period.get(yyyymm)
        if amount_dd is not None:
            ws.cell(row=pos["diferencias_divisas"], column=idx).value = float(amount_dd)
    # Fila 3: facturación diaria = Product sales / días transcurridos del mes
    # Meses pasados: dividir por días del mes. Mes actual: dividir por días+horas transcurridos.
    ps = pos["product_sales"]
    for col in [get_column_letter(i) for i in range(4, 16)]:
        d = f'DATE(VALUE(LEFT({col}$1,4)),VALUE(RIGHT({col}$1,2)),1)'
        ws[f"{col}3"] = (
            f'=IFERROR(IF({d}>TODAY(),"",'
            f'IF(EOMONTH({d},0)<TODAY(),'
            f'{col}{ps}/DAY(EOMONTH({d},0)),'
            f'{col}{ps}/MAX(NOW()-{d},1/24))),"")'
        )
    _apply_layout(
        ws,
        pos=pos,
        shopify_rows=shopify_rows,
        marketplace_rows=marketplace_rows,
        service_rows=service_rows,
        manufacturing_rows=manufacturing_rows,
        logistics_rows=logistics_rows,
        payment_fee_rows=payment_fee_rows,
        marketing_rows=marketing_rows,
        marketing_meta_detail_rows=marketing_meta_detail_rows,
        marketing_google_detail_rows=marketing_google_detail_rows,
        staff_rows=staff_rows,
        administration_rows=administration_rows,
        administration_detail_rows=administration_detail_rows,
        technology_rows=technology_rows,
    )
    _add_navigation_links(
        ws,
        pos=pos,
        shopify_rows=shopify_rows,
        marketplace_rows=marketplace_rows,
        service_rows=service_rows,
        manufacturing_rows=manufacturing_rows,
        logistics_rows=logistics_rows,
        payment_fee_rows=payment_fee_rows,
        marketing_rows=marketing_rows,
        marketing_meta_detail_rows=marketing_meta_detail_rows,
        marketing_google_detail_rows=marketing_google_detail_rows,
        staff_rows=staff_rows,
        administration_rows=administration_rows,
        administration_detail_rows=administration_detail_rows,
        technology_rows=technology_rows,
    )
    for row in range(3, pos["profit_pct"] + 1):
        for col in range(4, 18):
            ws.cell(row=row, column=col).number_format = (
                RATIO_FORMAT if row == pos["marketing_pct"] else
                PERCENT_FORMAT if row in {
                    pos["manufacturing_pct"], pos["logistics_pct"], pos["royalties_pct"], pos["payment_fees_pct"],
                    pos["gross_margin_pct"], pos["contributive_margin_pct"], pos["profit_pct"],
                } else MONEY_FORMAT
            )
    for col, width in (("A", 18), ("B", 20), ("C", 24)):
        ws.column_dimensions[col].width = width
    ws.column_dimensions["P"].width = 14
    ws.column_dimensions["Q"].width = 3
    for idx in range(4, 16):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.column_dimensions["O"].width = 14
    ws.freeze_panes = "D4"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False
    # Devolvemos pos para que build_pyg_sl_workbook pueda construir shared-services
    # DESPUÉS de _add_back_links (que hace insert_rows y desplazaría las referencias)
    return pos


def _shared_services_sheet(wb: Workbook, bundle: PygSlDataBundle, pos: dict[str, int]) -> None:
    """Pestaña de shared services: ajuste de gasto intracompañía SL → LTD (UK) e INC (US).

    Creada DESPUÉS de _add_back_links para que sus referencias de fórmula sean correctas.
    Fila 1: back-link manual (en lugar de insert_rows).
    Fila 2: yyyymm (oculta, referenciada como {col}$2 en las fórmulas).
    Fila 3: título + nombres de mes.
    Datos desde fila 4.
    """
    ws = wb.create_sheet("shared-services")

    # Fila 1: back-link manual (no usamos insert_rows para evitar corrupción de fórmulas)
    ws["A1"] = "<- Volver a P&G-SL"
    ws["A1"]._hyperlink = Hyperlink(ref="A1", location="'P&G-SL'!A1", display="<- Volver a P&G-SL")
    ws["A1"].font = Font(bold=True, color="1F1F1F")
    ws["A1"].fill = SUBTOTAL_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Fila 2 (oculta): yyyymm → referenciada como {col}$2 en fórmulas
    for idx, yyyymm in enumerate(month_keys(bundle.year), start=4):
        ws.cell(row=2, column=idx, value=yyyymm)
        ws.cell(row=3, column=idx, value=MONTH_NAMES_ES[idx - 4])
    ws["P2"] = "TOTAL"
    ws["P3"] = "Total"
    ws["A3"] = f"Shared Services SL → LTD / INC {bundle.year}"
    ws.merge_cells("A3:C3")

    # Posiciones de fila (datos desde fila 4)
    LTD_HDR   = 4
    LTD_MK    = 5
    LTD_RY    = 6
    LTD_ST    = 7
    LTD_AD    = 8
    LTD_EUR   = 9
    LTD_GBP   = 10
    INC_HDR   = 12
    INC_MK    = 13
    INC_RY    = 14
    INC_ST    = 15
    INC_AD    = 16
    INC_EUR   = 17
    INC_USD   = 18
    TOTAL_EUR = 20  # Total shared services SL→LTD+INC en EUR (fila 19 vacía)

    # Etiquetas
    ws[f"A{LTD_HDR}"] = "LTD (UK)"
    ws[f"A{LTD_MK}"]  = "Marketing"
    ws[f"A{LTD_RY}"]  = "Royalties"
    ws[f"A{LTD_ST}"]  = "Staff"
    ws[f"A{LTD_AD}"]  = "Administrativos"
    ws[f"A{LTD_EUR}"] = "TOTAL EUR"
    ws[f"A{LTD_GBP}"] = "TOTAL GBP"
    ws[f"A{INC_HDR}"] = "INC (US)"
    ws[f"A{INC_MK}"]  = "Marketing"
    ws[f"A{INC_RY}"]  = "Royalties"
    ws[f"A{INC_ST}"]  = "Staff"
    ws[f"A{INC_AD}"]  = "Administrativos"
    ws[f"A{INC_EUR}"] = "TOTAL EUR"
    ws[f"A{INC_USD}"] = "TOTAL USD"
    ws[f"A{TOTAL_EUR}"] = "TOTAL SHARED SERVICES EUR"

    st_row = pos["staff_header"]
    ad_row = pos["administration_header"]
    tech_row = pos["technology_header"]

    def p(key: str) -> str:
        """VLOOKUP al sheet params para obtener el parámetro configurable."""
        return f'VLOOKUP("{key}",params!$A:$B,2,0)'

    for col in [get_column_letter(i) for i in range(4, 16)]:
        # ── LTD (UK) ──────────────────────────────────────────────────────────
        # Marketing UK: gasto real de g-expenses-sl con subcategory="marketing" y detail="uk"
        ws[f"{col}{LTD_MK}"] = (
            f"=SUMIFS('g-expenses-sl'!$K:$K,'g-expenses-sl'!$A:$A,{col}$2,"
            f"'g-expenses-sl'!$D:$D,\"marketing\",'g-expenses-sl'!$F:$F,\"uk\")"
        )
        # Royalties UK: gasto real por scope desde i-royalties-scope-sl
        ws[f"{col}{LTD_RY}"] = (
            f"=SUMIFS('i-royalties-scope-sl'!$C:$C,'i-royalties-scope-sl'!$A:$A,{col}$2,"
            f"'i-royalties-scope-sl'!$B:$B,\"uk\")"
        )
        # Staff UK: % configurable del total staff SL
        ws[f"{col}{LTD_ST}"] = f"='P&G-SL'!{col}{st_row}*{p('pct_staff_uk')}"
        # Admin UK: % configurable del total (administration + technology) SL
        ws[f"{col}{LTD_AD}"] = f"=('P&G-SL'!{col}{ad_row}+'P&G-SL'!{col}{tech_row})*{p('pct_admin_uk')}"
        # Totales LTD
        ws[f"{col}{LTD_EUR}"] = f"=SUM({col}{LTD_MK}:{col}{LTD_AD})"
        ws[f"{col}{LTD_GBP}"] = (
            f"=IFERROR({col}{LTD_EUR}/AVERAGEIFS('fx-rates'!$F:$F,'fx-rates'!$A:$A,{col}$2,"
            f"'fx-rates'!$C:$C,\"GBP\",'fx-rates'!$D:$D,\"EUR\"),{col}{LTD_EUR})"
        )

        # ── INC (US) ──────────────────────────────────────────────────────────
        ws[f"{col}{INC_MK}"] = (
            f"=SUMIFS('g-expenses-sl'!$K:$K,'g-expenses-sl'!$A:$A,{col}$2,"
            f"'g-expenses-sl'!$D:$D,\"marketing\",'g-expenses-sl'!$F:$F,\"us\")"
        )
        ws[f"{col}{INC_RY}"] = (
            f"=SUMIFS('i-royalties-scope-sl'!$C:$C,'i-royalties-scope-sl'!$A:$A,{col}$2,"
            f"'i-royalties-scope-sl'!$B:$B,\"us\")"
        )
        ws[f"{col}{INC_ST}"] = f"='P&G-SL'!{col}{st_row}*{p('pct_staff_us')}"
        ws[f"{col}{INC_AD}"] = f"=('P&G-SL'!{col}{ad_row}+'P&G-SL'!{col}{tech_row})*{p('pct_admin_us')}"
        ws[f"{col}{INC_EUR}"] = f"=SUM({col}{INC_MK}:{col}{INC_AD})"
        ws[f"{col}{INC_USD}"] = (
            f"=IFERROR({col}{INC_EUR}/AVERAGEIFS('fx-rates'!$F:$F,'fx-rates'!$A:$A,{col}$2,"
            f"'fx-rates'!$C:$C,\"USD\",'fx-rates'!$D:$D,\"EUR\"),{col}{INC_EUR})"
        )
        ws[f"{col}{TOTAL_EUR}"] = f"={col}{LTD_EUR}+{col}{INC_EUR}"

    # Columna TOTAL (P)
    for row in (LTD_MK, LTD_RY, LTD_ST, LTD_AD, LTD_EUR, LTD_GBP,
                INC_MK, INC_RY, INC_ST, INC_AD, INC_EUR, INC_USD, TOTAL_EUR):
        ws[f"P{row}"] = f"=SUM(D{row}:O{row})"

    # ── Estilos ───────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = ROW_HEIGHT
    ws.row_dimensions[2].hidden = True
    for row_idx in range(1, TOTAL_EUR + 1):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    # Cabecera fila 3
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"].font = TITLE_FONT
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    # Secciones LTD / INC
    for hdr_row in (LTD_HDR, INC_HDR):
        for col in range(1, 18):
            ws.cell(row=hdr_row, column=col).fill = SECTION_FILL
        ws.cell(row=hdr_row, column=1).font = BOLD

    # Filas de total parcial
    for total_row in (LTD_EUR, LTD_GBP, INC_EUR, INC_USD):
        for col in range(1, 18):
            ws.cell(row=total_row, column=col).fill = SUBSECTION_FILL
        ws.cell(row=total_row, column=1).font = BOLD

    # Fila TOTAL SHARED SERVICES EUR — fondo KPI y fuente en negrita
    for col in range(1, 18):
        ws.cell(row=TOTAL_EUR, column=col).fill = KPI_FILL
    ws.cell(row=TOTAL_EUR, column=1).font = Font(bold=True, size=10)

    # Formato numérico (EUR / GBP / USD → MONEY_FORMAT)
    for row in list(range(4, INC_USD + 1)) + [TOTAL_EUR]:
        for col in range(4, 18):
            ws.cell(row=row, column=col).number_format = MONEY_FORMAT

    # Anchos de columna
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["P"].width = 14
    for idx in range(4, 16):
        ws.column_dimensions[get_column_letter(idx)].width = 14

    ws.freeze_panes = "D4"
    ws.sheet_view.showGridLines = False


def _fill_month_formulas(
    ws,
    *,
    pos: dict[str, int],
    shopify_rows: list[int],
    marketplace_rows: list[int],
    service_rows: list[int],
    manufacturing_rows: list[int],
    logistics_rows: list[int],
    payment_fee_rows: list[int],
    marketing_rows: list[int],
    marketing_meta_detail_rows: list[int],
    marketing_google_detail_rows: list[int],
    staff_rows: list[int],
    administration_rows: list[int],
    administration_detail_rows: dict[int, list[int]],
    technology_rows: list[int],
) -> None:
    for col in [get_column_letter(i) for i in range(4, 16)]:
        for row in shopify_rows:
            ws[f"{col}{row}"] = f'=SUMIFS(\'i-shopify-sl\'!$I:$I,\'i-shopify-sl\'!$A:$A,{col}$1,\'i-shopify-sl\'!$C:$C,$C{row})'
        ws[f"{col}{pos['shopify_header']}"] = f"=SUM({col}{shopify_rows[0]}:{col}{shopify_rows[-1]})"
        for row, sheet in (
            *[(r, "i-marketplaces-sl") for r in marketplace_rows],
            (pos["rappels_detail"], "i-rappels-sl"),
            (pos["supplies_detail"], "i-supplies-sl"),
            *[(r, "i-services-sl") for r in service_rows],
        ):
            ws[f"{col}{row}"] = f'=SUMIFS(\'{sheet}\'!$I:$I,\'{sheet}\'!$A:$A,{col}$1,\'{sheet}\'!$C:$C,$C{row})'
        ws[f"{col}{pos['marketplaces_header']}"] = f"=SUM({col}{marketplace_rows[0]}:{col}{marketplace_rows[-1]})"
        ws[f"{col}{pos['rappels_header']}"] = f"={col}{pos['rappels_detail']}"
        ws[f"{col}{pos['supplies_header']}"] = f"={col}{pos['supplies_detail']}"
        ws[f"{col}{pos['services_header']}"] = f"=SUM({col}{service_rows[0]}:{col}{service_rows[-1]})"
        ws[f"{col}{pos['product_sales']}"] = f"={col}{pos['shopify_header']}+{col}{pos['marketplaces_header']}+{col}{pos['rappels_header']}+{col}{pos['supplies_header']}"
        ws[f"{col}{pos['turnover']}"] = f"={col}{pos['product_sales']}+{col}{pos['services_header']}+{col}{pos['otros_ingresos']}"
        for row in manufacturing_rows + logistics_rows + [pos["royalties_detail"]] + payment_fee_rows + marketing_rows + marketing_meta_detail_rows + marketing_google_detail_rows + staff_rows + administration_rows + [detail_row for rows in administration_detail_rows.values() for detail_row in rows] + technology_rows:
            if ws[f"C{row}"].value:
                if row in payment_fee_rows:
                    ws[f"{col}{row}"] = f'=SUMIFS(\'g-payment-fees-sl\'!$H:$H,\'g-payment-fees-sl\'!$A:$A,{col}$1,\'g-payment-fees-sl\'!$C:$C,$C{row})'
                elif row in marketing_meta_detail_rows:
                    ws[f"{col}{row}"] = f'=SUMIFS(\'g-expenses-sl\'!$K:$K,\'g-expenses-sl\'!$A:$A,{col}$1,\'g-expenses-sl\'!$D:$D,"marketing",\'g-expenses-sl\'!$E:$E,"METAADS",\'g-expenses-sl\'!$F:$F,$C{row})'
                elif row in marketing_google_detail_rows:
                    ws[f"{col}{row}"] = f'=SUMIFS(\'g-expenses-sl\'!$K:$K,\'g-expenses-sl\'!$A:$A,{col}$1,\'g-expenses-sl\'!$D:$D,"marketing",\'g-expenses-sl\'!$E:$E,"GOOGLEADS",\'g-expenses-sl\'!$F:$F,$C{row})'
                elif row in [detail_row for rows in administration_detail_rows.values() for detail_row in rows]:
                    supplier_row = next(parent_row for parent_row, rows in administration_detail_rows.items() if row in rows)
                    ws[f"{col}{row}"] = f'=SUMIFS(\'g-expenses-sl\'!$K:$K,\'g-expenses-sl\'!$A:$A,{col}$1,\'g-expenses-sl\'!$D:$D,"administration",\'g-expenses-sl\'!$E:$E,$C{supplier_row},\'g-expenses-sl\'!$F:$F,$C{row})'
                else:
                    subcategory = (
                        "manufacturing" if row in manufacturing_rows else
                        "logistics" if row in logistics_rows else
                        "royalties" if row == pos["royalties_detail"] else
                        "marketing" if row in marketing_rows else
                        "staff" if row in staff_rows else
                        "administration" if row in administration_rows else
                        "technology"
                    )
                    ws[f"{col}{row}"] = f'=SUMIFS(\'g-expenses-sl\'!$K:$K,\'g-expenses-sl\'!$A:$A,{col}$1,\'g-expenses-sl\'!$D:$D,"{subcategory}",\'g-expenses-sl\'!$E:$E,$C{row})'
        ws[f"{col}{pos['manufacturing_header']}"] = f"=SUM({col}{manufacturing_rows[0]}:{col}{manufacturing_rows[-1]})"
        ws[f"{col}{pos['manufacturing_pct']}"] = f'=IFERROR({col}{pos["manufacturing_header"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['logistics_header']}"] = f"=SUM({col}{logistics_rows[0]}:{col}{logistics_rows[-1]})"
        ws[f"{col}{pos['logistics_pct']}"] = f'=IFERROR({col}{pos["logistics_header"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['royalties_header']}"] = f"={col}{pos['royalties_detail']}"
        ws[f"{col}{pos['royalties_pct']}"] = f'=IFERROR({col}{pos["royalties_header"]}/{col}{pos["product_sales"]},0)'
        for _scope_key, _scope_val in [("royalties_eu", "eu"), ("royalties_uk", "uk"), ("royalties_us", "us")]:
            ws[f"{col}{pos[_scope_key]}"] = (
                f"=SUMIFS('i-royalties-scope-sl'!$C:$C,'i-royalties-scope-sl'!$A:$A,{col}$1,"
                f"'i-royalties-scope-sl'!$B:$B,\"{_scope_val}\")"
            )
        ws[f"{col}{pos['payment_fees_header']}"] = f"=SUM({col}{payment_fee_rows[0]}:{col}{payment_fee_rows[-1]})"
        ws[f"{col}{pos['payment_fees_pct']}"] = f'=IFERROR({col}{pos["payment_fees_header"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['cogs']}"] = f"={col}{pos['manufacturing_header']}+{col}{pos['logistics_header']}+{col}{pos['royalties_header']}+{col}{pos['payment_fees_header']}"
        ws[f"{col}{pos['gross_margin']}"] = f"={col}{pos['product_sales']}-{col}{pos['manufacturing_header']}"
        ws[f"{col}{pos['gross_margin_pct']}"] = f'=IFERROR({col}{pos["gross_margin"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['contributive_margin']}"] = (
            f"={col}{pos['product_sales']}-{col}{pos['manufacturing_header']}-{col}{pos['logistics_header']}"
            f"-{col}{pos['royalties_eu']}-{col}{pos['payment_fees_header']}"
        )
        ws[f"{col}{pos['contributive_margin_pct']}"] = f'=IFERROR({col}{pos["contributive_margin"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['marketing_meta']}"] = f"=SUM({col}{marketing_meta_detail_rows[0]}:{col}{marketing_meta_detail_rows[-1]})"
        ws[f"{col}{pos['marketing_google']}"] = f"=SUM({col}{marketing_google_detail_rows[0]}:{col}{marketing_google_detail_rows[-1]})"
        ws[f"{col}{pos['marketing_header']}"] = f"={col}{pos['marketing_meta']}+{col}{pos['marketing_google']}"
        ws[f"{col}{pos['marketing_pct']}"] = f'=IFERROR(SUM({col}{shopify_rows[0]}:{col}{shopify_rows[-2]})/({col}{marketing_meta_detail_rows[0]}+{col}{marketing_google_detail_rows[0]}),0)'
        ws[f"{col}{pos['staff_header']}"] = f"=SUM({col}{staff_rows[0]}:{col}{staff_rows[-1]})"
        ws[f"{col}{pos['administration_header']}"] = "=" + "+".join(f"{col}{row}" for row in administration_rows)
        ws[f"{col}{pos['technology_header']}"] = f"=SUM({col}{technology_rows[0]}:{col}{technology_rows[-1]})"
        ws[f"{col}{pos['otros_gastos']}"] = f'=SUMIFS(\'g-expenses-sl\'!$K:$K,\'g-expenses-sl\'!$A:$A,{col}$1,\'g-expenses-sl\'!$D:$D,"otros_gastos")'
        ws[f"{col}{pos['opex']}"] = f"={col}{pos['marketing_header']}+{col}{pos['staff_header']}+{col}{pos['administration_header']}+{col}{pos['technology_header']}+{col}{pos['otros_gastos']}"
        ws[f"{col}{pos['expenses']}"] = f"={col}{pos['cogs']}+{col}{pos['opex']}"
        ws[f"{col}{pos['profit']}"] = f"={col}{pos['turnover']}-{col}{pos['cogs']}-{col}{pos['opex']}-{col}{pos['diferencias_divisas']}"
        ws[f"{col}{pos['profit_pct']}"] = f'=IFERROR({col}{pos["profit"]}/{col}{pos["product_sales"]},0)'
    for row in range(4, pos["profit_pct"] + 1):
        ws[f"P{row}"] = f"=SUM(D{row}:O{row})"
    # Percentage rows: use ratio on totals, not sum of monthly percentages
    ws[f"P{pos['manufacturing_pct']}"]       = f'=IFERROR(P{pos["manufacturing_header"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['logistics_pct']}"]           = f'=IFERROR(P{pos["logistics_header"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['royalties_pct']}"]           = f'=IFERROR(P{pos["royalties_header"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['payment_fees_pct']}"]        = f'=IFERROR(P{pos["payment_fees_header"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['gross_margin_pct']}"]        = f'=IFERROR(P{pos["gross_margin"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['contributive_margin_pct']}"] = f'=IFERROR(P{pos["contributive_margin"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['marketing_pct']}"]           = f'=IFERROR(SUM(P{shopify_rows[0]}:P{shopify_rows[-2]})/(P{marketing_meta_detail_rows[0]}+P{marketing_google_detail_rows[0]}),0)'
    ws[f"P{pos['profit_pct']}"]              = f'=IFERROR(P{pos["profit"]}/P{pos["product_sales"]},0)'


def _count_sheet_sl(wb: Workbook, bundle: PygSlDataBundle) -> None:
    """Mirror P&G-SL structure but show invoice counts instead of amounts."""
    from collections import defaultdict
    from openpyxl.formatting.rule import FormulaRule

    # ── Compute counts from bundle ──────────────────────────────────────────
    exp_sc: dict[tuple[str, str], dict[str, int]] = defaultdict(lambda: defaultdict(int))
    exp_sc_det: dict[tuple[str, str, str], dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in bundle.expense_rows:
        exp_sc[(r.subcategory, r.supplier_code)][r.yyyymm] += 1
        exp_sc_det[(r.subcategory, r.supplier_code, r.detail)][r.yyyymm] += 1

    svc: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in bundle.service_rows:
        svc[r.line_item][r.yyyymm] += 1

    mkt: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in bundle.marketplace_rows:
        mkt[r.line_item][r.yyyymm] += 1

    rap: dict[str, int] = defaultdict(int)
    for r in bundle.rappel_rows:
        rap[r.yyyymm] += 1

    sup: dict[str, int] = defaultdict(int)
    for r in bundle.supplies_rows:
        sup[r.yyyymm] += 1

    fee: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in bundle.payment_fee_rows:
        fee[r.supplier_code][r.yyyymm] += 1

    # ── Determine last populated month ──────────────────────────────────────
    all_yyyymm: set[str] = set()
    for d in list(exp_sc.values()) + list(svc.values()) + list(mkt.values()) + list(fee.values()):
        all_yyyymm.update(k for k, v in d.items() if v > 0)
    all_yyyymm.update(k for k, v in rap.items() if v > 0)
    all_yyyymm.update(k for k, v in sup.items() if v > 0)
    all_yyyymm.update(bundle.otros_ingresos_by_period.keys())
    all_yyyymm.update(bundle.diferencias_divisas_by_period.keys())
    valid_months = [m for m in all_yyyymm if m.startswith(str(bundle.year))]
    last_yyyymm = max(valid_months) if valid_months else None
    last_month_num = int(last_yyyymm[4:6]) if last_yyyymm else None
    prev_month_num = (last_month_num - 1) if last_month_num and last_month_num > 1 else None
    last_col_idx = (3 + last_month_num) if last_month_num else None
    prev_col_idx = (3 + prev_month_num) if prev_month_num else None

    # ── Create sheet ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Nº Facturas-SL")
    months = month_keys(bundle.year)
    for idx, yyyymm in enumerate(months, start=4):
        ws.cell(row=1, column=idx, value=yyyymm)
        ws.cell(row=2, column=idx, value=MONTH_NAMES_ES[idx - 4])
    ws["P1"] = "TOTAL"
    ws["P2"] = "Total"
    ws["A2"] = f"Nº Facturas {COMPANY_CODE} {bundle.year}"
    ws.merge_cells("A2:C2")
    for r in (1, 2):
        for cell in ws[r]:
            cell.fill = HEADER_FILL
            cell.font = WHITE_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = TITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 1
    ws.row_dimensions[2].height = ROW_HEIGHT

    # ── Build row layout (mirrors _main_sheet) ───────────────────────────────
    groups = _provider_groups(bundle.provider_catalog_rows)
    pos: dict[str, int] = {}
    leaf_rows: list[int] = []  # rows that get red-fill comparison

    def lbl(col: str, r: int, text: str, bold: bool = False) -> None:
        ws[f"{col}{r}"] = text
        if bold:
            ws[f"{col}{r}"].font = BOLD

    def write_counts(r: int, count_fn) -> None:
        """Write count values for all 12 months + total."""
        total = 0
        for idx, yyyymm in enumerate(months, start=4):
            v = count_fn(yyyymm)
            ws.cell(row=r, column=idx, value=v)
            total += v
        ws[f"P{r}"] = total

    def write_sum_rows(r: int, child_rows: list[int]) -> None:
        """Write SUM of child rows for aggregate rows."""
        for idx in range(4, 16):
            ws.cell(row=r, column=idx, value=sum(ws.cell(row=cr, column=idx).value or 0 for cr in child_rows))
        ws[f"P{r}"] = sum(ws[f"P{cr}"].value or 0 for cr in child_rows)

    row = 4
    pos["turnover"] = row; lbl("A", row, "Turnover", bold=True); row += 1
    pos["product_sales"] = row; lbl("B", row, "Product sales", bold=True); row += 1
    pos["shopify_header"] = row; lbl("C", row, "Shopify", bold=True); row += 1
    shopify_rows = list(range(row, row + len(bundle.shopify_markets)))
    for idx, market in enumerate(bundle.shopify_markets):
        lbl("C", row + idx, market)
    row += len(bundle.shopify_markets)
    pos["marketplaces_header"] = row; lbl("C", row, "Marketplaces", bold=True); row += 1
    marketplace_codes = ["HANNUN", "TOASTY", "CHOOSE"]
    marketplace_rows_list = list(range(row, row + len(marketplace_codes)))
    for idx, code in enumerate(marketplace_codes):
        lbl("C", row + idx, code)
    row += len(marketplace_codes)
    pos["rappels_header"] = row; lbl("C", row, "Rappels", bold=True); row += 1
    pos["rappels_detail"] = row; lbl("C", row, "LIVITUM"); row += 1
    pos["supplies_header"] = row; lbl("C", row, "Supplies", bold=True); row += 1
    pos["supplies_detail"] = row; lbl("C", row, "REVER"); row += 2
    pos["services_header"] = row; lbl("B", row, "Services", bold=True); row += 1
    service_codes = ["HANNUN", "QHANDS", "Ltd", "Inc"]
    service_rows_list = list(range(row, row + len(service_codes)))
    for idx, code in enumerate(service_codes):
        lbl("C", row + idx, code)
    row += len(service_codes)
    pos["otros_ingresos"] = row; lbl("B", row, "Uncategorized income", bold=True); row += 2

    pos["expenses"] = row; lbl("A", row, "Expenses", bold=True); row += 1
    pos["cogs"] = row; lbl("B", row, "COGS", bold=True); row += 1
    pos["manufacturing_header"] = row; lbl("C", row, "Manufacturing", bold=True); row += 1
    manufacturing_rows_list = list(range(row, row + len(groups["manufacturing"])))
    for idx, supplier in enumerate(groups["manufacturing"]):
        lbl("C", row + idx, supplier)
    row += len(manufacturing_rows_list)

    pos["logistics_header"] = row; lbl("C", row, "Logistics", bold=True); row += 1
    logistics_rows_list = list(range(row, row + len(groups["logistics"])))
    for idx, supplier in enumerate(groups["logistics"]):
        lbl("C", row + idx, supplier)
    row += len(logistics_rows_list)

    pos["royalties_header"] = row; lbl("C", row, "Royalties", bold=True); row += 1
    pos["royalties_detail"] = row; lbl("C", row, "ROYALTIES"); row += 1
    pos["royalties_eu"] = row; lbl("C", row, "eu"); row += 1
    pos["royalties_uk"] = row; lbl("C", row, "uk"); row += 1
    pos["royalties_us"] = row; lbl("C", row, "us"); row += 1

    pos["payment_fees_header"] = row; lbl("C", row, "Payment fees", bold=True); row += 1
    payment_fee_codes = ["SHOPIFY", "PAYPAL"]
    payment_fee_rows_list = list(range(row, row + len(payment_fee_codes)))
    for idx, code in enumerate(payment_fee_codes):
        lbl("C", row + idx, code)
    row += len(payment_fee_rows_list)
    row += 2  # blank + gross margin etc skipped

    pos["opex"] = row; lbl("B", row, "Opex", bold=True); row += 1
    pos["marketing_header"] = row; lbl("C", row, "Marketing", bold=True); row += 1
    pos["marketing_meta"] = row; lbl("C", row, "METAADS", bold=True); row += 1
    marketing_meta_detail_rows = list(range(row, row + len(DEFAULT_MARKETING_REGIONS)))
    for idx, code in enumerate(DEFAULT_MARKETING_REGIONS):
        lbl("C", row + idx, code)
    row += len(marketing_meta_detail_rows)
    pos["marketing_google"] = row; lbl("C", row, "GOOGLEADS", bold=True); row += 1
    marketing_google_detail_rows = list(range(row, row + len(DEFAULT_MARKETING_REGIONS)))
    for idx, code in enumerate(DEFAULT_MARKETING_REGIONS):
        lbl("C", row + idx, code)
    row += len(marketing_google_detail_rows)

    pos["staff_header"] = row; lbl("C", row, "Staff", bold=True); row += 1
    staff_rows_list = [row, row + 1]
    lbl("C", row, "PAYROLL"); lbl("C", row + 1, "DOSCONSULTING"); row += 2

    pos["administration_header"] = row; lbl("C", row, "Administration", bold=True); row += 1
    administration_rows_list: list[int] = []
    administration_detail_rows: dict[int, list[int]] = {}
    for supplier in groups["administration"]:
        lbl("C", row, supplier)
        administration_rows_list.append(row)
        row += 1
        detail_rows_for_supplier: list[int] = []
        for detail in ADMINISTRATION_DETAIL_LINES.get(supplier, []):
            lbl("C", row, detail)
            detail_rows_for_supplier.append(row)
            row += 1
        if detail_rows_for_supplier:
            administration_detail_rows[administration_rows_list[-1]] = detail_rows_for_supplier

    pos["technology_header"] = row; lbl("C", row, "Technology", bold=True); row += 1
    technology_rows_list = list(range(row, row + len(groups["technology"])))
    for idx, supplier in enumerate(groups["technology"]):
        lbl("C", row + idx, supplier)
    row += len(technology_rows_list)
    pos["otros_gastos"] = row; lbl("C", row, "Uncategorized Expenses", bold=True); row += 1
    pos["diferencias_divisas"] = row; lbl("B", row, "Currency Adjustment", bold=True); row += 1
    max_data_row = row - 1

    # ── Write counts ─────────────────────────────────────────────────────────
    # Shopify markets (count of sales records, typically 1 per market/month)
    for r, market in zip(shopify_rows, bundle.shopify_markets):
        write_counts(r, lambda yyyymm, m=market: sum(1 for sr in bundle.shopify_rows if sr.yyyymm == yyyymm and sr.line_item == m))
        leaf_rows.append(r)

    # Marketplaces
    for r, code in zip(marketplace_rows_list, marketplace_codes):
        write_counts(r, lambda yyyymm, c=code: mkt[c][yyyymm])
        leaf_rows.append(r)

    # Rappels
    write_counts(pos["rappels_detail"], lambda yyyymm: rap[yyyymm])
    leaf_rows.append(pos["rappels_detail"])

    # Supplies
    write_counts(pos["supplies_detail"], lambda yyyymm: sup[yyyymm])
    leaf_rows.append(pos["supplies_detail"])

    # Services
    for r, code in zip(service_rows_list, service_codes):
        write_counts(r, lambda yyyymm, c=code: svc[c][yyyymm])
        leaf_rows.append(r)

    # Otros ingresos
    write_counts(pos["otros_ingresos"], lambda yyyymm: 1 if yyyymm in bundle.otros_ingresos_by_period else 0)
    leaf_rows.append(pos["otros_ingresos"])

    # Manufacturing
    for r, supplier in zip(manufacturing_rows_list, groups["manufacturing"]):
        write_counts(r, lambda yyyymm, s=supplier: exp_sc[("manufacturing", s)][yyyymm])
        leaf_rows.append(r)

    # Logistics
    for r, supplier in zip(logistics_rows_list, groups["logistics"]):
        write_counts(r, lambda yyyymm, s=supplier: exp_sc[("logistics", s)][yyyymm])
        leaf_rows.append(r)

    # Royalties
    write_counts(pos["royalties_detail"], lambda yyyymm: exp_sc[("royalties", "ROYALTIES")][yyyymm])
    leaf_rows.append(pos["royalties_detail"])
    for _r, _scope in [(pos["royalties_eu"], "eu"), (pos["royalties_uk"], "uk"), (pos["royalties_us"], "us")]:
        write_counts(_r, lambda yyyymm, s=_scope: 1 if bundle.royalties_by_scope.get(s, {}).get(yyyymm) else 0)
        leaf_rows.append(_r)

    # Payment fees
    for r, code in zip(payment_fee_rows_list, payment_fee_codes):
        write_counts(r, lambda yyyymm, c=code: fee[c][yyyymm])
        leaf_rows.append(r)

    # Marketing METAADS detail
    for r, region in zip(marketing_meta_detail_rows, DEFAULT_MARKETING_REGIONS):
        write_counts(r, lambda yyyymm, reg=region: exp_sc_det[("marketing", "METAADS", reg)][yyyymm])
        leaf_rows.append(r)

    # Marketing GOOGLEADS detail
    for r, region in zip(marketing_google_detail_rows, DEFAULT_MARKETING_REGIONS):
        write_counts(r, lambda yyyymm, reg=region: exp_sc_det[("marketing", "GOOGLEADS", reg)][yyyymm])
        leaf_rows.append(r)

    # Staff
    write_counts(staff_rows_list[0], lambda yyyymm: exp_sc[("staff", "PAYROLL")][yyyymm])
    leaf_rows.append(staff_rows_list[0])
    write_counts(staff_rows_list[1], lambda yyyymm: exp_sc[("staff", "DOSCONSULTING")][yyyymm])
    leaf_rows.append(staff_rows_list[1])

    # Administration
    for admin_row in administration_rows_list:
        supplier = ws[f"C{admin_row}"].value
        detail_rows_for_this = administration_detail_rows.get(admin_row, [])
        if detail_rows_for_this:
            for detail_row in detail_rows_for_this:
                detail_val = ws[f"C{detail_row}"].value
                write_counts(detail_row, lambda yyyymm, s=supplier, d=detail_val: exp_sc_det[("administration", s, d)][yyyymm])
                leaf_rows.append(detail_row)
            write_sum_rows(admin_row, detail_rows_for_this)
        else:
            write_counts(admin_row, lambda yyyymm, s=supplier: exp_sc[("administration", s)][yyyymm])
            leaf_rows.append(admin_row)

    # Technology
    for r, supplier in zip(technology_rows_list, groups["technology"]):
        write_counts(r, lambda yyyymm, s=supplier: exp_sc[("technology", s)][yyyymm])
        leaf_rows.append(r)

    # Otros gastos (count rows with supplier_code='OTROSGASTOS')
    write_counts(pos["otros_gastos"], lambda yyyymm: min(1, exp_sc[("otros_gastos", "OTROSGASTOS")][yyyymm]))
    leaf_rows.append(pos["otros_gastos"])

    # Diferencias divisas
    write_counts(pos["diferencias_divisas"], lambda yyyymm: 1 if yyyymm in bundle.diferencias_divisas_by_period else 0)
    leaf_rows.append(pos["diferencias_divisas"])

    # ── Aggregate rows (SUM of children) ────────────────────────────────────
    write_sum_rows(pos["shopify_header"], shopify_rows)
    write_sum_rows(pos["marketplaces_header"], marketplace_rows_list)
    write_sum_rows(pos["rappels_header"], [pos["rappels_detail"]])
    write_sum_rows(pos["supplies_header"], [pos["supplies_detail"]])
    write_sum_rows(pos["services_header"], service_rows_list)
    write_sum_rows(pos["product_sales"], shopify_rows + marketplace_rows_list + [pos["rappels_detail"], pos["supplies_detail"]])
    write_sum_rows(pos["turnover"], [pos["product_sales"], pos["services_header"], pos["otros_ingresos"]])
    write_sum_rows(pos["manufacturing_header"], manufacturing_rows_list)
    write_sum_rows(pos["logistics_header"], logistics_rows_list)
    write_sum_rows(pos["royalties_header"], [pos["royalties_detail"]])
    write_sum_rows(pos["payment_fees_header"], payment_fee_rows_list)
    write_sum_rows(pos["cogs"], manufacturing_rows_list + logistics_rows_list + [pos["royalties_detail"]] + payment_fee_rows_list)
    write_sum_rows(pos["marketing_meta"], marketing_meta_detail_rows)
    write_sum_rows(pos["marketing_google"], marketing_google_detail_rows)
    write_sum_rows(pos["marketing_header"], [pos["marketing_meta"], pos["marketing_google"]])
    write_sum_rows(pos["staff_header"], staff_rows_list)
    all_admin_leaf = [d for dr in administration_detail_rows.values() for d in dr] or administration_rows_list
    write_sum_rows(pos["administration_header"], all_admin_leaf)
    write_sum_rows(pos["technology_header"], technology_rows_list)
    write_sum_rows(pos["opex"], [pos["marketing_header"], pos["staff_header"], pos["administration_header"], pos["technology_header"], pos["otros_gastos"]])
    write_sum_rows(pos["expenses"], [pos["cogs"], pos["opex"]])

    # ── Styling ──────────────────────────────────────────────────────────────
    major_rows = {pos["turnover"], pos["expenses"]}
    subtotal_rows = {pos["product_sales"], pos["services_header"], pos["cogs"], pos["opex"]}
    section_rows = {
        pos["shopify_header"], pos["marketplaces_header"], pos["rappels_header"], pos["supplies_header"],
        pos["manufacturing_header"], pos["logistics_header"], pos["royalties_header"], pos["payment_fees_header"],
        pos["marketing_header"], pos["marketing_meta"], pos["marketing_google"], pos["staff_header"],
        pos["administration_header"], pos["technology_header"], pos["otros_gastos"], pos["otros_ingresos"],
        pos["diferencias_divisas"],
    }
    for r in range(4, max_data_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT
        if r in major_rows:
            for c in range(1, 18):
                ws.cell(row=r, column=c).fill = SECTION_FILL
        elif r in section_rows:
            for c in range(1, 18):
                ws.cell(row=r, column=c).fill = SUBSECTION_FILL
        elif r in subtotal_rows:
            for c in range(1, 18):
                ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
    for c, w in (("A", 18), ("B", 20), ("C", 24)):
        ws.column_dimensions[c].width = w
    ws.column_dimensions["P"].width = 12
    for i in range(4, 17):
        ws.column_dimensions[get_column_letter(i)].width = 7
    ws.freeze_panes = "D4"
    ws.sheet_view.showGridLines = False

    # ── Red fill: last month, rows with 0 or < half of previous month ────────
    if last_col_idx and prev_col_idx:
        red_fill = PatternFill("solid", fgColor="FFAAAA")
        last_col = get_column_letter(last_col_idx)
        prev_col = get_column_letter(prev_col_idx)
        for r in leaf_rows:
            last_val = ws.cell(row=r, column=last_col_idx).value or 0
            prev_val = ws.cell(row=r, column=prev_col_idx).value or 0
            if prev_val > 0 and (last_val == 0 or last_val < prev_val / 2):
                ws.cell(row=r, column=last_col_idx).fill = red_fill


def _apply_layout(
    ws,
    *,
    pos: dict[str, int],
    shopify_rows: list[int],
    marketplace_rows: list[int],
    service_rows: list[int],
    manufacturing_rows: list[int],
    logistics_rows: list[int],
    payment_fee_rows: list[int],
    marketing_rows: list[int],
    marketing_meta_detail_rows: list[int],
    marketing_google_detail_rows: list[int],
    staff_rows: list[int],
    administration_rows: list[int],
    administration_detail_rows: dict[int, list[int]],
    technology_rows: list[int],
) -> None:
    flattened_admin_detail_rows = [detail_row for rows in administration_detail_rows.values() for detail_row in rows]
    percent_rows = {
        pos["manufacturing_pct"], pos["logistics_pct"], pos["royalties_pct"], pos["payment_fees_pct"],
        pos["gross_margin_pct"], pos["contributive_margin_pct"], pos["profit_pct"],
    }
    ratio_rows = {pos["marketing_pct"]}
    major_rows = {pos["turnover"], pos["expenses"], pos["gross_margin"], pos["contributive_margin"], pos["profit"]}
    subtotal_rows = {
        pos["product_sales"], pos["services_header"], pos["cogs"], pos["opex"],
    }
    section_rows = {
        pos["shopify_header"], pos["marketplaces_header"], pos["rappels_header"], pos["supplies_header"],
        pos["manufacturing_header"], pos["logistics_header"], pos["royalties_header"], pos["payment_fees_header"],
        pos["marketing_header"], pos["marketing_meta"], pos["marketing_google"], pos["staff_header"], pos["administration_header"], pos["technology_header"],
        pos["otros_gastos"],
        pos["otros_ingresos"],
        pos["diferencias_divisas"],
    }
    detail_rows = set(
        shopify_rows
        + marketplace_rows
        + [pos["rappels_detail"], pos["supplies_detail"]]
        + service_rows
        + manufacturing_rows
        + logistics_rows
        + [pos["royalties_detail"], pos["royalties_eu"], pos["royalties_uk"], pos["royalties_us"]]
        + payment_fee_rows
        + marketing_meta_detail_rows
        + marketing_google_detail_rows
        + staff_rows
        + administration_rows
        + flattened_admin_detail_rows
        + technology_rows
    )
    level_1_rows = subtotal_rows
    level_2_rows = section_rows | percent_rows | ratio_rows
    level_3_rows = detail_rows
    for row in range(4, pos["profit_pct"] + 1):
        if not ws.cell(row=row, column=1).value:
            ws.cell(row=row, column=1, value=ws.cell(row=row, column=2).value or ws.cell(row=row, column=3).value)
        raw_label = str(ws.cell(row=row, column=1).value or "")
        if row in major_rows:
            ws.cell(row=row, column=1, value=raw_label.strip())
        elif row in level_1_rows:
            ws.cell(row=row, column=1, value=_display_label(raw_label, level=1))
        elif row in level_2_rows:
            ws.cell(row=row, column=1, value=_display_label(raw_label, level=2))
        elif row in level_3_rows:
            ws.cell(row=row, column=1, value=_display_label(raw_label, level=3))
    for row in range(3, pos["profit_pct"] + 1):
        ws.row_dimensions[row].height = ROW_HEIGHT
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            if col <= 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(size=9, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color.type == "rgb" and cell.font.color.rgb or None)
        label_cell = ws.cell(row=row, column=1)
        if row in major_rows:
            label_cell.font = Font(bold=True, size=10)
            for col in range(1, 18):
                ws.cell(row=row, column=col).fill = SECTION_FILL if row in {pos["turnover"], pos["expenses"]} else KPI_FILL
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=0)
            _apply_border_row(ws, row, MEDIUM_TOP_BORDER)
        elif row in subtotal_rows:
            label_cell.font = Font(bold=True, size=9)
            for col in range(1, 18):
                ws.cell(row=row, column=col).fill = SUBTOTAL_FILL
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            _apply_border_row(ws, row, THIN_TOP_BORDER)
        elif row in section_rows:
            label_cell.font = Font(bold=True, size=9)
            for col in range(1, 18):
                ws.cell(row=row, column=col).fill = SUBSECTION_FILL
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            _apply_border_row(ws, row, THIN_TOP_BORDER)
        elif row in percent_rows:
            label_cell.font = Font(italic=True, color="666666", size=9)
            for col in range(1, 18):
                ws.cell(row=row, column=col).fill = PERCENT_ROW_FILL
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        elif row in ratio_rows:
            label_cell.font = Font(bold=True, color="666666", size=9)
            for col in range(1, 18):
                ws.cell(row=row, column=col).fill = PERCENT_ROW_FILL
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        elif row in detail_rows and ws.cell(row=row, column=1).value:
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=3)
            ws.cell(row=row, column=1).font = Font(size=9)
        if row in {pos["gross_margin_pct"], pos["contributive_margin_pct"], pos["profit_pct"]}:
            _apply_border_row(ws, row, THIN_TOP_BORDER)
    for row in range(4, pos["profit_pct"] + 1):
        ws[f"P{row}"].fill = TOTAL_FILL
        if row in major_rows | subtotal_rows:
            ws[f"P{row}"].font = BOLD

    _collapse_group(ws, shopify_rows, level=1)
    _collapse_group(ws, marketplace_rows, level=1)
    _collapse_group(ws, [pos["rappels_detail"]], level=1)
    _collapse_group(ws, [pos["supplies_detail"]], level=1)
    _collapse_group(ws, service_rows, level=1)
    _collapse_group(ws, manufacturing_rows, level=1)
    _collapse_group(ws, logistics_rows, level=1)
    _collapse_group(ws, [pos["royalties_detail"], pos["royalties_eu"], pos["royalties_uk"], pos["royalties_us"]], level=1)
    _collapse_group(ws, payment_fee_rows, level=1)
    _collapse_group(ws, [*marketing_rows, *marketing_meta_detail_rows, *marketing_google_detail_rows, pos["marketing_pct"]], level=1)
    _collapse_group(ws, staff_rows, level=1)
    for rows in administration_detail_rows.values():
        _collapse_group(ws, rows, level=1)
    _collapse_group(ws, administration_rows, level=1)
    _collapse_group(ws, technology_rows, level=1)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["C"].hidden = True
    for row in (
        pos["turnover"], pos["expenses"], pos["gross_margin"], pos["contributive_margin"], pos["profit"],
        pos["product_sales"], pos["services_header"], pos["cogs"], pos["opex"],
    ):
        ws.row_dimensions[row].height = ROW_HEIGHT


def _apply_border_row(ws, row: int, border: Border) -> None:
    for col in range(1, 17):
        ws.cell(row=row, column=col).border = border


def _collapse_group(ws, rows: list[int], *, level: int) -> None:
    if not rows:
        return
    for row in rows:
        ws.row_dimensions[row].outlineLevel = level
        ws.row_dimensions[row].hidden = True
    ws.row_dimensions[rows[-1]].collapsed = True


def _provider_groups(rows: tuple[ProviderCatalogRow, ...]) -> dict[str, list[str]]:
    groups = {"manufacturing": [], "logistics": [], "administration": [], "technology": []}
    for row in rows:
        if row.destination_path == "expenses/cogs/manufacturing":
            groups["manufacturing"].append(row.supplier_code)
        elif row.destination_path == "expenses/cogs/logistics":
            groups["logistics"].append(row.supplier_code)
        elif row.destination_path == "expenses/opex/administration":
            groups["administration"].append(row.supplier_code)
        elif row.destination_path == "expenses/opex/technology":
            groups["technology"].append(row.supplier_code)
    return groups


def _filter_periodified_documents(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    periodified_roots = {
        invoice_number[: invoice_number.index("_PERIODIFICADA_")]
        for row in rows
        if (invoice_number := str(row["invoice_number"] or "").strip())
        and "_PERIODIFICADA_" in invoice_number
    }
    if not periodified_roots:
        return rows

    filtered: list[dict[str, Any]] = []
    for row in rows:
        invoice_number = str(row["invoice_number"] or "").strip()
        parser_name = str(row.get("parser_name") or "").strip().lower()
        if invoice_number in periodified_roots and parser_name != "manual_periodificada":
            continue
        filtered.append(row)
    return filtered


def _add_navigation_links(
    ws,
    *,
    pos: dict[str, int],
    shopify_rows: list[int],
    marketplace_rows: list[int],
    service_rows: list[int],
    manufacturing_rows: list[int],
    logistics_rows: list[int],
    payment_fee_rows: list[int],
    marketing_rows: list[int],
    marketing_meta_detail_rows: list[int],
    marketing_google_detail_rows: list[int],
    staff_rows: list[int],
    administration_rows: list[int],
    administration_detail_rows: dict[int, list[int]],
    technology_rows: list[int],
) -> None:
    sheet_links = {
        pos["shopify_header"]: "i-shopify-sl",
        pos["marketplaces_header"]: "i-marketplaces-sl",
        pos["rappels_header"]: "i-rappels-sl",
        pos["supplies_header"]: "i-supplies-sl",
        pos["services_header"]: "i-services-sl",
        pos["manufacturing_header"]: "g-expenses-sl",
        pos["logistics_header"]: "g-expenses-sl",
        pos["royalties_header"]: "g-expenses-sl",
        pos["payment_fees_header"]: "g-payment-fees-sl",
        pos["marketing_header"]: "g-expenses-sl",
        pos["staff_header"]: "g-expenses-sl",
        pos["administration_header"]: "g-expenses-sl",
        pos["technology_header"]: "g-expenses-sl",
        pos["otros_gastos"]: "g-expenses-sl",
        pos["rappels_detail"]: "i-rappels-sl",
        pos["supplies_detail"]: "i-supplies-sl",
        pos["royalties_detail"]: "g-expenses-sl",
    }
    for row in shopify_rows:
        sheet_links[row] = "i-shopify-sl"
    for row in marketplace_rows:
        sheet_links[row] = "i-marketplaces-sl"
    for row in service_rows:
        sheet_links[row] = "i-services-sl"
    for row in manufacturing_rows + logistics_rows + marketing_rows + marketing_meta_detail_rows + marketing_google_detail_rows + staff_rows + administration_rows + [detail_row for rows in administration_detail_rows.values() for detail_row in rows] + technology_rows:
        sheet_links[row] = "g-expenses-sl"
    for row in payment_fee_rows:
        sheet_links[row] = "g-payment-fees-sl"

    for row, sheet_name in sheet_links.items():
        cell = ws[f"Q{row}"]
        if not ws[f"A{row}"].value:
            continue
        cell.value = "->"
        cell._hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet_name}'!A1", display="->")
        cell.font = Font(color="666666", bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _normalize_shopify_market(value: Any) -> str:
    country_code = str(value or "").upper()
    return country_code if country_code in set(DEFAULT_SHOPIFY_MARKETS[:-1]) else "XX"


def _display_label(value: str, *, level: int) -> str:
    clean = value.strip()
    return f"{'\u00A0' * (4 * level)}{clean}" if clean else clean


def _dedupe_stage_rows(rows: list[StageRow]) -> list[StageRow]:
    aggregated: dict[tuple[str, str, str, str, str, str, str], Decimal] = {}
    sources: dict[tuple[str, str, str, str, str, str, str], str] = {}
    for row in rows:
        key = (row.yyyymm, row.entity, row.line_item, row.detail, row.currency, row.invoice_number, row.drive_url)
        aggregated[key] = aggregated.get(key, Decimal("0")) + row.amount_net
        sources[key] = row.source
    return [
        StageRow(yyyymm, entity, line_item, detail, amount_net, currency, sources[(yyyymm, entity, line_item, detail, currency, invoice_number, drive_url)], invoice_number, drive_url)
        for (yyyymm, entity, line_item, detail, currency, invoice_number, drive_url), amount_net in sorted(aggregated.items())
    ]


def _decimal(value: Any) -> Decimal:
    return value if isinstance(value, Decimal) else Decimal(str(value or "0"))


def _period_from_timestamp(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        parsed = value
    else:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(DISPLAY_TIMEZONE)
    return parsed.strftime("%Y%m")
