"""Build a per-SKU stock detail workbook for a given fabricante / month.

Reads from:
  supply.frame_consumption_valued  — per-SKU consumption detail
  supply.frame_stock_monthly        — opening / closing summary

Generates an xlsx suitable for archiving in the Drive expenses/cogs/stock folder.

WAC note
--------
``unit_wac_opening`` is the WAC at the start of the month.  If a purchase
arrived mid-month the WAC changed part-way through, so the per-day amounts
were computed with two different WAC values.  ``amount_effective`` always
reflects the correct weighted cost regardless of intra-month purchases.

The "Effective WAC" column shows ``amount_effective / quantity_effective``
(i.e. the realised average cost per unit consumed that month).  When there
was no intra-month purchase it equals ``unit_wac_opening``; when there was,
it shows the blended rate actually charged to the P&L.  The "Opening WAC"
column is kept for reference.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, UTC
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:  # pragma: no cover
    psycopg = None  # type: ignore[assignment]
    dict_row = None  # type: ignore[assignment]


MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
MONEY_FMT = '#,##0.00;[Red](#,##0.00);-'
INT_FMT   = '#,##0;[Red](#,##0);-'
ROW_H     = 14

HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="2E74B5")
SUBTOT_FILL  = PatternFill("solid", fgColor="D9EAF7")
SUMMARY_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_BOLD   = Font(color="FFFFFF", bold=True)
BOLD         = Font(bold=True)
THIN         = Border(top=Side(style="thin", color="BFBFBF"))
MEDIUM_TOP   = Border(top=Side(style="medium", color="595959"))
CENTER       = Alignment(horizontal="center", vertical="center")
RIGHT        = Alignment(horizontal="right", vertical="center")
LEFT_VA      = Alignment(horizontal="left", vertical="center")


@dataclass
class StockDetailBundle:
    fabricante: str
    mes_yyyymm: str
    currency: str
    generated_at: datetime
    skus: list[dict[str, Any]] = field(default_factory=list)
    monthly: dict[str, Any] = field(default_factory=dict)


def collect_stock_detail(
    *,
    fabricante: str,
    mes_yyyymm: str,
    database_url: str,
) -> StockDetailBundle:
    """Query DB and return a StockDetailBundle."""
    if psycopg is None:
        raise RuntimeError("psycopg is not installed.")

    year  = int(mes_yyyymm[:4])
    month = int(mes_yyyymm[4:])
    from datetime import date
    month_start = date(year, month, 1)

    with psycopg.connect(database_url, row_factory=dict_row) as conn:
        # Aggregate daily rows → per-SKU monthly totals
        # Join with override table and opening WAC from frame_sku_wac
        skus = conn.execute(
            """
            WITH daily AS (
                SELECT frame_color, frame_size,
                       SUM(quantity) AS quantity_system,
                       SUM(amount)   AS amount_system
                FROM supply.frame_consumption_valued
                WHERE fabricante = %s AND mes_yyyymm = %s
                GROUP BY frame_color, frame_size
            ),
            overrides AS (
                SELECT frame_color, frame_size, quantity_override, opening_wac
                FROM supply.frame_consumption_override
                WHERE fabricante = %s AND mes_yyyymm = %s
            ),
            opening_wac AS (
                SELECT DISTINCT ON (frame_color, frame_size)
                       frame_color, frame_size, wac
                FROM supply.frame_sku_wac
                WHERE fabricante = %s AND effective_from < %s
                ORDER BY frame_color, frame_size, effective_from DESC, id DESC
            )
            SELECT
                COALESCE(d.frame_color, ov.frame_color) AS frame_color,
                COALESCE(d.frame_size,  ov.frame_size)  AS frame_size,
                COALESCE(d.quantity_system, 0)          AS quantity_system,
                ov.quantity_override,
                COALESCE(ov.quantity_override, d.quantity_system, 0) AS quantity_effective,
                COALESCE(ow.wac, 0)                     AS unit_wac_opening,
                COALESCE(d.amount_system, 0)            AS amount_system,
                CASE
                    WHEN ov.quantity_override IS NOT NULL
                        THEN ov.quantity_override * COALESCE(ov.opening_wac, ow.wac, 0)
                    ELSE COALESCE(d.amount_system, 0)
                END                                      AS amount_effective
            FROM daily d
            FULL OUTER JOIN overrides ov
                ON ov.frame_color = d.frame_color AND ov.frame_size = d.frame_size
            LEFT JOIN opening_wac ow
                ON ow.frame_color = COALESCE(d.frame_color, ov.frame_color)
               AND ow.frame_size  = COALESCE(d.frame_size,  ov.frame_size)
            ORDER BY frame_color, frame_size
            """,
            (fabricante, mes_yyyymm, fabricante, mes_yyyymm, fabricante, month_start),
        ).fetchall()

        monthly = conn.execute(
            """
            SELECT currency, opening_units, opening_value,
                   purchased_units, purchased_value,
                   consumed_units, consumed_value,
                   closing_units, closing_value,
                   calculated_at
            FROM supply.frame_stock_monthly
            WHERE fabricante = %s AND mes_yyyymm = %s
            """,
            (fabricante, mes_yyyymm),
        ).fetchone()

    currency = (monthly or {}).get("currency", "USD")
    return StockDetailBundle(
        fabricante=fabricante,
        mes_yyyymm=mes_yyyymm,
        currency=str(currency),
        generated_at=datetime.now(UTC),
        skus=[dict(r) for r in skus],
        monthly=dict(monthly) if monthly else {},
    )


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def build_stock_detail_workbook(bundle: StockDetailBundle, output_path: Path) -> Path:
    """Write xlsx to output_path and return it."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _build_wb(bundle).save(str(output_path))
    return output_path


def build_stock_detail_bytes(bundle: StockDetailBundle) -> bytes:
    """Return xlsx as bytes (no disk I/O)."""
    buf = BytesIO()
    _build_wb(bundle).save(buf)
    return buf.getvalue()


def _effective_wac(sku: dict[str, Any]) -> Decimal | None:
    """Return amount_effective / quantity_effective, or None if qty == 0."""
    qty = sku.get("quantity_effective") or 0
    if qty == 0:
        return None
    amt = Decimal(str(sku["amount_effective"]))
    return (amt / Decimal(qty)).quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)


def _build_wb(bundle: StockDetailBundle) -> Workbook:
    wb = Workbook()
    ws = wb.active

    year  = int(bundle.mes_yyyymm[:4])
    month = int(bundle.mes_yyyymm[4:])
    title = f"Frame Stock — {bundle.fabricante} — {MONTH_NAMES[month - 1]} {year}"
    ws.title = bundle.mes_yyyymm

    # Columns: Color | Size | Sys Qty | Override | Eff Qty | Opening WAC | Eff WAC | Sys Amt | Eff Amt
    col_widths = [14, 8, 11, 10, 10, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- title ----
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = title
    c.font = Font(size=13, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = CENTER

    # ---- meta ----
    ws.row_dimensions[2].height = 13
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (
        f"Generated: {bundle.generated_at.strftime('%Y-%m-%d %H:%M UTC')}   |   "
        f"Currency: {bundle.currency}   |   "
        f"Effective WAC = Total amount / Units consumed (blended if mid-month purchase)"
    )
    c.font = Font(size=9, italic=True, color="595959")
    c.alignment = CENTER

    # ---- column headers ----
    cur = bundle.currency
    headers = [
        "Color", "Size",
        "System Qty", "Override", "Effective Qty",
        f"Opening WAC ({cur})", f"Effective WAC ({cur})",
        f"System Amount ({cur})", f"Effective Amount ({cur})",
    ]
    col_comments = [
        "Frame colour group (e.g. 1.Blanco, 2.Negro, 3.Roble…).\nDefined by the supplier; used as the primary grouping key.",
        "Frame size in cm, width × height (e.g. 50x70).\nAlways stored in the canonical format used by consumption data.",
        "Units consumed according to the source system (supply.consumo_marcos_diario).\nThis figure is never edited manually and is preserved even when an override is applied.",
        "Manual override quantity set by the operations team (e.g. after a physical stock count).\nBlank = no override; the system quantity is used as-is.\nCan be negative to represent a stock return or correction.",
        "The quantity that goes to the P&L.\nRules: if Override is set → Effective = Override; otherwise Effective = System Qty.\nThis is the figure used to value consumption and update closing stock.",
        f"Weighted Average Cost ({cur}) per unit at the START of this month,\nbefore any intra-month purchases.\nFormula: (units_on_hand × prev_WAC + qty_purchased × purchase_price) / (units_on_hand + qty_purchased).\nUsed as a reference; the P&L charge may differ if a purchase arrived mid-month.",
        f"Realised average cost ({cur}) per unit consumed this month.\nFormula: Effective Amount / Effective Qty.\nWhen no purchase occurred mid-month this equals Opening WAC.\nWhen a purchase changed the WAC part-way through, this shows the blended\nrate actually charged to the P&L (weighted by daily consumption).",
        f"Gross consumption amount ({cur}) based on System Qty × daily WAC.\nFormula: Σ(daily_system_qty × WAC_in_effect_that_day).\nFor reference only; not used in the P&L when an override is active.",
        f"Net consumption amount ({cur}) that flows to the P&L.\nFormula: Σ(daily_effective_qty × WAC_in_effect_that_day).\nWhen an override is set, the daily breakdown is replaced by:\n  Override Qty × Opening WAC.\nThis is the definitive cost-of-goods figure for this SKU and month.",
    ]
    row = 4
    ws.row_dimensions[row].height = ROW_H
    for col, (h, note) in enumerate(zip(headers, col_comments), 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = WHITE_BOLD
        c.fill = SECTION_FILL
        c.alignment = CENTER
        c.border = THIN
        c.comment = Comment(note, "Stock Detail")

    # ---- data rows ----
    row += 1
    color_groups: dict[str, list[dict]] = {}
    for sku in bundle.skus:
        color_groups.setdefault(sku["frame_color"], []).append(sku)

    for color, skus in color_groups.items():
        for sku in skus:
            ws.row_dimensions[row].height = ROW_H
            eff_wac = _effective_wac(sku)
            vals: list[Any] = [
                sku["frame_color"],
                sku["frame_size"],
                sku["quantity_system"],
                sku["quantity_override"],   # None renders as blank
                sku["quantity_effective"],
                Decimal(str(sku["unit_wac_opening"])),
                eff_wac,                    # None → blank cell
                Decimal(str(sku["amount_system"])),
                Decimal(str(sku["amount_effective"])),
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = LEFT_VA if col <= 2 else RIGHT
                c.border = THIN
                if col in (6, 7, 8, 9):
                    c.number_format = MONEY_FMT
                elif col in (3, 4, 5):
                    c.number_format = INT_FMT
            row += 1

        # subtotal per color
        ws.row_dimensions[row].height = ROW_H
        sum_sys     = sum(s["quantity_system"] or 0 for s in skus)
        sum_eff     = sum(s["quantity_effective"] or 0 for s in skus)
        sum_amt_sys = sum(Decimal(str(s["amount_system"])) for s in skus)
        sum_amt_eff = sum(Decimal(str(s["amount_effective"])) for s in skus)
        # Blended effective WAC for the color group
        blended_eff_wac: Decimal | None = None
        if sum_eff != 0:
            blended_eff_wac = (sum_amt_eff / Decimal(sum_eff)).quantize(
                Decimal("0.000001"), rounding=ROUND_HALF_UP
            )

        subtot_vals: list[Any] = [
            f"  Subtotal {color}", "",
            sum_sys, "", sum_eff,
            "", blended_eff_wac,
            sum_amt_sys, sum_amt_eff,
        ]
        for col, v in enumerate(subtot_vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = BOLD
            c.fill = SUBTOT_FILL
            c.border = MEDIUM_TOP
            c.alignment = LEFT_VA if col <= 2 else RIGHT
            if col in (7, 8, 9):
                c.number_format = MONEY_FMT
            elif col in (3, 5):
                c.number_format = INT_FMT
        row += 1

    # ---- grand total ----
    ws.row_dimensions[row].height = ROW_H + 2
    total_sys     = sum(s["quantity_system"] or 0 for s in bundle.skus)
    total_eff     = sum(s["quantity_effective"] or 0 for s in bundle.skus)
    total_amt_sys = sum(Decimal(str(s["amount_system"])) for s in bundle.skus)
    total_amt_eff = sum(Decimal(str(s["amount_effective"])) for s in bundle.skus)
    grand_eff_wac: Decimal | None = None
    if total_eff != 0:
        grand_eff_wac = (total_amt_eff / Decimal(total_eff)).quantize(
            Decimal("0.000001"), rounding=ROUND_HALF_UP
        )

    total_vals: list[Any] = [
        "TOTAL CONSUMPTION", "",
        total_sys, "", total_eff,
        "", grand_eff_wac,
        total_amt_sys, total_amt_eff,
    ]
    for col, v in enumerate(total_vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.border = MEDIUM_TOP
        c.alignment = LEFT_VA if col <= 2 else RIGHT
        if col in (7, 8, 9):
            c.number_format = MONEY_FMT
        elif col in (3, 5):
            c.number_format = INT_FMT
    row += 2

    # ---- monthly summary ----
    m = bundle.monthly
    if m:
        ws.row_dimensions[row].height = ROW_H
        ws.merge_cells(f"A{row}:I{row}")
        c = ws.cell(row=row, column=1, value="MONTHLY SUMMARY")
        c.font = WHITE_BOLD
        c.fill = SECTION_FILL
        c.alignment = CENTER
        row += 1

        # Header
        for col, h in enumerate(["Item", "Units", f"Value ({cur})"], 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = WHITE_BOLD
            c.fill = SECTION_FILL
            c.alignment = CENTER
            c.border = THIN
        row += 1

        # Rows — NOTE: labels must NOT start with "=" (Excel interprets as formula)
        summary_rows: list[tuple[str, Any, Any]] = [
            ("Opening stock",       m.get("opening_units", 0),   m.get("opening_value",   Decimal("0"))),
            ("Purchases",           m.get("purchased_units", 0), m.get("purchased_value", Decimal("0"))),
            ("Consumption (net)",   m.get("consumed_units", 0),  m.get("consumed_value",  Decimal("0"))),
            ("Closing stock",       m.get("closing_units", 0),   m.get("closing_value",   Decimal("0"))),
        ]
        for i, (label, units, value) in enumerate(summary_rows):
            ws.row_dimensions[row].height = ROW_H
            is_closing = i == 3
            for col, v in enumerate([label, units, Decimal(str(value))], 1):
                c = ws.cell(row=row, column=col, value=v)
                if is_closing:
                    c.fill = SUMMARY_FILL
                    c.font = BOLD
                c.border = THIN
                c.alignment = LEFT_VA if col == 1 else RIGHT
                if col == 2:
                    c.number_format = INT_FMT
                elif col == 3:
                    c.number_format = MONEY_FMT
            row += 1

    ws.freeze_panes = "A5"
    return wb
