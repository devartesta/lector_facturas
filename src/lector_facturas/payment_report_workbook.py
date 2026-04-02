"""Payment tracking report workbook.

Builds an Excel file with one row per invoice, showing payment status,
due date, and a colour-coded semaphore.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_HEADER_FILL    = PatternFill("solid", fgColor="1F4E78")
_PAID_FILL      = PatternFill("solid", fgColor="C6EFCE")   # green
_OVERDUE_FILL   = PatternFill("solid", fgColor="FFC7CE")   # red
_PENDING_FILL   = PatternFill("solid", fgColor="FFEB9C")   # yellow
_DD_FILL        = PatternFill("solid", fgColor="BDD7EE")   # blue (direct debit)
_PARTIAL_FILL   = PatternFill("solid", fgColor="FFDAB3")   # orange

_WHITE_BOLD = Font(color="FFFFFF", bold=True)
_BOLD       = Font(bold=True)
_THIN       = Side(style="thin", color="D9D9D9")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

HEADERS = [
    "Company",
    "Period",
    "Supplier",
    "Invoice #",
    "Invoice date",
    "Net amount",
    "Currency",
    "Due date",
    "Status",
    "Payment date",
    "Method",
    "Payment amount",
    "Days overdue",
    "PDF",
]

COL_WIDTHS = [8, 8, 12, 16, 13, 13, 9, 13, 14, 13, 15, 15, 13, 8]


def _row_fill(row: dict) -> PatternFill | None:
    status = row.get("payment_status", "pending")
    is_overdue = row.get("is_overdue", False)
    if status == "paid":
        return _PAID_FILL
    if status == "direct_debit":
        return _DD_FILL
    if status == "partial":
        return _PARTIAL_FILL
    if is_overdue:
        return _OVERDUE_FILL
    return _PENDING_FILL


def _today() -> date:
    return date.today()


def _enrich(row: dict) -> dict:
    """Add computed is_overdue, is_settled, days_overdue fields."""
    today = _today()
    status = row.get("payment_status", "pending")
    due = row.get("payment_due_date")
    is_overdue = bool(due and due < today and status not in ("paid",))
    is_settled = bool(status == "direct_debit" and due and due <= today)
    days_overdue: int | None = None
    if due:
        delta = (today - due).days
        days_overdue = delta  # positive = overdue, negative = remaining
    return {**row, "is_overdue": is_overdue, "is_settled": is_settled, "days_overdue": days_overdue}


def build_payment_report(rows: list[dict], output: Path | BytesIO | None = None) -> BytesIO:
    """Build payment report workbook from list of document dicts.

    Returns a BytesIO with the xlsx content (also saves to *output* if given).
    """
    enriched = [_enrich(r) for r in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Status"
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 18

    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, row in enumerate(enriched, start=2):
        fill = _row_fill(row)

        def _cell(col: int, value):
            c = ws.cell(row=row_idx, column=col, value=value)
            if fill:
                c.fill = fill
            c.border = _THIN_BORDER
            c.alignment = Alignment(vertical="center")
            return c

        _cell(1,  row.get("company_code", ""))
        _cell(2,  row.get("period_yyyymm", ""))
        _cell(3,  row.get("supplier_code", ""))
        _cell(4,  row.get("invoice_number", ""))
        _cell(5,  row.get("invoice_date"))
        amt_cell = _cell(6, _decimal_or_none(row.get("net_amount")))
        amt_cell.number_format = '#,##0.00'
        _cell(7,  row.get("currency_code", ""))
        _cell(8,  row.get("payment_due_date"))
        _cell(9,  row.get("payment_status", "pending"))
        _cell(10, row.get("payment_date"))
        _cell(11, row.get("payment_method", ""))
        paid_amt = _cell(12, _decimal_or_none(row.get("payment_amount")))
        paid_amt.number_format = '#,##0.00'
        days = row.get("days_overdue")
        _cell(13, days)

        # PDF hyperlink
        drive_url = row.get("drive_url", "")
        pdf_cell = _cell(14, "PDF" if drive_url else "")
        if drive_url:
            pdf_cell.hyperlink = drive_url
            pdf_cell.font = Font(color="0563C1", underline="single",
                                 bold=(fill == _OVERDUE_FILL))
            pdf_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 15

    # Summary row
    summary_row = len(enriched) + 2
    ws.cell(row=summary_row, column=1, value="TOTAL").font = _BOLD

    buf = BytesIO()
    wb.save(buf)
    if isinstance(output, Path):
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_bytes(buf.getvalue())
    buf.seek(0)
    return buf


def _decimal_or_none(value) -> float | None:
    if value is None:
        return None
    try:
        return float(Decimal(str(value)))
    except Exception:
        return None
