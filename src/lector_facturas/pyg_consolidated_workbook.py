from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from lector_facturas.fx_rates import EcbFxService, FxRateAuditRow
from lector_facturas.pyg_inc_workbook import PygIncDataBundle, collect_pyg_inc_data
from lector_facturas.pyg_ltd_workbook import PygLtdDataBundle, collect_pyg_ltd_data
from lector_facturas.pyg_sl_workbook import PygSlDataBundle, collect_pyg_sl_data

REPORTING_CURRENCY = "EUR"
DISPLAY_TIMEZONE = ZoneInfo("Europe/Madrid")
MONTH_NAMES_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]
MONEY_FORMAT = '#,##0.00;[Red](#,##0.00);-'
PERCENT_FORMAT = '0.0%;[Red](0.0%);-'
ROW_HEIGHT = 12
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
SUBSECTION_FILL = PatternFill("solid", fgColor="EAF3F8")
KPI_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="D9D9D9")
PERCENT_ROW_FILL = PatternFill("solid", fgColor="FAFAFA")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
TITLE_FONT = Font(size=12, bold=True, color="FFFFFF")
THIN_TOP_BORDER = Border(top=Side(style="thin", color="A6A6A6"))
MEDIUM_TOP_BORDER = Border(top=Side(style="medium", color="7F7F7F"))


@dataclass(frozen=True)
class ConsolidatedPygBundle:
    year: int
    generated_at: datetime
    sl_bundle: PygSlDataBundle | None = field(default=None)
    ltd_bundle: PygLtdDataBundle | None = field(default=None)
    inc_bundle: PygIncDataBundle | None = field(default=None)


def default_output_path(root: Path, year: int) -> Path:
    return root / "output" / "spreadsheet" / f"pyg_consolidado_{year}.xlsx"


def month_keys(year: int) -> list[str]:
    return [f"{year}{month:02d}" for month in range(1, 13)]


def collect_pyg_consolidated_data(*, year: int, database_url: str | None) -> ConsolidatedPygBundle:
    if not database_url:
        return ConsolidatedPygBundle(year=year, generated_at=datetime.now(UTC))
    sl = collect_pyg_sl_data(year=year, database_url=database_url)
    ltd = collect_pyg_ltd_data(year=year, database_url=database_url)
    inc = collect_pyg_inc_data(year=year, database_url=database_url)
    return ConsolidatedPygBundle(year=year, generated_at=datetime.now(UTC), sl_bundle=sl, ltd_bundle=ltd, inc_bundle=inc)


def build_pyg_consolidated_workbook(bundle: ConsolidatedPygBundle, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    fx = EcbFxService()

    # ── Aggregate data into flat (yyyymm, line_key, amount) rows ───────────
    sl_rows, ltd_rows, inc_rows, fx_rate_rows = _aggregate_all(bundle, fx)

    # ── Write data sheets ──────────────────────────────────────────────────
    _write_data_sheet(wb, "i-sl", "amount_eur", sl_rows)
    _write_data_sheet(wb, "i-ltd", "amount_gbp", ltd_rows)
    _write_data_sheet(wb, "i-inc", "amount_usd", inc_rows)
    _write_fx_rates_sheet(wb, fx_rate_rows)

    # ── Build P&G with formulas ────────────────────────────────────────────
    _build_main_sheet(wb, bundle)
    _build_quarterly_sheet(wb, bundle)

    wb.save(output_path)
    return output_path


# ── Data aggregation ───────────────────────────────────────────────────────

def _aggregate_all(
    bundle: ConsolidatedPygBundle,
    fx: EcbFxService,
) -> tuple[list[list[Any]], list[list[Any]], list[list[Any]], list[list[Any]]]:
    sl_amounts: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    ltd_amounts: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    inc_amounts: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    fx_audit: list[FxRateAuditRow] = []

    if bundle.sl_bundle:
        b = bundle.sl_bundle
        for row in b.shopify_rows:
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency="EUR", yyyymm=row.yyyymm)
            sl_amounts[(row.yyyymm, "shopify")] += c.amount_reporting; fx_audit.append(a)
        for row in b.marketplace_rows:
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency="EUR", yyyymm=row.yyyymm)
            sl_amounts[(row.yyyymm, "marketplaces")] += c.amount_reporting; fx_audit.append(a)
        for row in b.rappel_rows:
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency="EUR", yyyymm=row.yyyymm)
            sl_amounts[(row.yyyymm, "rappels")] += c.amount_reporting; fx_audit.append(a)
        for row in b.supplies_rows:
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency="EUR", yyyymm=row.yyyymm)
            sl_amounts[(row.yyyymm, "supplies")] += c.amount_reporting; fx_audit.append(a)
        for row in b.service_rows:
            if row.line_item in {"Ltd", "Inc"}:
                # Shared services interco: excluded entirely from consolidated
                continue
            if row.detail == "renting_cnc":
                # CNC renting interco: income from HANNUN/QHANDS, offset by BBVACNC lease expense
                continue
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency="EUR", yyyymm=row.yyyymm)
            sl_amounts[(row.yyyymm, "services_ext")] += c.amount_reporting; fx_audit.append(a)
        for row in b.expense_rows:
            if row.subcategory in {"manufacturing", "logistics", "royalties", "marketing", "staff", "administration", "technology", "otros_gastos"}:
                # Exclude BBVACNC (CNC machine leasing): offset by renting_cnc income above
                if row.supplier_code == "BBVACNC":
                    continue
                c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency="EUR", yyyymm=row.yyyymm)
                sl_amounts[(row.yyyymm, row.subcategory)] += c.amount_reporting; fx_audit.append(a)
        for row in b.payment_fee_rows:
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency="EUR", yyyymm=row.yyyymm)
            sl_amounts[(row.yyyymm, "payment_fees")] += c.amount_reporting; fx_audit.append(a)
        for yyyymm, amt in b.otros_ingresos_by_period.items():
            sl_amounts[(yyyymm, "otros_ingresos")] += amt
        for yyyymm, amt in b.diferencias_divisas_by_period.items():
            sl_amounts[(yyyymm, "diferencias_divisas")] += amt

    if bundle.ltd_bundle:
        b = bundle.ltd_bundle
        RC = "GBP"
        for row in b.sales_rows:
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency=RC, yyyymm=row.yyyymm)
            ltd_amounts[(row.yyyymm, "product_sales")] += c.amount_reporting; fx_audit.append(a)
        for row in b.expense_rows:
            if row.subcategory in {"manufacturing", "logistics", "administration", "technology", "otros_gastos"}:
                # shared_services excluded entirely from consolidated
                c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency=RC, yyyymm=row.yyyymm)
                ltd_amounts[(row.yyyymm, row.subcategory)] += c.amount_reporting; fx_audit.append(a)
        for row in b.payment_fee_rows:
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency=RC, yyyymm=row.yyyymm)
            ltd_amounts[(row.yyyymm, "payment_fees")] += c.amount_reporting; fx_audit.append(a)
        for yyyymm, amt in b.otros_ingresos_by_period.items():
            c, a = fx.convert(amount=amt, source_currency="EUR", reporting_currency=RC, yyyymm=yyyymm)
            ltd_amounts[(yyyymm, "otros_ingresos")] += c.amount_reporting; fx_audit.append(a)
        for yyyymm, amt in b.diferencias_divisas_by_period.items():
            c, a = fx.convert(amount=amt, source_currency="EUR", reporting_currency=RC, yyyymm=yyyymm)
            ltd_amounts[(yyyymm, "diferencias_divisas")] += c.amount_reporting; fx_audit.append(a)
        for yyyymm, amt in b.frame_consumed_by_period.items():
            ltd_amounts[(yyyymm, "manufacturing")] += amt

    if bundle.inc_bundle:
        b = bundle.inc_bundle
        RC = "USD"
        for row in b.sales_rows:
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency=RC, yyyymm=row.yyyymm)
            inc_amounts[(row.yyyymm, "product_sales")] += c.amount_reporting; fx_audit.append(a)
        for row in b.expense_rows:
            if row.subcategory in {"manufacturing", "logistics", "administration", "technology", "otros_gastos"}:
                # shared_services excluded entirely from consolidated
                c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency=RC, yyyymm=row.yyyymm)
                inc_amounts[(row.yyyymm, row.subcategory)] += c.amount_reporting; fx_audit.append(a)
        for row in b.payment_fee_rows:
            c, a = fx.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency=RC, yyyymm=row.yyyymm)
            inc_amounts[(row.yyyymm, "payment_fees")] += c.amount_reporting; fx_audit.append(a)
        for yyyymm, amt in b.otros_ingresos_by_period.items():
            c, a = fx.convert(amount=amt, source_currency="EUR", reporting_currency=RC, yyyymm=yyyymm)
            inc_amounts[(yyyymm, "otros_ingresos")] += c.amount_reporting; fx_audit.append(a)
        for yyyymm, amt in b.diferencias_divisas_by_period.items():
            c, a = fx.convert(amount=amt, source_currency="EUR", reporting_currency=RC, yyyymm=yyyymm)
            inc_amounts[(yyyymm, "diferencias_divisas")] += c.amount_reporting; fx_audit.append(a)
        for yyyymm, amt in b.frame_consumed_by_period.items():
            inc_amounts[(yyyymm, "manufacturing")] += amt

    # Collect EOM GBP/EUR and USD/EUR rates for all months
    year = bundle.year
    fx_rates_rows: list[list[Any]] = []
    seen: set[tuple[str, str]] = set()
    for a in fx_audit:
        if a.currency_original in {"GBP", "USD"} and a.reporting_currency == "EUR":
            key = (a.yyyymm, a.currency_original)
            if key not in seen:
                seen.add(key)
                fx_rates_rows.append([a.yyyymm, a.currency_original, float(a.reference_rate)])
    # Ensure all 12 months have rates for GBP and USD
    for yyyymm in month_keys(year):
        for currency in ["GBP", "USD"]:
            if (yyyymm, currency) not in seen:
                _, a = fx.convert(amount=Decimal("1"), source_currency=currency, reporting_currency="EUR", yyyymm=yyyymm)
                seen.add((yyyymm, currency))
                fx_rates_rows.append([yyyymm, currency, float(a.reference_rate)])
    fx_rates_rows.sort(key=lambda r: (r[0], r[1]))

    def to_rows(amounts: dict[tuple[str, str], Decimal]) -> list[list[Any]]:
        return sorted(
            [[yyyymm, key, float(amt)] for (yyyymm, key), amt in amounts.items()],
            key=lambda r: (r[0], r[1]),
        )

    return to_rows(sl_amounts), to_rows(ltd_amounts), to_rows(inc_amounts), fx_rates_rows


# ── Sheet writers ──────────────────────────────────────────────────────────

def _write_data_sheet(wb: Workbook, title: str, amount_col: str, rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    headers = ["yyyymm", "line_key", amount_col]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(header) + 2)
        if header == amount_col:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).number_format = MONEY_FORMAT
    ws.freeze_panes = "A2"
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT


def _write_fx_rates_sheet(wb: Workbook, rows: list[list[Any]]) -> None:
    ws = wb.create_sheet("fx-rates")
    headers = ["yyyymm", "currency", "reference_rate"]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = "0.00000000"
        ws.row_dimensions[row_idx].height = ROW_HEIGHT
    ws.row_dimensions[1].height = ROW_HEIGHT
    ws.freeze_panes = "A2"
    # Note: reference_rate = foreign units per 1 EUR (ECB convention)
    # To convert GBP→EUR in formulas: amount_gbp / reference_rate
    # To convert USD→EUR in formulas: amount_usd / reference_rate


# ── Main P&G sheet ─────────────────────────────────────────────────────────

# P&G row definitions: (key, label, indent_level, row_type)
# row_type: "major" | "subtotal" | "section" | "detail" | "percent" | "blank"
_ROWS: list[tuple[str, str, int, str]] = [
    ("turnover",                "TURNOVER",                                  0, "major"),
    ("product_sales",           "Product sales",                             1, "subtotal"),
    ("shopify",                 "Shopify",                                   2, "section"),
    ("marketplaces",            "Marketplaces",                              2, "section"),
    ("services",                "Services",                                  2, "section"),
    ("rappels",                 "Rappels",                                   2, "section"),
    ("supplies",                "Supplies",                                  2, "section"),
    ("otros_ingresos",          "Otros ingresos",                            2, "section"),
    ("expenses",                "EXPENSES",                                  0, "major"),
    ("cogs",                    "COGS",                                      1, "subtotal"),
    ("manufacturing",           "Manufacturing",                             2, "section"),
    ("logistics",               "Logistics",                                 2, "section"),
    ("royalties",               "Royalties",                                 2, "section"),
    ("payment_fees",            "Payment fees",                              2, "section"),
    ("gross_margin",            "GROSS MARGIN (SALES-MANUFACTURING)",        0, "kpi"),
    ("gross_margin_pct",        "% GROSS MARGIN",                           0, "kpi"),
    ("contributive_margin",     "CONTRIBUTIVE MARGIN (TURNOVER-COGS)",       0, "kpi"),
    ("contributive_margin_pct", "% CONTRIBUTIVE MARGIN",                    0, "kpi"),
    ("opex",                    "OPEX",                                      1, "subtotal"),
    ("marketing",               "Marketing",                                 2, "section"),
    ("staff",                   "Staff",                                     2, "section"),
    ("administration",          "Administration",                            2, "section"),
    ("technology",              "Technology",                                2, "section"),
    ("otros_gastos",            "Otros gastos",                              2, "section"),
    ("diferencias_divisas",     "Diferencias divisas",                       1, "section"),
    ("profit",                  "PROFIT",                                    0, "kpi"),
    ("profit_pct",              "% Profit / product sales",                  0, "kpi"),
]


def _build_main_sheet(wb: Workbook, bundle: ConsolidatedPygBundle) -> None:
    ws = wb.create_sheet("P&G-CONSOLIDADO", 0)

    generated_at_local = (
        bundle.generated_at.astimezone(DISPLAY_TIMEZONE)
        if bundle.generated_at.tzinfo
        else bundle.generated_at.replace(tzinfo=UTC).astimezone(DISPLAY_TIMEZONE)
    )

    # Row 1: yyyymm (hidden)
    for idx, yyyymm in enumerate(month_keys(bundle.year), start=4):
        ws.cell(row=1, column=idx, value=yyyymm)
    ws["P1"] = "TOTAL"

    # Row 2: header
    ws["A2"] = f"P&G CONSOLIDADO {bundle.year} ({REPORTING_CURRENCY})"
    ws.merge_cells("A2:C2")
    for idx, yyyymm in enumerate(month_keys(bundle.year), start=4):
        ws.cell(row=2, column=idx, value=MONTH_NAMES_ES[idx - 4])
    ws["P2"] = "Total"

    # Row 3: generated_at
    ws["A3"] = f"Actualizado: {generated_at_local.strftime('%d/%m/%Y %H:%M')} h"
    ws.merge_cells("A3:C3")

    # Build row_map
    row_map: dict[str, int] = {}
    r = 4
    for key, label, _indent, _rtype in _ROWS:
        row_map[key] = r
        ws[f"A{r}"] = label
        r += 1

    # Write SUMIFS formulas for each month column
    for col_idx in range(4, 16):
        col = get_column_letter(col_idx)
        _write_col_formulas(ws, col, row_map)

    # Total column P
    for key in row_map:
        rr = row_map[key]
        ws[f"P{rr}"] = f"=SUM(D{rr}:O{rr})"
    # Override percentage rows in total column
    ws[f"P{row_map['gross_margin_pct']}"]        = f"=IFERROR(P{row_map['gross_margin']}/P{row_map['product_sales']},0)"
    ws[f"P{row_map['contributive_margin_pct']}"] = f"=IFERROR(P{row_map['contributive_margin']}/P{row_map['product_sales']},0)"
    ws[f"P{row_map['profit_pct']}"]              = f"=IFERROR(P{row_map['profit']}/P{row_map['product_sales']},0)"

    # Apply styles
    _apply_main_styles(ws, row_map)


def _write_col_formulas(ws, col: str, row_map: dict[str, int]) -> None:
    # FX rate lookups (reference_rate = foreign per 1 EUR)
    GBP = f"AVERAGEIFS('fx-rates'!$C:$C,'fx-rates'!$A:$A,{col}$1,'fx-rates'!$B:$B,\"GBP\")"
    USD = f"AVERAGEIFS('fx-rates'!$C:$C,'fx-rates'!$A:$A,{col}$1,'fx-rates'!$B:$B,\"USD\")"

    def sl(k: str) -> str:
        return f"SUMIFS('i-sl'!$C:$C,'i-sl'!$A:$A,{col}$1,'i-sl'!$B:$B,\"{k}\")"

    def ltd(k: str) -> str:
        return f"IFERROR(SUMIFS('i-ltd'!$C:$C,'i-ltd'!$A:$A,{col}$1,'i-ltd'!$B:$B,\"{k}\")/{GBP},0)"

    def inc(k: str) -> str:
        return f"IFERROR(SUMIFS('i-inc'!$C:$C,'i-inc'!$A:$A,{col}$1,'i-inc'!$B:$B,\"{k}\")/{USD},0)"

    rm = row_map

    # ── Data rows (SUMIFS on data sheets with inline FX conversion) ─────────
    ws[f"{col}{rm['shopify']}"]      = f"={sl('shopify')}+{ltd('product_sales')}+{inc('product_sales')}"
    ws[f"{col}{rm['product_sales']}"] = (
        f"={col}{rm['shopify']}+{col}{rm['marketplaces']}"
        f"+{col}{rm['rappels']}+{col}{rm['supplies']}"
    )
    ws[f"{col}{rm['marketplaces']}"]    = f"={sl('marketplaces')}"
    ws[f"{col}{rm['services']}"]        = f"={sl('services_ext')}"
    ws[f"{col}{rm['rappels']}"]         = f"={sl('rappels')}"
    ws[f"{col}{rm['supplies']}"]        = f"={sl('supplies')}"
    ws[f"{col}{rm['otros_ingresos']}"]  = f"={sl('otros_ingresos')}+{ltd('otros_ingresos')}+{inc('otros_ingresos')}"
    ws[f"{col}{rm['manufacturing']}"]   = f"={sl('manufacturing')}+{ltd('manufacturing')}+{inc('manufacturing')}"
    ws[f"{col}{rm['logistics']}"]       = f"={sl('logistics')}+{ltd('logistics')}+{inc('logistics')}"
    ws[f"{col}{rm['royalties']}"]       = f"={sl('royalties')}"
    ws[f"{col}{rm['payment_fees']}"]    = f"={sl('payment_fees')}+{ltd('payment_fees')}+{inc('payment_fees')}"
    ws[f"{col}{rm['marketing']}"]       = f"={sl('marketing')}"
    ws[f"{col}{rm['staff']}"]           = f"={sl('staff')}"
    ws[f"{col}{rm['administration']}"]  = f"={sl('administration')}+{ltd('administration')}+{inc('administration')}"
    ws[f"{col}{rm['technology']}"]      = f"={sl('technology')}+{ltd('technology')}+{inc('technology')}"
    ws[f"{col}{rm['otros_gastos']}"]    = f"={sl('otros_gastos')}+{ltd('otros_gastos')}+{inc('otros_gastos')}"
    ws[f"{col}{rm['diferencias_divisas']}"] = (
        f"={sl('diferencias_divisas')}+{ltd('diferencias_divisas')}+{inc('diferencias_divisas')}"
    )

    # ── Derived rows (formulas referencing other cells in this sheet) ────────
    ws[f"{col}{rm['turnover']}"] = f"={col}{rm['product_sales']}+{col}{rm['services']}+{col}{rm['otros_ingresos']}"
    ws[f"{col}{rm['cogs']}"] = (
        f"={col}{rm['manufacturing']}+{col}{rm['logistics']}"
        f"+{col}{rm['royalties']}+{col}{rm['payment_fees']}"
    )
    ws[f"{col}{rm['opex']}"] = (
        f"={col}{rm['marketing']}+{col}{rm['staff']}+{col}{rm['administration']}"
        f"+{col}{rm['technology']}+{col}{rm['otros_gastos']}"
    )
    ws[f"{col}{rm['expenses']}"]  = f"={col}{rm['cogs']}+{col}{rm['opex']}"
    ws[f"{col}{rm['gross_margin']}"] = f"={col}{rm['product_sales']}-{col}{rm['manufacturing']}"
    ws[f"{col}{rm['gross_margin_pct']}"] = (
        f"=IFERROR({col}{rm['gross_margin']}/{col}{rm['product_sales']},0)"
    )
    ws[f"{col}{rm['contributive_margin']}"] = f"={col}{rm['turnover']}-{col}{rm['cogs']}"
    ws[f"{col}{rm['contributive_margin_pct']}"] = (
        f"=IFERROR({col}{rm['contributive_margin']}/{col}{rm['product_sales']},0)"
    )
    ws[f"{col}{rm['profit']}"] = (
        f"={col}{rm['turnover']}-{col}{rm['cogs']}-{col}{rm['opex']}"
        f"-{col}{rm['diferencias_divisas']}"
    )
    ws[f"{col}{rm['profit_pct']}"] = (
        f"=IFERROR({col}{rm['profit']}/{col}{rm['product_sales']},0)"
    )


def _apply_main_styles(ws, row_map: dict[str, int]) -> None:
    # Build sets for styling
    key_to_rtype = {key: rtype for key, _lbl, _ind, rtype in _ROWS}
    key_to_indent = {key: ind for key, _lbl, ind, _rtype in _ROWS}

    major_rows    = {row_map[k] for k, _, _, t in _ROWS if t == "major"}
    kpi_rows      = {row_map[k] for k, _, _, t in _ROWS if t == "kpi"}
    subtotal_rows = {row_map[k] for k, _, _, t in _ROWS if t == "subtotal"}
    section_rows  = {row_map[k] for k, _, _, t in _ROWS if t == "section"}
    pct_rows      = {row_map[k] for k in ("gross_margin_pct", "contributive_margin_pct", "profit_pct")}

    # Header rows (1 and 2)
    ws.row_dimensions[1].hidden = True
    for cell in ws[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = TITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(size=9, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    max_row = max(row_map.values())
    for rr in range(2, max_row + 1):
        ws.row_dimensions[rr].height = ROW_HEIGHT

    for rr in range(4, max_row + 1):
        label_cell = ws.cell(row=rr, column=1)
        raw_label = str(label_cell.value or "")
        key = next((k for k, v in row_map.items() if v == rr), None)
        indent = key_to_indent.get(key, 0) if key else 0
        rtype = key_to_rtype.get(key, "section") if key else "section"

        if rtype == "major":
            for c in range(1, 17):
                ws.cell(row=rr, column=c).fill = SECTION_FILL
            label_cell.font = Font(bold=True, size=10)
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=0)
            _border_row(ws, rr, MEDIUM_TOP_BORDER)
        elif rtype == "kpi":
            for c in range(1, 17):
                ws.cell(row=rr, column=c).fill = KPI_FILL
            label_cell.font = Font(bold=True, size=10) if rr not in pct_rows else Font(italic=True, size=9)
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=0)
            _border_row(ws, rr, MEDIUM_TOP_BORDER)
        elif rtype == "subtotal":
            for c in range(1, 17):
                ws.cell(row=rr, column=c).fill = SUBSECTION_FILL
            label_cell.font = Font(bold=True, size=9)
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            _border_row(ws, rr, THIN_TOP_BORDER)
        elif rtype == "section":
            label_cell.font = Font(size=9)
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent + 1)

        # Number format
        for c in range(4, 17):
            cell = ws.cell(row=rr, column=c)
            cell.number_format = PERCENT_FORMAT if rr in pct_rows else MONEY_FORMAT
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(size=9)

        # Total column (P)
        ws.cell(row=rr, column=16).fill = TOTAL_FILL
        if rr in major_rows | subtotal_rows | kpi_rows:
            ws.cell(row=rr, column=16).font = BOLD

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["P"].width = 14
    for i in range(4, 16):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "D4"
    ws.sheet_view.showGridLines = False


def _border_row(ws, row: int, border) -> None:
    for c in range(1, 17):
        ws.cell(row=row, column=c).border = border


def _build_quarterly_sheet(wb: Workbook, bundle: ConsolidatedPygBundle) -> None:
    """Quarterly summary referencing P&G-CONSOLIDADO monthly cells via SUM formulas."""
    ws = wb.create_sheet("P&G-TRIMESTRAL")
    src = "'P&G-CONSOLIDADO'"

    # In P&G-CONSOLIDADO: D=Jan, E=Feb, F=Mar, G=Apr, H=May, I=Jun,
    #                      J=Jul, K=Aug, L=Sep, M=Oct, N=Nov, O=Dec
    # Quarters (column ranges in source sheet):
    QUARTERS = [
        ("Q1", "D", "F"),  # Jan-Mar
        ("Q2", "G", "I"),  # Apr-Jun
        ("Q3", "J", "L"),  # Jul-Sep
        ("Q4", "M", "O"),  # Oct-Dec
    ]
    # Output columns: A=label, B=Q1, C=Q2, D=Q3, E=Q4, F=Total
    Q_COLS = ["B", "C", "D", "E"]
    TOTAL_COL = "F"

    year = bundle.year

    # ── Build row_map (same rows as main sheet, starting at row 4) ───────────
    row_map: dict[str, int] = {}
    r = 4
    for key, label, _indent, _rtype in _ROWS:
        row_map[key] = r
        ws[f"A{r}"] = label
        r += 1

    pct_keys = {"gross_margin_pct", "contributive_margin_pct", "profit_pct"}

    # ── Row 1: hidden quarter labels (for reference) ─────────────────────────
    ws.row_dimensions[1].hidden = True

    # ── Row 2: header ─────────────────────────────────────────────────────────
    ws["A2"] = f"P&G CONSOLIDADO TRIMESTRAL {year} ({REPORTING_CURRENCY})"
    for col, (q_label, _, __) in zip(Q_COLS, QUARTERS):
        ws[f"{col}2"] = q_label
    ws[f"{TOTAL_COL}2"] = "Total"

    # ── Row 3: generated_at (pulled from P&G-CONSOLIDADO) ────────────────────
    ws["A3"] = f"={src}!A3"

    # ── Data rows: SUM quarterly ranges from main sheet ──────────────────────
    for key, _label, _indent, _rtype in _ROWS:
        src_row = row_map[key]
        is_pct = key in pct_keys

        for q_col, (_q_label, c_from, c_to) in zip(Q_COLS, QUARTERS):
            if is_pct:
                # Percentages: recompute from quarterly totals
                if key == "gross_margin_pct":
                    num_key, den_key = "gross_margin", "product_sales"
                elif key == "contributive_margin_pct":
                    num_key, den_key = "contributive_margin", "product_sales"
                else:
                    num_key, den_key = "profit", "product_sales"
                num_row = row_map[num_key]
                den_row = row_map[den_key]
                ws[f"{q_col}{src_row}"] = (
                    f"=IFERROR(SUM({src}!{c_from}{num_row}:{c_to}{num_row})"
                    f"/SUM({src}!{c_from}{den_row}:{c_to}{den_row}),0)"
                )
            else:
                ws[f"{q_col}{src_row}"] = f"=SUM({src}!{c_from}{src_row}:{c_to}{src_row})"

        # Total column
        if is_pct:
            if key == "gross_margin_pct":
                num_key, den_key = "gross_margin", "product_sales"
            elif key == "contributive_margin_pct":
                num_key, den_key = "contributive_margin", "product_sales"
            else:
                num_key, den_key = "profit", "product_sales"
            ws[f"{TOTAL_COL}{src_row}"] = (
                f"=IFERROR({src}!P{row_map[num_key]}/{src}!P{row_map[den_key]},0)"
            )
        else:
            ws[f"{TOTAL_COL}{src_row}"] = f"={src}!P{src_row}"

    # ── Styles (reuse same logic as main sheet) ────────────────────────────────
    key_to_rtype  = {key: rtype  for key, _lbl, _ind, rtype  in _ROWS}
    key_to_indent = {key: indent for key, _lbl, indent, _rtype in _ROWS}

    major_rows    = {row_map[k] for k, _, _, t in _ROWS if t == "major"}
    kpi_rows      = {row_map[k] for k, _, _, t in _ROWS if t == "kpi"}
    subtotal_rows = {row_map[k] for k, _, _, t in _ROWS if t == "subtotal"}
    pct_rows      = {row_map[k] for k in pct_keys}

    # Header row 2
    for cell in ws[2]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = TITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(size=9, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    max_row = max(row_map.values())
    for rr in range(2, max_row + 1):
        ws.row_dimensions[rr].height = ROW_HEIGHT

    ALL_COLS = Q_COLS + [TOTAL_COL]
    n_cols = len(ALL_COLS) + 1  # A + Q1-Q4 + Total

    for rr in range(4, max_row + 1):
        label_cell = ws.cell(row=rr, column=1)
        key = next((k for k, v in row_map.items() if v == rr), None)
        indent = key_to_indent.get(key, 0) if key else 0
        rtype = key_to_rtype.get(key, "section") if key else "section"

        if rtype == "major":
            for c in range(1, n_cols + 1):
                ws.cell(row=rr, column=c).fill = SECTION_FILL
            label_cell.font = Font(bold=True, size=10)
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=0)
            for c in range(1, n_cols + 1):
                ws.cell(row=rr, column=c).border = MEDIUM_TOP_BORDER
        elif rtype == "kpi":
            for c in range(1, n_cols + 1):
                ws.cell(row=rr, column=c).fill = KPI_FILL
            label_cell.font = Font(bold=True, size=10) if rr not in pct_rows else Font(italic=True, size=9)
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=0)
            for c in range(1, n_cols + 1):
                ws.cell(row=rr, column=c).border = MEDIUM_TOP_BORDER
        elif rtype == "subtotal":
            for c in range(1, n_cols + 1):
                ws.cell(row=rr, column=c).fill = SUBSECTION_FILL
            label_cell.font = Font(bold=True, size=9)
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for c in range(1, n_cols + 1):
                ws.cell(row=rr, column=c).border = THIN_TOP_BORDER
        else:
            label_cell.font = Font(size=9)
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent + 1)

        for col_letter in ALL_COLS:
            cell = ws[f"{col_letter}{rr}"]
            cell.number_format = PERCENT_FORMAT if rr in pct_rows else MONEY_FORMAT
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(size=9)

        # Total column styling
        ws[f"{TOTAL_COL}{rr}"].fill = TOTAL_FILL
        if rr in major_rows | subtotal_rows | kpi_rows:
            ws[f"{TOTAL_COL}{rr}"].font = BOLD

    # Column widths
    ws.column_dimensions["A"].width = 40
    for col_letter in ALL_COLS:
        ws.column_dimensions[col_letter].width = 16
    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
