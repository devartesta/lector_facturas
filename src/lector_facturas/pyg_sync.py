from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from lector_facturas.pyg_consolidated_workbook import build_pyg_consolidated_workbook, collect_pyg_consolidated_data, default_output_path as default_output_path_consolidated
from lector_facturas.google_drive import GoogleDriveClient
from lector_facturas.pyg_inc_workbook import build_pyg_inc_workbook, collect_pyg_inc_data, default_output_path as default_output_path_inc
from lector_facturas.pyg_ltd_workbook import build_pyg_ltd_workbook, collect_pyg_ltd_data, default_output_path as default_output_path_ltd
from lector_facturas.pyg_sl_workbook import build_pyg_sl_workbook, collect_pyg_sl_data, default_output_path
from lector_facturas.settings import AppSettings
from lector_facturas.stock_detail_workbook import StockDetailBundle, collect_stock_detail, build_stock_detail_bytes
from lector_facturas.payment_reconciliation import build_reconciliation
from lector_facturas.payment_reconciliation_workbook import build_reconciliation_workbook
from lector_facturas.folder_structure import ENTITY_ALIASES


@dataclass(frozen=True)
class PygSlSyncResult:
    year: int
    drive_folder_id: str
    drive_file_id: str
    drive_file_name: str
    drive_file_url: str
    local_output_path: str
    replaced_file_ids: tuple[str, ...]


@dataclass(frozen=True)
class PygIncSyncResult:
    year: int
    drive_folder_id: str
    drive_file_id: str
    drive_file_name: str
    drive_file_url: str
    local_output_path: str
    replaced_file_ids: tuple[str, ...]


@dataclass(frozen=True)
class PygLtdSyncResult:
    year: int
    drive_folder_id: str
    drive_file_id: str
    drive_file_name: str
    drive_file_url: str
    local_output_path: str
    replaced_file_ids: tuple[str, ...]


@dataclass(frozen=True)
class PygConsolidatedSyncResult:
    year: int
    drive_folder_id: str
    drive_file_id: str
    drive_file_name: str
    drive_file_url: str
    local_output_path: str
    replaced_file_ids: tuple[str, ...]


def sync_pyg_sl_to_drive(
    *,
    settings: AppSettings,
    year: int,
    drive_folder_id: str | None = None,
    file_name: str | None = None,
    output_root: Path | None = None,
) -> PygSlSyncResult:
    if not settings.google_oauth_ready:
        raise RuntimeError("Google OAuth is not configured.")
    if not settings.drive_root_folder_id and not drive_folder_id:
        raise RuntimeError("Google Drive root folder is not configured.")
    database_url = _database_url()
    if not database_url:
        raise RuntimeError("DATABASE_URL is not configured.")

    root = output_root or Path(__file__).resolve().parents[2]
    output_path = default_output_path(root, year)
    bundle = collect_pyg_sl_data(year=year, database_url=database_url)
    build_pyg_sl_workbook(bundle, output_path)

    client = GoogleDriveClient(settings.to_drive_config())
    target_folder_id = drive_folder_id or settings.drive_root_folder_id
    target_name = file_name or f"pyg_sl_{year}.xlsx"
    existing = client.list_files(parent_id=target_folder_id, name=target_name)
    for item in existing:
        client.trash_file(file_id=str(item["id"]))
    created = client.upload_file(
        name=target_name,
        parent_id=target_folder_id,
        content=output_path.read_bytes(),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return PygSlSyncResult(
        year=year,
        drive_folder_id=target_folder_id,
        drive_file_id=str(created["id"]),
        drive_file_name=str(created["name"]),
        drive_file_url=str(created.get("webViewLink", "")),
        local_output_path=str(output_path),
        replaced_file_ids=tuple(str(item["id"]) for item in existing),
    )


def sync_pyg_inc_to_drive(
    *,
    settings: AppSettings,
    year: int,
    drive_folder_id: str | None = None,
    file_name: str | None = None,
    output_root: Path | None = None,
) -> PygIncSyncResult:
    if not settings.google_oauth_ready:
        raise RuntimeError("Google OAuth is not configured.")
    if not settings.drive_root_folder_id and not drive_folder_id:
        raise RuntimeError("Google Drive root folder is not configured.")
    database_url = _database_url()
    if not database_url:
        raise RuntimeError("DATABASE_URL is not configured.")

    root = output_root or Path(__file__).resolve().parents[2]
    output_path = default_output_path_inc(root, year)
    bundle = collect_pyg_inc_data(year=year, database_url=database_url)
    build_pyg_inc_workbook(bundle, output_path)

    client = GoogleDriveClient(settings.to_drive_config())
    target_folder_id = drive_folder_id or settings.drive_root_folder_id
    target_name = file_name or f"pyg_inc_{year}.xlsx"
    existing = client.list_files(parent_id=target_folder_id, name=target_name)
    for item in existing:
        client.trash_file(file_id=str(item["id"]))
    created = client.upload_file(
        name=target_name,
        parent_id=target_folder_id,
        content=output_path.read_bytes(),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return PygIncSyncResult(
        year=year,
        drive_folder_id=target_folder_id,
        drive_file_id=str(created["id"]),
        drive_file_name=str(created["name"]),
        drive_file_url=str(created.get("webViewLink", "")),
        local_output_path=str(output_path),
        replaced_file_ids=tuple(str(item["id"]) for item in existing),
    )


def sync_pyg_ltd_to_drive(
    *,
    settings: AppSettings,
    year: int,
    drive_folder_id: str | None = None,
    file_name: str | None = None,
    output_root: Path | None = None,
) -> PygLtdSyncResult:
    if not settings.google_oauth_ready:
        raise RuntimeError("Google OAuth is not configured.")
    if not settings.drive_root_folder_id and not drive_folder_id:
        raise RuntimeError("Google Drive root folder is not configured.")
    database_url = _database_url()
    if not database_url:
        raise RuntimeError("DATABASE_URL is not configured.")

    root = output_root or Path(__file__).resolve().parents[2]
    output_path = default_output_path_ltd(root, year)
    bundle = collect_pyg_ltd_data(year=year, database_url=database_url)
    build_pyg_ltd_workbook(bundle, output_path)

    client = GoogleDriveClient(settings.to_drive_config())
    target_folder_id = drive_folder_id or settings.drive_root_folder_id
    target_name = file_name or f"pyg_ltd_{year}.xlsx"
    existing = client.list_files(parent_id=target_folder_id, name=target_name)
    for item in existing:
        client.trash_file(file_id=str(item["id"]))
    created = client.upload_file(
        name=target_name,
        parent_id=target_folder_id,
        content=output_path.read_bytes(),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return PygLtdSyncResult(
        year=year,
        drive_folder_id=target_folder_id,
        drive_file_id=str(created["id"]),
        drive_file_name=str(created["name"]),
        drive_file_url=str(created.get("webViewLink", "")),
        local_output_path=str(output_path),
        replaced_file_ids=tuple(str(item["id"]) for item in existing),
    )


def sync_pyg_consolidated_to_drive(
    *,
    settings: AppSettings,
    year: int,
    drive_folder_id: str | None = None,
    file_name: str | None = None,
    output_root: Path | None = None,
) -> PygConsolidatedSyncResult:
    if not settings.google_oauth_ready:
        raise RuntimeError("Google OAuth is not configured.")
    if not settings.drive_root_folder_id and not drive_folder_id:
        raise RuntimeError("Google Drive root folder is not configured.")
    database_url = _database_url()
    if not database_url:
        raise RuntimeError("DATABASE_URL is not configured.")

    root = output_root or Path(__file__).resolve().parents[2]
    output_path = default_output_path_consolidated(root, year)
    bundle = collect_pyg_consolidated_data(year=year, database_url=database_url)
    build_pyg_consolidated_workbook(bundle, output_path)

    client = GoogleDriveClient(settings.to_drive_config())
    target_folder_id = drive_folder_id or settings.drive_root_folder_id
    target_name = file_name or f"pyg_consolidado_{year}.xlsx"
    existing = client.list_files(parent_id=target_folder_id, name=target_name)
    for item in existing:
        client.trash_file(file_id=str(item["id"]))
    created = client.upload_file(
        name=target_name,
        parent_id=target_folder_id,
        content=output_path.read_bytes(),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return PygConsolidatedSyncResult(
        year=year,
        drive_folder_id=target_folder_id,
        drive_file_id=str(created["id"]),
        drive_file_name=str(created["name"]),
        drive_file_url=str(created.get("webViewLink", "")),
        local_output_path=str(output_path),
        replaced_file_ids=tuple(str(item["id"]) for item in existing),
    )


# ---------------------------------------------------------------------------
# Stock detail sync
# ---------------------------------------------------------------------------

# Maps fabricante -> Drive entity key (used in ENTITY_ALIASES)
FABRICANTE_ENTITY: dict[str, str] = {
    "TGI": "Inc",
    "Proco": "Ltd",
}

DRIVE_ROOT_NAME = "ARTESTA - 6. Finances"


@dataclass(frozen=True)
class StockDetailSyncResult:
    fabricante: str
    mes_yyyymm: str
    drive_folder_id: str
    drive_file_id: str
    drive_file_name: str
    drive_file_url: str


def sync_stock_detail_to_drive(
    *,
    settings: AppSettings,
    fabricante: str,
    mes_yyyymm: str,
) -> StockDetailSyncResult:
    """Build a per-SKU stock detail xlsx and upload it to the month's
    expenses/cogs/stock Drive folder.

    Path: <root> / <entity> / <year> / <yyyymm> / expenses / cogs / stock
    File: stock_{fabricante.lower()}_{yyyymm}.xlsx
    """
    if not settings.google_oauth_ready:
        raise RuntimeError("Google OAuth is not configured.")
    if not settings.drive_root_folder_id:
        raise RuntimeError("GOOGLE_DRIVE_ROOT_FOLDER_ID is not configured.")

    database_url = _database_url()
    if not database_url:
        raise RuntimeError("DATABASE_URL is not configured.")

    entity_key = FABRICANTE_ENTITY.get(fabricante)
    if not entity_key:
        raise ValueError(f"Unknown fabricante '{fabricante}'. Expected one of: {list(FABRICANTE_ENTITY)}")

    entity_name = ENTITY_ALIASES[entity_key]
    year = mes_yyyymm[:4]
    file_name = f"stock_{fabricante.lower()}_{mes_yyyymm}.xlsx"

    # Navigate Drive path: root -> entity -> year -> yyyymm -> expenses -> cogs -> stock
    client = GoogleDriveClient(settings.to_drive_config())
    folder_id = settings.drive_root_folder_id
    for part in [entity_name, year, mes_yyyymm, "expenses", "cogs", "stock"]:
        folder_id = str(client.ensure_folder(name=part, parent_id=folder_id)["id"])

    # Build xlsx bytes
    bundle = collect_stock_detail(
        fabricante=fabricante,
        mes_yyyymm=mes_yyyymm,
        database_url=database_url,
    )
    content = build_stock_detail_bytes(bundle)

    # Replace existing file (if any) and upload new version
    existing = client.list_files(parent_id=folder_id, name=file_name)
    for item in existing:
        client.trash_file(file_id=str(item["id"]))
    created = client.upload_file(
        name=file_name,
        parent_id=folder_id,
        content=content,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    return StockDetailSyncResult(
        fabricante=fabricante,
        mes_yyyymm=mes_yyyymm,
        drive_folder_id=folder_id,
        drive_file_id=str(created["id"]),
        drive_file_name=str(created["name"]),
        drive_file_url=str(created.get("webViewLink", "")),
    )


@dataclass(frozen=True)
class PaymentReconciliationSyncResult:
    company_code: str
    period_yyyymm: str
    drive_folder_id: str
    drive_file_id: str
    drive_file_name: str
    drive_file_url: str
    shopify_only_accounting: int
    shopify_only_payment: int
    shopify_amount_diff: int
    paypal_only_accounting: int
    paypal_only_payment: int
    paypal_amount_diff: int


# Maps company_code → Drive entity folder name (matches ENTITY_ALIASES / folder_structure)
_COMPANY_ENTITY: dict[str, str] = {
    "SL":  "Artesta Store, S.L",
    "LTD": "Artesta Stores (UK) Ltd",
    "INC": "Artesta Inc",
}

# Drive path under <entity>/<year>/<yyyymm>/
_RECON_FOLDER_PATH = ["income", "sales", "shopify"]


def sync_payment_reconciliation_to_drive(
    *,
    settings: AppSettings,
    company_code: str,
    period_yyyymm: str,
    drive_folder_id: str | None = None,
) -> PaymentReconciliationSyncResult:
    """Build the payment reconciliation xlsx and upload it to Drive.

    Default upload path (when drive_folder_id is not supplied):
      <root> / <entity> / <year> / <yyyymm> / income / sales / shopify
      e.g. ARTESTA - 6. Finances / Artesta Store, S.L / 2026 / 202602 / income / sales / shopify

    File name: cotejo_pagos_{company_code.lower()}_{period_yyyymm}.xlsx
    """
    if not settings.google_oauth_ready:
        raise RuntimeError("Google OAuth is not configured.")
    if not settings.drive_root_folder_id and not drive_folder_id:
        raise RuntimeError("GOOGLE_DRIVE_ROOT_FOLDER_ID is not configured.")

    database_url = _database_url()
    if not database_url:
        raise RuntimeError("DATABASE_URL is not configured.")

    shopify_config = settings.to_shopify_config() if settings.shopify_ready else None
    report = build_reconciliation(
        database_url=database_url,
        company_code=company_code,
        period_yyyymm=period_yyyymm,
        shopify_config=shopify_config,
    )

    client = GoogleDriveClient(settings.to_drive_config())

    if drive_folder_id:
        target_folder_id = drive_folder_id
    else:
        # Navigate: root → entity → year → yyyymm → income → sales → shopify
        entity_name = _COMPANY_ENTITY.get(company_code.upper())
        if not entity_name:
            raise ValueError(
                f"Unknown company_code '{company_code}'. "
                f"Expected one of: {list(_COMPANY_ENTITY)}"
            )
        year = period_yyyymm[:4]
        folder_id = settings.drive_root_folder_id
        for part in [entity_name, year, period_yyyymm] + _RECON_FOLDER_PATH:
            folder_id = str(client.ensure_folder(name=part, parent_id=folder_id)["id"])
        target_folder_id = folder_id

    file_name = f"payment_reconciliation_{company_code.lower()}_{period_yyyymm}.xlsx"
    existing = client.list_files(parent_id=target_folder_id, name=file_name)

    # Before trashing the existing file, read the Bank Transfer sheet to
    # preserve manually entered Pagado / Fecha cobro values.
    existing_payments: dict[str, tuple[str, str | None]] = {}
    if existing:
        try:
            from io import BytesIO as _BytesIO
            from openpyxl import load_workbook as _load_workbook
            existing_bytes = client.download_file_bytes(file_id=str(existing[0]["id"]))
            wb_prev = _load_workbook(_BytesIO(existing_bytes), data_only=True)
            if "Bank Transfers" in wb_prev.sheetnames or "Bank Transfer" in wb_prev.sheetnames:
                ws_bt = wb_prev["Bank Transfers"] if "Bank Transfers" in wb_prev.sheetnames else wb_prev["Bank Transfer"]
                # Rows 1-2 are title/subtitle, row 3 is column headers, data starts row 4.
                # Col A = order_name, Col J (10) = Pagado, Col K (11) = Fecha cobro.
                for bt_row in ws_bt.iter_rows(min_row=4, values_only=True):
                    order_name = bt_row[0]
                    if not order_name:
                        continue
                    order_name_str = str(order_name).strip()
                    # Skip non-order rows (totals, summary lines, etc.)
                    if not (order_name_str.startswith("AS-") or order_name_str.startswith("UK-")):
                        continue
                    pagado      = str(bt_row[9]).strip()  if bt_row[9]  else "Pending"
                    fecha_cobro = str(bt_row[10]).strip() if bt_row[10] else None
                    # Backwards-compat: map old Spanish values to English
                    if pagado in ("Sí", "Si"):   pagado = "Yes"
                    if pagado == "Pendiente":     pagado = "Pending"
                    existing_payments[order_name_str] = (pagado, fecha_cobro)
        except Exception as exc:
            # Non-fatal: if reading fails we regenerate without preserved data
            print(
                f"[recon] WARNING: could not read existing Bank Transfer data: {exc}",
                flush=True,
            )

    content = build_reconciliation_workbook(report, existing_payments=existing_payments or None)

    for item in existing:
        client.trash_file(file_id=str(item["id"]))
    created = client.upload_file(
        name=file_name,
        parent_id=target_folder_id,
        content=content,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    return PaymentReconciliationSyncResult(
        company_code=company_code,
        period_yyyymm=period_yyyymm,
        drive_folder_id=target_folder_id,
        drive_file_id=str(created["id"]),
        drive_file_name=str(created["name"]),
        drive_file_url=str(created.get("webViewLink", "")),
        shopify_only_accounting=len(report.shopify.only_accounting),
        shopify_only_payment=len(report.shopify.only_payment),
        shopify_amount_diff=len(report.shopify.amount_diff),
        paypal_only_accounting=len(report.paypal.only_accounting),
        paypal_only_payment=len(report.paypal.only_payment),
        paypal_amount_diff=len(report.paypal.amount_diff),
    )


@dataclass(frozen=True)
class GestoriaSyncResult:
    company_code: str
    period_yyyymm: str
    drive_folder_id: str
    drive_file_id: str
    drive_file_name: str
    drive_file_url: str
    n_resumen_rows: int
    n_detalle_rows: int


_GESTORIA_FOLDER_PATH = ["income", "sales", "vat"]


def sync_gestoria_to_drive(
    *,
    settings: AppSettings,
    company_code: str,
    period_yyyymm: str,
    drive_folder_id: str | None = None,
) -> GestoriaSyncResult:
    """Build the VAT gestoría xlsx and upload to Drive.

    Default upload path (when drive_folder_id is not supplied):
      <root> / <entity> / <year> / <yyyymm> / income / sales / vat
      e.g. ARTESTA - 6. Finances / Artesta Store, S.L / 2026 / 202602 / income / sales / vat

    File name: shopify_sales_{company_code.lower()}_{period_yyyymm}.xlsx
    """
    if not settings.google_oauth_ready:
        raise RuntimeError("Google OAuth is not configured.")
    if not settings.drive_root_folder_id and not drive_folder_id:
        raise RuntimeError("GOOGLE_DRIVE_ROOT_FOLDER_ID is not configured.")

    database_url = _database_url()
    if not database_url:
        raise RuntimeError("DATABASE_URL is not configured.")

    from lector_facturas.gestoria_workbook import collect_gestoria_data, build_gestoria_workbook

    report = collect_gestoria_data(
        database_url=database_url,
        company_code=company_code,
        period_yyyymm=period_yyyymm,
    )

    client = GoogleDriveClient(settings.to_drive_config())

    if drive_folder_id:
        target_folder_id = drive_folder_id
    else:
        # Navigate: root → entity → year → yyyymm → income → sales → vat
        entity_name = _COMPANY_ENTITY.get(company_code.upper())
        if not entity_name:
            raise ValueError(
                f"Unknown company_code '{company_code}'. "
                f"Expected one of: {list(_COMPANY_ENTITY)}"
            )
        year = period_yyyymm[:4]
        folder_id = settings.drive_root_folder_id
        for part in [entity_name, year, period_yyyymm] + _GESTORIA_FOLDER_PATH:
            folder_id = str(client.ensure_folder(name=part, parent_id=folder_id)["id"])
        target_folder_id = folder_id

    file_name = f"shopify_sales_{company_code.lower()}_{period_yyyymm}.xlsx"
    existing  = client.list_files(parent_id=target_folder_id, name=file_name)

    content = build_gestoria_workbook(report)

    for item in existing:
        client.trash_file(file_id=str(item["id"]))
    created = client.upload_file(
        name=file_name,
        parent_id=target_folder_id,
        content=content,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    return GestoriaSyncResult(
        company_code=company_code,
        period_yyyymm=period_yyyymm,
        drive_folder_id=target_folder_id,
        drive_file_id=str(created["id"]),
        drive_file_name=str(created["name"]),
        drive_file_url=str(created.get("webViewLink", "")),
        n_resumen_rows=len(report.resumen_rows),
        n_detalle_rows=len(report.detalle_rows),
    )


def _database_url() -> str:
    import os

    return os.environ.get("DATABASE_URL", "").strip()
