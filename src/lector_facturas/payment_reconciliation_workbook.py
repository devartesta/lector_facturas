"""Build the payment reconciliation Excel workbook.

Four sheets:
  1. Resumen       — overview table (counts + totals) and legend
  2. Chargebacks   — inventory of all chargebacks last 12 months (open + won)
  3. Shopify       — detailed cotejo rows for Shopify Payments channel
  4. PayPal        — detailed cotejo rows for PayPal channel

Each detail sheet (Shopify / PayPal) has three sections:
  · Solo en ventas  — accounting entry with no matching payment
  · Solo en pago    — payment charge with no accounting entry this period
  · Diferencias     — both present but amounts differ > 0.01

Columns per section:
  Nº Pedido | Fecha | País | Moneda | Importe ventas | Importe pago | Diferencia
  | Link Shopify | T. regalo | Estado disp. | Comentarios
"""
from __future__ import annotations

from datetime import datetime, UTC
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from lector_facturas.payment_reconciliation import (
    B2BOrderRow,
    CB_LOST,
    CB_LOST_DAYS,
    CB_OPEN,
    CB_WON,
    ChannelReconciliation,
    ChargebackInventoryRow,
    GiftCardRow,
    ReconciliationReport,
    ReconciliationRow,
)

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")   # dark blue   – title row
SECTION_FILL = PatternFill("solid", fgColor="2E74B5")   # mid blue    – section header
COL_HDR_FILL = PatternFill("solid", fgColor="D6E4F0")   # light blue  – column headers
WARN_FILL    = PatternFill("solid", fgColor="FFF2CC")   # yellow      – diff row highlight
CB_OPEN_FILL = PatternFill("solid", fgColor="FCE4D6")   # orange      – dispute open
CB_WON_FILL  = PatternFill("solid", fgColor="E2EFDA")   # light green – dispute won
CB_LOST_FILL = PatternFill("solid", fgColor="FF0000")   # red         – dispute lost (no reversal > 75d)
LEGEND_FILL  = PatternFill("solid", fgColor="F2F2F2")   # grey        – legend background
TOTAL_FILL   = PatternFill("solid", fgColor="BDD7EE")   # blue-grey   – totals row
OK_FILL      = PatternFill("solid", fgColor="E2EFDA")   # light green – zero issues

WHITE_BOLD  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
BOLD        = Font(bold=True, name="Calibri", size=10)
BOLD_SMALL  = Font(bold=True, name="Calibri", size=9)
NORMAL      = Font(name="Calibri", size=10)
ITALIC_GREY = Font(name="Calibri", size=9, italic=True, color="595959")
LINK_FONT   = Font(name="Calibri", size=10, color="0563C1", underline="single")
THIN_SIDE   = Side(style="thin", color="BFBFBF")
THIN        = Border(top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=False)
LEFT_WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
RIGHT       = Alignment(horizontal="right",  vertical="center", wrap_text=False)
MONEY_FMT   = '#,##0.00;[Red](#,##0.00);-'
ROW_H       = 14

MONTH_NAMES_EN = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# Detail columns for Shopify / PayPal sheets
COLUMNS = [
    ("Order #",         16, LEFT,   None),
    ("Date",            12, CENTER, None),
    ("Country",          6, CENTER, None),
    ("Currency",         7, CENTER, None),
    ("Accounting amt.", 14, RIGHT,  MONEY_FMT),
    ("Payment amt.",    14, RIGHT,  MONEY_FMT),
    ("Difference",      12, RIGHT,  MONEY_FMT),
    ("Shopify link",    14, LEFT,   None),
    ("Gift card",        9, CENTER, None),
    ("Chargeback",      18, CENTER, None),
    ("Comments",        35, LEFT,   None),
]

# Section definitions: (attribute, title, explanation)
SECTION_DEFS = [
    (
        "only_accounting",
        "ACCOUNTING ONLY",
        "Orders recorded in accounting with no matching charge in the payment channel this month."
        " May indicate a pending payment, payment via another method (bank transfer, etc.)"
        " or a gateway assignment error.",
    ),
    (
        "only_payment",
        "PAYMENT ONLY",
        "Charges in the payment channel with no accounting entry for this period."
        " Check whether an invoice is missing or the order belongs to a different accounting month.",
    ),
    (
        "amount_diff",
        "AMOUNT DIFFERENCES",
        "The amount settled by the payment channel does not match the accounting entry."
        " Common causes: partial refunds, chargebacks, unrecorded fees or exchange-rate differences.",
    ),
]

# Chargeback inventory columns
CB_COLUMNS = [
    ("Channel",          10, CENTER, None),
    ("Order #",          16, LEFT,   None),
    ("Order date",       12, CENTER, None),
    ("Country",           6, CENTER, None),
    ("Currency",          7, CENTER, None),
    ("Sale amount",      14, RIGHT,  MONEY_FMT),
    ("Hold date",        13, CENTER, None),
    ("Amount held",      15, RIGHT,  MONEY_FMT),
    ("Resolution date",  13, CENTER, None),
    ("Amount recovered", 16, RIGHT,  MONEY_FMT),
    ("Net impact",       13, RIGHT,  MONEY_FMT),
    ("Status",           13, CENTER, None),
    ("Shopify link",     14, LEFT,   None),
    ("Comments",         35, LEFT,   None),
]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_reconciliation_workbook(
    report: ReconciliationReport,
    existing_payments: dict[str, tuple[str, str | None]] | None = None,
) -> bytes:
    """Return the xlsx as bytes.

    existing_payments maps order_name → (pagado, fecha_cobro) from a previously
    generated workbook, so manual user edits in the Bank Transfer sheet survive
    monthly regenerations.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    year        = int(report.period_yyyymm[:4])
    month       = int(report.period_yyyymm[4:])
    month_label = f"{MONTH_NAMES_EN[month - 1]} {year}"

    _add_summary_sheet(wb, report, month_label, existing_payments=existing_payments)
    _add_channel_sheet(wb, "Shopify Payments", report.shopify, report, month_label)
    _add_channel_sheet(wb, "PayPal",           report.paypal,  report, month_label)
    _add_bank_transfer_sheet(wb, report, month_label, existing_payments=existing_payments)
    _add_chargeback_sheet(wb, report, month_label)
    _add_gift_card_sheet(wb, report, month_label)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _add_summary_sheet(
    wb: Workbook,
    report: ReconciliationReport,
    month_label: str,
    existing_payments: dict[str, tuple[str, str | None]] | None = None,
) -> None:
    ws = wb.create_sheet("Summary")

    # Columns: A=label, B=count, C=Shopify, D=PayPal, E=Bank Transfer, F=Total, G=notas
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 40
    LAST_COL = "G"
    NCOLS    = 7

    def _sh(r: int, h: int = ROW_H) -> None:
        ws.row_dimensions[r].height = h

    def _cell(r: int, col: int, value: Any, *,
              font: Font | None = None,
              fill: PatternFill | None = None,
              align: Alignment = LEFT,
              fmt: str | None = None,
              border: Border | None = THIN) -> None:
        c = ws.cell(row=r, column=col, value=value)
        c.font      = font  or NORMAL
        c.alignment = align
        if fill:   c.fill   = fill
        if border: c.border = border
        if fmt:    c.number_format = fmt

    def _section_hdr(r: int, title: str) -> None:
        _sh(r, ROW_H + 2)
        ws.merge_cells(f"A{r}:{LAST_COL}{r}")
        _cell(r, 1, title, font=WHITE_BOLD, fill=SECTION_FILL, align=LEFT)

    def _blank_row(r: int) -> None:
        _sh(r, 6)
        ws.merge_cells(f"A{r}:{LAST_COL}{r}")
        ws.cell(row=r, column=1)

    row = 1

    # ── Title ──────────────────────────────────────────────────────────────────
    _sh(row, 22)
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    _cell(row, 1,
          f"Payment reconciliation — {report.company_code} — {month_label}"
          f"   |   Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}",
          font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
          fill=HEADER_FILL, align=CENTER)
    row += 1

    # ── Subtitle ───────────────────────────────────────────────────────────────
    _sh(row, 22)
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    _cell(row, 1,
          "Accounting vs payment channel comparison (Shopify Payments, PayPal and Bank Transfer). "
          "Bank Transfer: accounting = manual orders for the period; collected = orders marked 'Yes' in the Bank Transfer tab. "
          "Differences (Shopify / PayPal) are classified by gift card, chargeback and other.",
          font=ITALIC_GREY, align=LEFT_WRAP, border=None)
    row += 1
    _blank_row(row); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1 — TOTALES DEL PERÍODO
    # ══════════════════════════════════════════════════════════════════════════
    _section_hdr(row, "PERIOD TOTALS"); row += 1

    # Column sub-headers: A=label, B=count, C=Shopify, D=PayPal, E=Bank Transfer, F=Total, G=notes
    _sh(row)
    for col, hdr in enumerate(
        ["", "# orders", "Shopify Payments", "PayPal", "Bank Transfer", "Total", ""], 1
    ):
        _cell(row, col, hdr, font=BOLD, fill=COL_HDR_FILL, align=CENTER)
    row += 1

    _D0 = Decimal("0")
    sa  = report.shopify_accounting_total
    sp  = report.shopify_payment_total
    pa  = report.paypal_accounting_total
    pp  = report.paypal_payment_total

    # Bank Transfer totals from b2b_orders + existing_payments
    ep  = existing_payments or {}
    b2b_acct = sum((o.total for o in report.b2b_orders), _D0).quantize(Decimal("0.01"))
    b2b_paid = sum(
        (o.total for o in report.b2b_orders if ep.get(o.order_name, ("Pendiente",))[0] == "Sí"),
        _D0,
    ).quantize(Decimal("0.01"))
    b2b_diff = (b2b_paid - b2b_acct).quantize(Decimal("0.01"))

    # Each tuple: (label, n_orders, shopify_val, paypal_val, b2b_val, total_val, note)
    totals_rows = [
        ("Accounting amount",
         None, sa, pa, b2b_acct,
         (sa + pa + b2b_acct).quantize(Decimal("0.01")),
         "Net accounting sales (including refunds). Bank Transfer = manual orders for the period."),
        ("Amount collected",
         None, sp, pp, b2b_paid,
         (sp + pp + b2b_paid).quantize(Decimal("0.01")),
         "Shopify/PayPal: settled by the channel. Bank Transfer: orders marked 'Yes' in the tab."),
        ("Difference",
         None,
         (sp - sa).quantize(Decimal("0.01")),
         (pp - pa).quantize(Decimal("0.01")),
         b2b_diff,
         ((sp + pp + b2b_paid) - (sa + pa + b2b_acct)).quantize(Decimal("0.01")),
         "Positive = collected more than accounted; negative = pending or unreconciled."),
    ]
    for label, cnt, sh_val, pp_val, b2b_val, tot_val, note in totals_rows:
        _sh(row)
        is_diff_row = (label == "Difference")
        row_fill    = (WARN_FILL if is_diff_row and abs(tot_val) > Decimal("0.01") else
                       OK_FILL   if is_diff_row else None)
        font_row    = BOLD if is_diff_row else NORMAL
        _cell(row, 1, label,  font=BOLD,      fill=row_fill)
        _cell(row, 2, cnt,    font=font_row,  fill=row_fill, align=CENTER)
        _cell(row, 3, float(sh_val)  if sh_val  is not None else None,
              font=font_row, fill=row_fill, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 4, float(pp_val)  if pp_val  is not None else None,
              font=font_row, fill=row_fill, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 5, float(b2b_val) if b2b_val is not None else None,
              font=font_row, fill=row_fill, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 6, float(tot_val) if tot_val is not None else None,
              font=BOLD,     fill=row_fill, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 7, note,   font=ITALIC_GREY, fill=row_fill, align=LEFT_WRAP)
        row += 1

    _blank_row(row); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2 — DESGLOSE DE DIFERENCIAS
    # ══════════════════════════════════════════════════════════════════════════
    _section_hdr(row, "DIFFERENCE BREAKDOWN"); row += 1

    # Explanation
    _sh(row, 26)
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    _cell(row, 1,
          "Orders with a difference between accounting and payment channel, classified by cause. "
          "Gift card: gateway includes gift_card. "
          "Chargeback: open, won or likely-lost dispute. "
          "Bank Transfer: orders not yet marked as collected ('Yes'). "
          "Other: differences with no known cause.",
          font=Font(name="Calibri", size=9, italic=True, color="404040"),
          fill=LEGEND_FILL, align=LEFT_WRAP)
    row += 1

    # Column headers: A=Category, B=#orders, C=Shopify diff, D=PayPal diff, E=BT diff, F=Total, G=Action
    _sh(row)
    for col, hdr in enumerate(["Category", "# orders", "Shopify diff", "PayPal diff",
                                "Bank Transfer diff", "Total diff", "Recommended action"], 1):
        _cell(row, col, hdr, font=BOLD, fill=COL_HDR_FILL, align=CENTER)
    row += 1

    # Compute effective diff per row across all buckets+channels
    def _eff_diff(r: ReconciliationRow, bucket: str) -> Decimal:
        if bucket == "only_accounting":
            return -(r.accounting_amount or _D0)
        if bucket == "only_payment":
            return  (r.payment_amount    or _D0)
        return (r.diff or _D0)  # amount_diff

    # Gather all rows with differences, tagged by channel and bucket
    all_diff: list[tuple[str, str, ReconciliationRow]] = []
    for ch_label, recon in (("Shopify", report.shopify), ("PayPal", report.paypal)):
        for bucket in ("only_accounting", "only_payment", "amount_diff"):
            for r in getattr(recon, bucket):
                all_diff.append((ch_label, bucket, r))

    def _classify(r: ReconciliationRow) -> str:
        if r.is_gift_card:   return "gift"
        if r.is_chargeback:  return "cb"
        return "other"

    # Bank Transfer pending: orders not marked 'Yes'
    n_bt_pending = sum(
        1 for o in report.b2b_orders if ep.get(o.order_name, ("Pending",))[0] != "Yes"
    )
    bt_pending_diff = sum(
        (-(o.total) for o in report.b2b_orders if ep.get(o.order_name, ("Pending",))[0] != "Yes"),
        _D0,
    ).quantize(Decimal("0.01"))

    diff_cats: dict[str, dict[str, Any]] = {
        "gift":  {"label": "Gift card",      "sh": _D0, "pp": _D0, "bt": _D0, "n": 0,
                  "fill": LEGEND_FILL,
                  "note": "Orders partially paid with gift card (channel settles less than accounting)."},
        "cb":    {"label": "Chargeback",     "sh": _D0, "pp": _D0, "bt": _D0, "n": 0,
                  "fill": CB_OPEN_FILL,
                  "note": "Open, won or likely-lost disputes. See Chargebacks tab."},
        "other": {"label": "Other",          "sh": _D0, "pp": _D0, "bt": bt_pending_diff, "n": n_bt_pending,
                  "fill": WARN_FILL,
                  "note": "Review in the Shopify / PayPal / Bank Transfer tabs."},
    }
    for ch_label, bucket, r in all_diff:
        cat  = _classify(r)
        d    = _eff_diff(r, bucket)
        diff_cats[cat]["n"] += 1
        if ch_label == "Shopify":
            diff_cats[cat]["sh"] = (diff_cats[cat]["sh"] + d).quantize(Decimal("0.01"))
        else:
            diff_cats[cat]["pp"] = (diff_cats[cat]["pp"] + d).quantize(Decimal("0.01"))

    tot_n  = 0
    tot_sh = _D0
    tot_pp = _D0
    tot_bt = _D0
    for cat_key in ("gift", "cb", "other"):
        d      = diff_cats[cat_key]
        n      = d["n"]
        sh_d   = d["sh"].quantize(Decimal("0.01"))
        pp_d   = d["pp"].quantize(Decimal("0.01"))
        bt_d   = d["bt"].quantize(Decimal("0.01"))
        tot_d  = (sh_d + pp_d + bt_d).quantize(Decimal("0.01"))
        tot_n += n
        tot_sh = (tot_sh + sh_d).quantize(Decimal("0.01"))
        tot_pp = (tot_pp + pp_d).quantize(Decimal("0.01"))
        tot_bt = (tot_bt + bt_d).quantize(Decimal("0.01"))
        r_fill = d["fill"] if n > 0 else OK_FILL
        _sh(row)
        _cell(row, 1, d["label"], font=BOLD, fill=r_fill)
        _cell(row, 2, n,          font=BOLD if n > 0 else NORMAL, fill=r_fill, align=CENTER)
        _cell(row, 3, float(sh_d) if sh_d else None,
              font=NORMAL, fill=r_fill, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 4, float(pp_d) if pp_d else None,
              font=NORMAL, fill=r_fill, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 5, float(bt_d) if bt_d else None,
              font=NORMAL, fill=r_fill, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 6, float(tot_d) if tot_d else None,
              font=BOLD, fill=r_fill, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 7, "✓ No differences in this category" if n == 0 else d["note"],
              font=ITALIC_GREY, fill=r_fill, align=LEFT_WRAP)
        row += 1

    # Total row
    tot_tot = (tot_sh + tot_pp + tot_bt).quantize(Decimal("0.01"))
    _sh(row)
    _cell(row, 1, f"TOTAL  ({tot_n} orders)", font=BOLD, fill=TOTAL_FILL)
    _cell(row, 2, tot_n,           font=BOLD, fill=TOTAL_FILL, align=CENTER)
    _cell(row, 3, float(tot_sh) if tot_sh else None, font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row, 4, float(tot_pp) if tot_pp else None, font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row, 5, float(tot_bt) if tot_bt else None, font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row, 6, float(tot_tot),  font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row, 7, None,            font=NORMAL, fill=TOTAL_FILL)
    row += 1

    _blank_row(row); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 3 — CHARGEBACKS (últimos 12 meses)
    # ══════════════════════════════════════════════════════════════════════════
    _section_hdr(row, "CHARGEBACKS (last 12 months)"); row += 1

    shopify_cbs = [r for r in report.chargeback_inventory if r.channel == "Shopify"]
    paypal_cbs  = [r for r in report.chargeback_inventory if r.channel == "PayPal"]

    def _cb_impact(items: list, status: str) -> Decimal:
        return sum(
            ((r.net_impact or _D0) for r in items if r.status == status), _D0,
        ).quantize(Decimal("0.01"))

    cb_summary = [
        ("Shopify Payments", CB_OPEN_FILL,
         sum(1 for r in shopify_cbs if r.status == CB_OPEN),
         f"In dispute (< {CB_LOST_DAYS}d): amount held {_cb_impact(shopify_cbs, CB_OPEN):,.2f}"),
        ("Shopify Payments", CB_LOST_FILL,
         sum(1 for r in shopify_cbs if r.status == CB_LOST),
         f"Likely lost (> {CB_LOST_DAYS}d no reversal): impact {_cb_impact(shopify_cbs, CB_LOST):,.2f}"),
        ("Shopify Payments", CB_WON_FILL,
         sum(1 for r in shopify_cbs if r.status == CB_WON),
         "Won — funds recovered."),
        ("PayPal", CB_OPEN_FILL,
         sum(1 for r in paypal_cbs if r.status == CB_OPEN),
         f"In dispute (< {CB_LOST_DAYS}d) — confirm status in PayPal portal."),
        ("PayPal", CB_LOST_FILL,
         sum(1 for r in paypal_cbs if r.status == CB_LOST),
         f"Likely lost (> {CB_LOST_DAYS}d no T1112): impact {_cb_impact(paypal_cbs, CB_LOST):,.2f} — confirm in PayPal."),
    ]
    for ch_lbl, r_fill, count, note in cb_summary:
        _sh(row)
        _cell(row, 1, ch_lbl, font=BOLD,                               fill=r_fill)
        _cell(row, 2, count,  font=BOLD if count > 0 else NORMAL,       fill=r_fill, align=CENTER)
        ws.merge_cells(f"C{row}:{LAST_COL}{row}")
        _cell(row, 3, note,   font=Font(name="Calibri", size=9),        fill=r_fill, align=LEFT_WRAP)
        for col in (4, 5, 6, 7):
            _cell(row, col, None, fill=r_fill)
        row += 1

    _blank_row(row); row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # LEGEND
    # ══════════════════════════════════════════════════════════════════════════
    _section_hdr(row, "COLOR LEGEND"); row += 1

    legend_items = [
        (CB_OPEN_FILL, "Orange — Open dispute",
         "Active chargeback: payment channel has held the funds. Respond before the deadline."),
        (CB_LOST_FILL, "Red — Likely lost",
         f"Dispute with no reversal after {CB_LOST_DAYS} days. Confirm loss in Shopify/PayPal."),
        (CB_WON_FILL,  "Green — Won dispute / No differences",
         "Dispute resolved in our favour, or category with no issues."),
        (WARN_FILL,    "Yellow — Differences to review",
         "Amount differences or orders only on one side. See Shopify / PayPal tabs."),
    ]
    for fill, label, explanation in legend_items:
        _sh(row, 30)
        _cell(row, 1, "", fill=fill)
        _cell(row, 2, label, font=BOLD_SMALL, fill=fill, align=LEFT)
        ws.merge_cells(f"C{row}:{LAST_COL}{row}")
        _cell(row, 3, explanation,
              font=Font(name="Calibri", size=9), fill=LEGEND_FILL, align=LEFT_WRAP)
        for col in (4, 5, 6, 7):
            _cell(row, col, None, fill=LEGEND_FILL)
        row += 1

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Chargeback inventory sheet
# ---------------------------------------------------------------------------

def _add_chargeback_sheet(
    wb: Workbook,
    report: ReconciliationReport,
    month_label: str,
) -> None:
    ws = wb.create_sheet("Chargebacks")

    for i, (_, w, _, _) in enumerate(CB_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ncols    = len(CB_COLUMNS)
    last_col = get_column_letter(ncols)
    row      = 1

    # Title
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1,
        value=(f"Chargeback inventory — {report.company_code}"
               f"   |   Last 12 months   |   Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"))
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL; c.alignment = CENTER
    row += 1

    # Summary line
    inventory   = report.chargeback_inventory
    shopify_cbs = [r for r in inventory if r.channel == "Shopify"]
    paypal_cbs  = [r for r in inventory if r.channel == "PayPal"]
    n_shopify_open = sum(1 for r in shopify_cbs if r.status == CB_OPEN)
    n_shopify_won  = sum(1 for r in shopify_cbs if r.status == CB_WON)

    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{last_col}{row}")
    n_shopify_lost2 = sum(1 for r in shopify_cbs if r.status == CB_LOST)
    n_paypal_lost2  = sum(1 for r in paypal_cbs  if r.status == CB_LOST)
    c = ws.cell(row=row, column=1,
        value=(f"Shopify: {n_shopify_open} in dispute · {n_shopify_lost2} likely lost · {n_shopify_won} won"
               f"   |   PayPal: {len(paypal_cbs)} total ({n_paypal_lost2} likely lost > {CB_LOST_DAYS}d) — confirm in PayPal"
               f"   |   Red = lost · Orange = in dispute · Green = won"))
    c.font = ITALIC_GREY; c.alignment = LEFT_WRAP
    row += 2

    if not inventory:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws.cell(row=row, column=1, value="✓ No chargebacks in the last 12 months")
        c.font = Font(name="Calibri", size=10, italic=True, color="548235")
        c.alignment = LEFT
        return

    # ---- Shopify section ----
    row = _write_cb_section(
        ws, row, last_col,
        title="SHOPIFY PAYMENTS",
        note=("Status detected automatically: 'In dispute' = active hold, no reversal yet; "
              "'Won' = channel returned the funds (dispute_reversal received)."),
        items=shopify_cbs,
        paypal_warning=False,
    )

    # ---- PayPal section ----
    row = _write_cb_section(
        ws, row, last_col,
        title="PAYPAL",
        note=("⚠ PayPal does not generate a 'dispute resolved' transaction in our data (no T1112). "
              "All disputes appear as 'In dispute' even if already closed. "
              "Confirm actual status in the PayPal Resolution Center."),
        items=paypal_cbs,
        paypal_warning=True,
    )

    ws.freeze_panes = "A4"


def _write_cb_section(
    ws: Any,
    row: int,
    last_col: str,
    title: str,
    note: str,
    items: list[ChargebackInventoryRow],
    paypal_warning: bool,
) -> int:
    """Write one channel block (header + col headers + rows + totals) in the chargeback sheet."""
    ncols = len(CB_COLUMNS) - 1  # CB_COLUMNS includes Canal which we drop here
    # We use CB_COLUMNS[1:] (skip the Canal column since it's in the section header)
    detail_cols = CB_COLUMNS[1:]
    detail_last = get_column_letter(len(detail_cols))

    # Section header
    ws.row_dimensions[row].height = ROW_H + 2
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value=title)
    c.font = WHITE_BOLD; c.fill = SECTION_FILL; c.alignment = LEFT
    row += 1

    # Note
    ws.row_dimensions[row].height = 30
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value=note)
    c.font = Font(name="Calibri", size=9,
                  italic=True,
                  color="843C0C" if paypal_warning else "404040")
    c.fill = LEGEND_FILL; c.alignment = LEFT_WRAP
    row += 1

    if not items:
        ws.row_dimensions[row].height = ROW_H
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws.cell(row=row, column=1, value="✓ No chargebacks in the last 12 months")
        c.font = Font(name="Calibri", size=10, italic=True, color="548235")
        c.alignment = LEFT
        return row + 2

    # Column headers (skip Canal column)
    ws.row_dimensions[row].height = ROW_H
    for col, (header, _, _, _) in enumerate(detail_cols, 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = BOLD; c.fill = COL_HDR_FILL; c.alignment = CENTER; c.border = THIN
    row += 1

    # Data rows
    for r in items:
        ws.row_dimensions[row].height = ROW_H
        fill = _cb_fill(r.status, True)

        rev_date_val  = r.reversal_date if r.reversal_date else "—"
        rev_date_font = ITALIC_GREY if not r.reversal_date else None

        # Status label including days open
        if paypal_warning and r.status == CB_LOST:
            status_label = f"Likely lost ({r.days_open}d)"
        elif paypal_warning and r.status == CB_OPEN:
            status_label = f"In dispute ({r.days_open}d)"
        elif r.status == CB_LOST:
            status_label = f"Likely lost ({r.days_open}d)"
        elif r.status == CB_OPEN:
            status_label = f"In dispute ({r.days_open}d)"
        else:
            status_label = r.status  # Won

        vals: list[tuple[Any, Alignment, str | None, Font | None]] = [
            (r.order_name,            LEFT,   None,      BOLD),
            (r.order_date,            CENTER, None,      None),
            (r.shipping_country_code, CENTER, None,      None),
            (r.currency,              CENTER, None,      None),
            (r.accounting_amount,     RIGHT,  MONEY_FMT, None),
            (r.withdrawal_date,       CENTER, None,      None),
            (r.withdrawal_amount,     RIGHT,  MONEY_FMT, None),
            (rev_date_val,            CENTER, None,      rev_date_font),
            (r.reversal_amount,       RIGHT,  MONEY_FMT, None),
            (r.net_impact,            RIGHT,  MONEY_FMT, BOLD),
            (status_label,            CENTER, None,      BOLD),
            (None,                    LEFT,   None,      LINK_FONT),
            ("",                      LEFT,   None,      None),
        ]
        for col, (value, align, fmt, font) in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=value)
            c.alignment = align; c.border = THIN; c.fill = fill
            c.font = font if font else NORMAL
            if fmt: c.number_format = fmt

        # Hyperlink (column 12 = Link Shopify, now without Canal col)
        if r.shopify_url:
            lc = ws.cell(row=row, column=12)
            lc.value = "Ver pedido"; lc.hyperlink = r.shopify_url; lc.font = LINK_FONT

        row += 1

    # Subtotals
    ws.row_dimensions[row].height = ROW_H
    _D0     = Decimal("0")
    t_acct  = sum(((r.accounting_amount or _D0) for r in items), _D0).quantize(Decimal("0.01"))
    t_w     = sum(((r.withdrawal_amount or _D0) for r in items), _D0).quantize(Decimal("0.01"))
    t_v     = sum(((r.reversal_amount   or _D0) for r in items), _D0).quantize(Decimal("0.01"))
    t_net   = sum(((r.net_impact        or _D0) for r in items), _D0).quantize(Decimal("0.01"))
    total_vals: list[tuple[Any, Alignment, str | None]] = [
        (f"TOTAL {title} ({len(items)} orders)", LEFT, None),
        (None, CENTER, None), (None, CENTER, None), (None, CENTER, None),
        (t_acct, RIGHT, MONEY_FMT),
        (None, CENTER, None),
        (t_w,   RIGHT, MONEY_FMT),
        (None, CENTER, None),
        (t_v,   RIGHT, MONEY_FMT),
        (t_net, RIGHT, MONEY_FMT),
        (None, CENTER, None), (None, LEFT, None), (None, LEFT, None),
    ]
    for col, (val, align, fmt) in enumerate(total_vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = BOLD; c.fill = TOTAL_FILL; c.alignment = align; c.border = THIN
        if fmt: c.number_format = fmt

    return row + 2


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cb_label(status: str | None, is_chargeback: bool) -> str:
    """Human-readable chargeback label for the detail sheet column."""
    if status == CB_WON:
        return "Chargeback — Won"
    if status == CB_LOST:
        return "Chargeback — Likely lost"
    if status == CB_OPEN:
        return "Chargeback — In dispute"
    if is_chargeback:
        return "Chargeback"
    return ""


def _cb_fill(status: str | None, is_chargeback: bool) -> PatternFill | None:
    """Row fill for chargeback status."""
    if status == CB_LOST:
        return CB_LOST_FILL
    if status == CB_WON:
        return CB_WON_FILL
    if status == CB_OPEN or is_chargeback:
        return CB_OPEN_FILL
    return None


# ---------------------------------------------------------------------------
# Detail sheet builder
# ---------------------------------------------------------------------------

def _add_channel_sheet(
    wb: Workbook,
    channel: str,
    recon: ChannelReconciliation,
    report: ReconciliationReport,
    month_label: str,
) -> None:
    ws = wb.create_sheet(channel)

    for i, (_, w, _, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{get_column_letter(len(COLUMNS))}{row}")
    c = ws.cell(row=row, column=1)
    c.value = (f"Accounting vs {channel} — {report.company_code} — {month_label}  "
               f"| Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL; c.alignment = CENTER
    row += 1

    # Quick summary
    only_a = len(recon.only_accounting)
    only_p = len(recon.only_payment)
    diff_n = len(recon.amount_diff)
    ws.row_dimensions[row].height = ROW_H
    ws.merge_cells(f"A{row}:{get_column_letter(len(COLUMNS))}{row}")
    c = ws.cell(row=row, column=1)
    c.value = (f"Accounting only: {only_a}   |   Payment only: {only_p}   |   "
               f"Amount differences: {diff_n}   "
               f"{'✓ All matched' if only_a + only_p + diff_n == 0 else '⚠ Review differences — see Summary tab'}")
    c.font = ITALIC_GREY; c.alignment = CENTER
    row += 2

    for attr, sec_title, sec_explanation in SECTION_DEFS:
        rows_data: list[ReconciliationRow] = getattr(recon, attr)
        row = _write_section(ws, row, sec_title, sec_explanation, rows_data)

    ws.freeze_panes = "A4"


def _write_section(
    ws: Any,
    row: int,
    title: str,
    explanation: str,
    rows: list[ReconciliationRow],
) -> int:
    ncols    = len(COLUMNS)
    last_col = get_column_letter(ncols)

    # Section title
    ws.row_dimensions[row].height = ROW_H + 2
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value=title)
    c.font = WHITE_BOLD; c.fill = SECTION_FILL; c.alignment = LEFT
    row += 1

    # Explanation
    ws.row_dimensions[row].height = 28
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value=explanation)
    c.font = Font(name="Calibri", size=9, italic=True, color="404040")
    c.fill = LEGEND_FILL; c.alignment = LEFT_WRAP
    row += 1

    if not rows:
        ws.row_dimensions[row].height = ROW_H
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws.cell(row=row, column=1, value="✓ No differences")
        c.font = Font(name="Calibri", size=10, italic=True, color="548235")
        c.alignment = LEFT
        row += 2
        return row

    # Column headers
    ws.row_dimensions[row].height = ROW_H
    for col, (header, _, _, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = BOLD; c.fill = COL_HDR_FILL; c.alignment = CENTER; c.border = THIN
    row += 1

    # Data rows
    for data_row in rows:
        ws.row_dimensions[row].height = ROW_H
        fill = _cb_fill(data_row.chargeback_status, data_row.is_chargeback)
        if fill is None and data_row.diff is not None and abs(data_row.diff) > Decimal("10"):
            fill = WARN_FILL

        vals: list[tuple[Any, Alignment, str | None, Font | None]] = [
            (data_row.order_name,             LEFT,   None,      None),
            (data_row.order_date,             CENTER, None,      None),
            (data_row.shipping_country_code,  CENTER, None,      None),
            (data_row.currency,               CENTER, None,      None),
            (data_row.accounting_amount,      RIGHT,  MONEY_FMT, None),
            (data_row.payment_amount,         RIGHT,  MONEY_FMT, None),
            (data_row.diff,                   RIGHT,  MONEY_FMT, None),
            (None,                            LEFT,   None,      LINK_FONT),
            ("Sí" if data_row.is_gift_card else "",                CENTER, None, None),
            (_cb_label(data_row.chargeback_status, data_row.is_chargeback), CENTER, None, BOLD if data_row.is_chargeback else None),
            ("",                              LEFT,   None,      None),
        ]
        for col, (value, align, num_fmt, font_override) in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=value)
            c.alignment = align; c.border = THIN
            if num_fmt: c.number_format = num_fmt
            c.font = font_override if font_override else NORMAL
            if fill: c.fill = fill

        if data_row.shopify_url:
            lc = ws.cell(row=row, column=8)
            lc.value = "Ver pedido"; lc.hyperlink = data_row.shopify_url; lc.font = LINK_FONT

        row += 1

    # Totals row
    ws.row_dimensions[row].height = ROW_H
    sum_acct = sum((r.accounting_amount or Decimal("0")) for r in rows).quantize(Decimal("0.01"))
    sum_pay  = sum((r.payment_amount    or Decimal("0")) for r in rows).quantize(Decimal("0.01"))
    sum_diff = sum((r.diff              or Decimal("0")) for r in rows).quantize(Decimal("0.01"))
    total_vals: list[tuple[Any, Alignment, str | None]] = [
        (f"TOTAL ({len(rows)} orders)", LEFT, None),
        (None, CENTER, None), (None, CENTER, None), (None, CENTER, None),
        (sum_acct, RIGHT, MONEY_FMT),
        (sum_pay,  RIGHT, MONEY_FMT),
        (sum_diff, RIGHT, MONEY_FMT),
        (None, LEFT, None), (None, CENTER, None), (None, CENTER, None), (None, LEFT, None),
    ]
    for col, (val, align, fmt) in enumerate(total_vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = BOLD; c.fill = TOTAL_FILL; c.alignment = align; c.border = THIN
        if fmt: c.number_format = fmt
    row += 2

    return row


# ---------------------------------------------------------------------------
# B2B / manual payments sheet
# ---------------------------------------------------------------------------

B2B_COLS = [
    ("Order #",          14, LEFT,   None),
    ("Order date",       12, CENTER, None),
    ("Country",           6, CENTER, None),
    ("Category",         12, CENTER, None),
    ("VAT %",             7, CENTER, "0%"),
    ("Net amount",       15, RIGHT,  MONEY_FMT),
    ("VAT",              12, RIGHT,  MONEY_FMT),
    ("Total (incl. VAT)",16, RIGHT,  MONEY_FMT),
    ("Shopify link",     14, LEFT,   None),
    ("Paid",             12, CENTER, None),
    ("Payment date",     14, CENTER, None),
    ("Comments",         35, LEFT,   None),
]

YES_FILL  = PatternFill("solid", fgColor="E2EFDA")   # green  – paid
PEND_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow – pending


def _add_bank_transfer_sheet(
    wb: Workbook,
    report: ReconciliationReport,
    month_label: str,
    existing_payments: dict[str, tuple[str, str | None]] | None = None,
) -> None:
    """Sheet listing manual-gateway (Bank Transfer) orders with payment tracking columns.

    existing_payments maps order_name → (paid, payment_date) from a previously
    generated version so that manually entered data survives regeneration.
    Rows with paid == "Yes" get green fill; all others get yellow.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet("Bank Transfers")
    orders = report.b2b_orders

    ncols    = len(B2B_COLS)
    last_col = get_column_letter(ncols)

    for i, (_, w, _, _) in enumerate(B2B_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1,
        value=(f"Bank Transfer — {report.company_code} — {month_label}"
               f"   |   Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"))
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL; c.alignment = CENTER
    row += 1

    # Subtitle
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1,
        value=('Orders with gateway "manual", [] or [""] (bank transfer), '
               'excluding Hannun and Rever. Source: finance.informe_vat_gestorias_detalle. '
               'Update "Paid" and "Payment date" as payments arrive. '
               'Green = collected · Yellow = pending.'))
    c.font = ITALIC_GREY; c.alignment = LEFT_WRAP
    row += 1

    if not orders:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws.cell(row=row, column=1,
            value="✓ No bank transfer orders for this period.")
        c.font = Font(name="Calibri", size=10, italic=True, color="548235")
        c.alignment = LEFT
        return

    # Column headers
    ws.row_dimensions[row].height = 15
    for col, (header, _, _, _) in enumerate(B2B_COLS, 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = BOLD; c.fill = COL_HDR_FILL; c.alignment = CENTER; c.border = THIN
    row += 1
    first_data = row

    # Data rows
    ep = existing_payments or {}
    for o in orders:
        ws.row_dimensions[row].height = 14
        # Restore manually entered payment data from prior version (if any)
        pagado, fecha_cobro = ep.get(o.order_name, ("Pending", None))
        # Backwards-compat: map old Spanish values to English
        if pagado in ("Sí", "Si"):   pagado = "Yes"
        if pagado == "Pendiente":    pagado = "Pending"
        row_fill = YES_FILL if pagado == "Yes" else PEND_FILL

        vals: list[tuple[Any, Alignment, str | None, Font | None]] = [
            (o.order_name,              LEFT,   None,      NORMAL),
            (o.order_date,              CENTER, None,      NORMAL),
            (o.shipping_country_code,   CENTER, None,      NORMAL),
            (o.category,                CENTER, None,      NORMAL),
            (float(o.vat_pct),          CENTER, "0%",      NORMAL),
            (float(o.base),             RIGHT,  MONEY_FMT, NORMAL),
            (float(o.vat),              RIGHT,  MONEY_FMT, NORMAL),
            (float(o.total),            RIGHT,  MONEY_FMT, BOLD),
            (None,                      LEFT,   None,      LINK_FONT),  # link placeholder
            (pagado,                    CENTER, None,      BOLD),        # Pagado
            (fecha_cobro,               CENTER, None,      NORMAL),      # Fecha cobro
            ("",                        LEFT,   None,      NORMAL),      # Comentarios
        ]
        for col, (value, align, fmt, font) in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=value)
            c.alignment = align; c.border = THIN; c.fill = row_fill
            c.font = font
            if fmt: c.number_format = fmt

        if o.shopify_url:
            lc = ws.cell(row=row, column=9)
            lc.value = "Ver pedido"; lc.hyperlink = o.shopify_url; lc.font = LINK_FONT

        row += 1

    last_data = row - 1

    # Totals
    ws.row_dimensions[row].height = 15
    _D0       = Decimal("0")
    tot_base  = sum((o.base  for o in orders), _D0).quantize(Decimal("0.01"))
    tot_vat   = sum((o.vat   for o in orders), _D0).quantize(Decimal("0.01"))
    tot_total = sum((o.total for o in orders), _D0).quantize(Decimal("0.01"))
    # Recompute paid/pending from existing_payments for accurate summary
    tot_paid  = sum(
        (o.total for o in orders if ep.get(o.order_name, ("Pending", None))[0] == "Yes"),
        _D0,
    ).quantize(Decimal("0.01"))
    tot_pend  = (tot_total - tot_paid).quantize(Decimal("0.01"))
    tot_row: list[tuple[Any, Alignment, str | None]] = [
        (f"TOTAL  ({len(orders)} orders)", LEFT, None),
        (None, CENTER, None), (None, CENTER, None), (None, CENTER, None), (None, CENTER, None),
        (float(tot_base),  RIGHT, MONEY_FMT),
        (float(tot_vat),   RIGHT, MONEY_FMT),
        (float(tot_total), RIGHT, MONEY_FMT),
        (None, LEFT, None), (None, CENTER, None), (None, CENTER, None), (None, LEFT, None),
    ]
    for col, (val, align, fmt) in enumerate(tot_row, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = BOLD; c.fill = TOTAL_FILL; c.alignment = align; c.border = THIN
        if fmt: c.number_format = fmt
    row += 2

    # Collected / Pending summary
    for label, amount, fill in [
        ("Collected",          tot_paid,  YES_FILL),
        ("Pending collection", tot_pend,  PEND_FILL),
    ]:
        ws.row_dimensions[row].height = 14
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = BOLD; c.fill = fill; c.alignment = LEFT; c.border = THIN
        c2 = ws.cell(row=row, column=6, value=float(amount))
        c2.font = BOLD; c2.fill = fill; c2.alignment = RIGHT
        c2.border = THIN; c2.number_format = MONEY_FMT
        pct = float(amount / tot_total) if tot_total else 0.0
        c3 = ws.cell(row=row, column=7, value=pct)
        c3.font = BOLD; c3.fill = fill; c3.alignment = CENTER
        c3.border = THIN; c3.number_format = "0.0%"
        for col in range(8, ncols + 1):
            cx = ws.cell(row=row, column=col)
            cx.fill = fill; cx.border = THIN
        row += 1

    # Data validation for Paid (col 10)
    dv = DataValidation(
        type="list",
        formula1='"Yes,No,Pending"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.sqref = f"J{first_data}:J{last_data}"
    ws.add_data_validation(dv)

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Gift card inventory sheet
# ---------------------------------------------------------------------------

GC_COLS = [
    ("Date issued",      12, CENTER, None),
    ("Code (last 4)",    12, CENTER, None),
    ("Currency",          8, CENTER, None),
    ("Order #",          14, LEFT,   None),
    ("Initial value",    14, RIGHT,  MONEY_FMT),
    ("Amount used",      14, RIGHT,  MONEY_FMT),
    ("Balance remaining",16, RIGHT,  MONEY_FMT),
    ("% used",           10, CENTER, "0.0%"),
    ("Expiry date",      12, CENTER, None),
    ("Shopify link",     14, LEFT,   None),
]


def _add_gift_card_sheet(
    wb: Workbook,
    report: ReconciliationReport,
    month_label: str,
) -> None:
    """Sheet showing gift cards issued in the last 12 months with usage stats.

    Data comes from the Shopify Admin API (gift_cards endpoint).
    If the Shopify API was not configured or returned no data, shows a notice.
    """
    ws = wb.create_sheet("Gift Cards")

    ncols    = len(GC_COLS)
    last_col = get_column_letter(ncols)

    for i, (_, w, _, _) in enumerate(GC_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1,
        value=(f"Gift card inventory — {report.company_code}"
               f"   |   Last 12 months   |   Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"))
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL; c.alignment = CENTER
    row += 1

    inventory = report.gift_card_inventory

    # Subtitle
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1,
        value=("Gift cards issued in the last 12 months. "
               "Source: Shopify Admin API (/admin/api/gift_cards). "
               "Initial value, current balance and amount used are fetched live. "
               "% used = amount used / initial value."))
    c.font = ITALIC_GREY; c.alignment = LEFT_WRAP
    row += 1

    if not inventory:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws.cell(row=row, column=1,
            value=("✓ No gift cards found in the last 12 months, "
                   "or Shopify API is not configured "
                   "(set SHOPIFY_SHOP / SHOPIFY_CLIENT_ID / SHOPIFY_CLIENT_SECRET)."))
        c.font = Font(name="Calibri", size=10, italic=True, color="595959")
        c.alignment = LEFT
        return

    # Column headers
    ws.row_dimensions[row].height = 15
    for col, (header, _, _, _) in enumerate(GC_COLS, 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = BOLD; c.fill = COL_HDR_FILL; c.alignment = CENTER; c.border = THIN
    row += 1

    _D0 = Decimal("0")

    for gc in inventory:
        ws.row_dimensions[row].height = ROW_H

        # Row fill: fully-used = green, partially used = yellow, unused = grey
        if gc.balance == _D0:
            row_fill = CB_WON_FILL   # green — fully spent
        elif gc.amount_used > _D0:
            row_fill = WARN_FILL     # yellow — partially spent
        else:
            row_fill = LEGEND_FILL   # grey — not yet spent

        vals: list[tuple[Any, Alignment, str | None, Font | None]] = [
            (gc.created_at,              CENTER, None,      None),
            (gc.code_last4,              CENTER, None,      BOLD),
            (gc.currency,                CENTER, None,      None),
            (gc.order_name or "—",       LEFT,   None,      None),
            (float(gc.initial_value),    RIGHT,  MONEY_FMT, None),
            (float(gc.amount_used),      RIGHT,  MONEY_FMT, BOLD if gc.amount_used > _D0 else None),
            (float(gc.balance),          RIGHT,  MONEY_FMT, None),
            (float(gc.pct_used),         CENTER, "0.0%",    BOLD),
            (gc.expires_on or "—",       CENTER, None,      ITALIC_GREY if not gc.expires_on else None),
            (None,                       LEFT,   None,      LINK_FONT),   # Shopify link placeholder
        ]
        for col, (value, align, fmt, font) in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=value)
            c.alignment = align; c.border = THIN; c.fill = row_fill
            c.font = font if font else NORMAL
            if fmt: c.number_format = fmt

        if gc.shopify_url:
            lc = ws.cell(row=row, column=10)
            lc.value = "View order"; lc.hyperlink = gc.shopify_url; lc.font = LINK_FONT

        row += 1

    # Totals row
    ws.row_dimensions[row].height = ROW_H
    tot_initial = sum((gc.initial_value for gc in inventory), _D0).quantize(Decimal("0.01"))
    tot_used    = sum((gc.amount_used   for gc in inventory), _D0).quantize(Decimal("0.01"))
    tot_balance = sum((gc.balance       for gc in inventory), _D0).quantize(Decimal("0.01"))
    tot_pct     = float(tot_used / tot_initial) if tot_initial != _D0 else 0.0

    tot_vals: list[tuple[Any, Alignment, str | None]] = [
        (f"TOTAL  ({len(inventory)} gift cards)", LEFT, None),
        (None, CENTER, None),
        (None, CENTER, None),
        (None, LEFT,   None),
        (float(tot_initial), RIGHT, MONEY_FMT),
        (float(tot_used),    RIGHT, MONEY_FMT),
        (float(tot_balance), RIGHT, MONEY_FMT),
        (tot_pct,            CENTER, "0.0%"),
        (None, CENTER, None),
        (None, LEFT,   None),
    ]
    for col, (val, align, fmt) in enumerate(tot_vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = BOLD; c.fill = TOTAL_FILL; c.alignment = align; c.border = THIN
        if fmt: c.number_format = fmt
    row += 2

    # Legend
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1,
        value=("Green = fully spent (balance = 0)   "
               "Yellow = partially used   "
               "Grey = not yet used"))
    c.font = ITALIC_GREY; c.fill = LEGEND_FILL; c.alignment = LEFT

    ws.freeze_panes = "A4"
