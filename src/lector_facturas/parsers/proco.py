from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import calendar
import re

from openpyxl import load_workbook
from pypdf import PdfReader


COMPANY_NAME = "ARTESTA STORES (UK) LTD"
ISSUER_COMPANY_NAME = "PRECISION PRINTING CO. LTD"
SUPPLIER_CODE = "PROCO"
TWOPLACES = Decimal("0.01")

MONTHS_EN = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


@dataclass(frozen=True)
class ProcoInvoice:
    supplier_code: str
    supplier_name: str
    issuer_company_name: str
    billed_company_name: str
    invoice_number: str
    invoice_date: date
    billing_period_start: date
    billing_period_end: date
    period_yyyymm: str
    currency_code: str
    vat_percent: Decimal
    gross_amount: Decimal
    vat_amount: Decimal
    net_amount: Decimal
    manufacturing_net_amount: Decimal
    manufacturing_vat_amount: Decimal
    manufacturing_gross_amount: Decimal
    logistics_net_amount: Decimal
    logistics_vat_amount: Decimal
    logistics_gross_amount: Decimal
    original_filename: str
    invoice_line_items: dict[str, str]
    detail_summary: dict[str, object]
    payment_due_date: date | None = None
    parser_name: str = "proco"
    parser_confidence: Decimal = Decimal("0.9950")

    @property
    def extracted_raw(self) -> dict[str, object]:
        return {
            "issuer_company_name": self.issuer_company_name,
            "billed_company_name": self.billed_company_name,
            "billing_period_start": self.billing_period_start.isoformat(),
            "billing_period_end": self.billing_period_end.isoformat(),
            "period_yyyymm": self.period_yyyymm,
            "currency_code": self.currency_code,
            "vat_percent": format(self.vat_percent, "f"),
            "gross_amount": format(self.gross_amount, "f"),
            "vat_amount": format(self.vat_amount, "f"),
            "net_amount": format(self.net_amount, "f"),
            "manufacturing_net_amount": format(self.manufacturing_net_amount, "f"),
            "manufacturing_vat_amount": format(self.manufacturing_vat_amount, "f"),
            "manufacturing_gross_amount": format(self.manufacturing_gross_amount, "f"),
            "logistics_net_amount": format(self.logistics_net_amount, "f"),
            "logistics_vat_amount": format(self.logistics_vat_amount, "f"),
            "logistics_gross_amount": format(self.logistics_gross_amount, "f"),
            "invoice_line_items": self.invoice_line_items,
            "detail_summary": self.detail_summary,
            "payment_due_date": self.payment_due_date.isoformat() if self.payment_due_date else None,
        }


def parse_proco_bundle(pdf_path: Path, detail_path: Path) -> ProcoInvoice:
    text = "\n".join((page.extract_text() or "") for page in PdfReader(str(pdf_path)).pages)
    summary = parse_proco_detail_summary(detail_path)
    return parse_proco_text_and_summary(text, detail_summary=summary, original_filename=pdf_path.name)


def parse_proco_pdf(pdf_path: Path) -> ProcoInvoice:
    text = "\n".join((page.extract_text() or "") for page in PdfReader(str(pdf_path)).pages)
    return parse_proco_text(text, original_filename=pdf_path.name)


def parse_proco_text(text: str, *, original_filename: str) -> ProcoInvoice:
    normalized = text.replace("\xa0", " ").replace("\r", "")
    invoice_number = _extract_invoice_number(normalized)
    invoice_date = _extract_invoice_date(normalized)
    vat_percent, net_amount, vat_amount, gross_amount = _extract_totals(normalized)
    invoice_lines = _extract_invoice_lines(normalized)
    logistics_net_amount = _quantize(
        _parse_optional_amount(invoice_lines.get("carriage"))
        + _parse_optional_amount(invoice_lines.get("postage"))
    )
    manufacturing_net_amount = _quantize(net_amount - logistics_net_amount)
    manufacturing_vat_amount = _quantize(manufacturing_net_amount * vat_percent / Decimal("100"))
    logistics_vat_amount = _quantize(logistics_net_amount * vat_percent / Decimal("100"))
    manufacturing_gross_amount = _quantize(manufacturing_net_amount + manufacturing_vat_amount)
    logistics_gross_amount = _quantize(logistics_net_amount + logistics_vat_amount)
    billing_period_start = date(invoice_date.year, invoice_date.month, 1)
    billing_period_end = date(invoice_date.year, invoice_date.month, calendar.monthrange(invoice_date.year, invoice_date.month)[1])
    payment_due_date = _extract_payment_due_date(normalized, billing_period_end)

    return ProcoInvoice(
        supplier_code=SUPPLIER_CODE,
        supplier_name=SUPPLIER_CODE,
        issuer_company_name=ISSUER_COMPANY_NAME,
        billed_company_name=COMPANY_NAME,
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        billing_period_start=billing_period_start,
        billing_period_end=billing_period_end,
        period_yyyymm=_period_with_most_days(billing_period_start, billing_period_end),
        currency_code="GBP",
        vat_percent=vat_percent,
        net_amount=net_amount,
        vat_amount=vat_amount,
        gross_amount=gross_amount,
        manufacturing_net_amount=manufacturing_net_amount,
        manufacturing_vat_amount=manufacturing_vat_amount,
        manufacturing_gross_amount=manufacturing_gross_amount,
        logistics_net_amount=logistics_net_amount,
        logistics_vat_amount=logistics_vat_amount,
        logistics_gross_amount=logistics_gross_amount,
        original_filename=original_filename,
        invoice_line_items=invoice_lines,
        detail_summary={"source": "pdf_only"},
        payment_due_date=payment_due_date,
    )


def parse_proco_text_and_summary(
    text: str,
    *,
    detail_summary: dict[str, object],
    original_filename: str,
) -> ProcoInvoice:
    normalized = text.replace("\xa0", " ").replace("\r", "")
    invoice_number = _extract_invoice_number(normalized)
    invoice_date = _extract_invoice_date(normalized)
    vat_percent, net_amount, vat_amount, gross_amount = _extract_totals(normalized)
    invoice_lines = _extract_invoice_lines(normalized)
    billing_period_start, billing_period_end = _extract_period_range(detail_summary)
    logistics_net_amount = _decimal(detail_summary["shipments_total"])
    manufacturing_net_amount = _quantize(net_amount - logistics_net_amount)
    manufacturing_vat_amount = _quantize(manufacturing_net_amount * vat_percent / Decimal("100"))
    logistics_vat_amount = _quantize(logistics_net_amount * vat_percent / Decimal("100"))
    manufacturing_gross_amount = _quantize(manufacturing_net_amount + manufacturing_vat_amount)
    logistics_gross_amount = _quantize(logistics_net_amount + logistics_vat_amount)

    if _quantize(manufacturing_net_amount + logistics_net_amount) != _quantize(net_amount):
        raise ValueError("PROCO manufacturing/logistics split does not match invoice net amount.")

    payment_due_date = _extract_payment_due_date(normalized, billing_period_end)

    return ProcoInvoice(
        supplier_code=SUPPLIER_CODE,
        supplier_name=SUPPLIER_CODE,
        issuer_company_name=ISSUER_COMPANY_NAME,
        billed_company_name=COMPANY_NAME,
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        billing_period_start=billing_period_start,
        billing_period_end=billing_period_end,
        period_yyyymm=_period_with_most_days(billing_period_start, billing_period_end),
        currency_code="GBP",
        vat_percent=vat_percent,
        net_amount=net_amount,
        vat_amount=vat_amount,
        gross_amount=gross_amount,
        manufacturing_net_amount=manufacturing_net_amount,
        manufacturing_vat_amount=manufacturing_vat_amount,
        manufacturing_gross_amount=manufacturing_gross_amount,
        logistics_net_amount=logistics_net_amount,
        logistics_vat_amount=logistics_vat_amount,
        logistics_gross_amount=logistics_gross_amount,
        original_filename=original_filename,
        invoice_line_items=invoice_lines,
        detail_summary=_stringify_summary(detail_summary),
        payment_due_date=payment_due_date,
    )


def parse_proco_detail_summary(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    title = str(worksheet.cell(1, 1).value or "")
    month_name, year = _extract_month_year(title)
    last_day = calendar.monthrange(year, month_name)[1]

    summary: dict[str, Decimal] = {
        "storage_fee_total": Decimal("0.00"),
        "shipments_total": Decimal("0.00"),
        "posters_total": Decimal("0.00"),
        "frames_total": Decimal("0.00"),
        "passpartout_total": Decimal("0.00"),
        "pick_pack_passpartout_total": Decimal("0.00"),
        "pick_pack_material_total": Decimal("0.00"),
        "detail_total": Decimal("0.00"),
    }

    for row in worksheet.iter_rows(min_row=1, max_row=20, values_only=True):
        label = _extract_label(row)
        if not label:
            continue
        total = _extract_last_decimal(row)
        if total is None:
            continue
        normalized_label = _normalize_label(label)
        if normalized_label.startswith("storage fee"):
            summary["storage_fee_total"] = total
        elif normalized_label == "shipments":
            summary["shipments_total"] = total
        elif normalized_label == "posters":
            summary["posters_total"] = total
        elif normalized_label == "frames":
            summary["frames_total"] = total
        elif normalized_label == "passpertout":
            summary["passpartout_total"] = total
        elif normalized_label == "pick/pack - passpartout":
            summary["pick_pack_passpartout_total"] = total
        elif normalized_label == "pick/pack & packing material":
            summary["pick_pack_material_total"] = total

    for row in worksheet.iter_rows(min_row=1, max_row=20, values_only=True):
        values = [value for value in row if value not in (None, "")]
        if len(values) == 1 and isinstance(values[0], (int, float)):
            summary["detail_total"] = _decimal(values[0])
            break

    summary["manufacturing_total"] = _quantize(
        summary["storage_fee_total"]
        + summary["posters_total"]
        + summary["frames_total"]
        + summary["passpartout_total"]
        + summary["pick_pack_passpartout_total"]
        + summary["pick_pack_material_total"]
    )
    summary["period_start"] = date(year, month_name, 1)
    summary["period_end"] = date(year, month_name, last_day)
    return summary


def _extract_invoice_number(text: str) -> str:
    match = re.search(r"\n(\d{7})\nAccounts", text)
    if not match:
        raise ValueError("Could not extract PROCO invoice number.")
    return match.group(1)


def _extract_invoice_date(text: str) -> date:
    match = re.search(r"(\d{2}/\d{2}/\d{4})\n\d{7}\nAccounts", text)
    if not match:
        raise ValueError("Could not extract PROCO invoice date.")
    return datetime.strptime(match.group(1), "%d/%m/%Y").date()


def _extract_totals(text: str) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    match = re.search(
        r"01\s+\D*([\d,]+\.\d{2})\s+\D*([\d,]+\.\d{2})\s+20\.00",
        text,
        flags=re.DOTALL,
    )
    if not match:
        raise ValueError("Could not extract PROCO totals.")
    net_amount = _parse_currency(match.group(1))
    vat_amount = _parse_currency(match.group(2))
    gross_amount = _quantize(net_amount + vat_amount)
    return Decimal("20.00"), net_amount, vat_amount, gross_amount


def _extract_invoice_lines(text: str) -> dict[str, str]:
    descriptions = ("Print", "Carriage", "Postage", "Direct Mailing", "Storage")
    result: dict[str, str] = {}
    for description in descriptions:
        match = re.search(rf"\D([\d,]+\.\d{{2}})\s+1 01{re.escape(description)}", text)
        if match:
            result[description.lower().replace(" ", "_")] = format(_parse_currency(match.group(1)), "f")
    return result


def _extract_period_range(detail_summary: dict[str, object]) -> tuple[date, date]:
    return detail_summary["period_start"], detail_summary["period_end"]


def _extract_month_year(title: str) -> tuple[int, int]:
    match = re.search(r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})", title, flags=re.IGNORECASE)
    if not match:
        raise ValueError("Could not extract PROCO billing period from detail workbook.")
    month = MONTHS_EN[match.group(1).lower()]
    year = int(match.group(2))
    return month, year


def _extract_label(row: tuple[object, ...]) -> str:
    for value in row:
        if isinstance(value, str):
            candidate = value.strip()
            if not candidate or candidate == "-":
                continue
            if re.search(r"[A-Za-z]", candidate):
                return candidate
    return ""


def _extract_last_decimal(row: tuple[object, ...]) -> Decimal | None:
    numbers: list[Decimal] = []
    for value in row:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            numbers.append(_decimal(value))
        elif isinstance(value, str):
            cleaned = value.strip().replace("Ł", "").replace(",", "")
            if re.fullmatch(r"-?\d+(?:\.\d+)?", cleaned):
                numbers.append(_decimal(cleaned))
    return numbers[-1] if numbers else None


def _normalize_label(value: str) -> str:
    return " ".join(value.strip().lower().split())


def _period_with_most_days(start: date, end: date) -> str:
    counts: dict[str, int] = {}
    current = start
    while current <= end:
        key = current.strftime("%Y%m")
        counts[key] = counts.get(key, 0) + 1
        current = date.fromordinal(current.toordinal() + 1)
    return max(counts.items(), key=lambda item: (item[1], item[0]))[0]


def _parse_currency(raw_value: str) -> Decimal:
    return Decimal(raw_value.replace(",", "")).quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def _parse_optional_amount(raw_value: str | None) -> Decimal:
    if not raw_value:
        return Decimal("0.00")
    return _parse_currency(raw_value)


def _decimal(value: object) -> Decimal:
    return Decimal(str(value)).quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def _quantize(value: Decimal) -> Decimal:
    return value.quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def _extract_payment_due_date(text: str, billing_period_end: date) -> date | None:
    """Extract or calculate payment due date from PROCO invoice text.

    PROCO standard terms: 'TERMS 30 DAYS FROM END OF MONTH'
    → due_date = end_of_month(billing_period_end) + 30 days.
    """
    from datetime import timedelta
    normalized_upper = text.upper()
    # Standard PROCO terms: 30 DAYS FROM END OF MONTH
    match = re.search(r"TERMS\s+(\d+)\s+DAYS\s+FROM\s+END\s+OF\s+MONTH", normalized_upper)
    if match:
        days = int(match.group(1))
        last_day = calendar.monthrange(billing_period_end.year, billing_period_end.month)[1]
        end_of_month = date(billing_period_end.year, billing_period_end.month, last_day)
        return end_of_month + timedelta(days=days)
    # Fallback: look for explicit DUE DATE dd/mm/yyyy or similar
    match = re.search(r"DUE\s+DATE[:\s]+(\d{2}/\d{2}/\d{4})", normalized_upper)
    if match:
        from datetime import datetime as _dt
        return _dt.strptime(match.group(1), "%d/%m/%Y").date()
    return None


def _stringify_summary(summary: dict[str, object]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in summary.items():
        if isinstance(value, Decimal):
            result[key] = format(value, "f")
        elif isinstance(value, date):
            result[key] = value.isoformat()
        else:
            result[key] = value
    return result
