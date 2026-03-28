"""Build the shopify_sales gestoría Excel workbook.

Generates a two-sheet .xlsx for a single company + month that is
uploaded to Google Drive by ``pyg_sync.sync_gestoria_to_drive()``.

Output file name: ``shopify_sales_{company_code}_{yyyymm}.xlsx``
Drive path:       ``{root}/{entity}/{year}/{yyyymm}/income/sales/shopify/``

Sheets
------
1. **Summary** — aggregated view by country × VAT rate.
   Rows are grouped by country (or US state for INC) with section headers,
   per-rate data rows, country subtotals and a grand total.
   Cells in WARN_FILL (yellow) indicate a discrepancy between the Shopify
   VAT rate and the theoretical rate (only flagged when both > 0 and
   |diff| > 0.001 — zero-rated orders are not flagged).
   For INC a ``Shopify fee`` column is added with fees per state sourced
   from ``invoices.shopify_payout_transactions``.  The grand-total fee
   uses ``invoices.payment_fee_monthly_summary.total_cost_amount`` so it
   matches the "SHOPIFY" line in the INC P&G workbook exactly.

2. **Detail** — one row per order.
   Columns: Date · Order# · Country · State · VAT rate ×3 · Gross ·
   VAT/Tax · Net · [Shopify fee — INC only] · Discrepancy ·
   Payment gateways · Hannun · Rever.
   Rows with |discrepancy| > 0.01 are highlighted in yellow.

Data sources
------------
``finance.informe_vat_gestorias_resumen_{yyyymm}``
    Partitioned table (one per month).  Aggregated by country × VAT rate ×
    payment method.  Filtered by ``payment_currency`` and
    ``is_hannun_tag = 0``.

``finance.informe_vat_gestorias_detalle``
    Single table with one row per order.  Filtered by
    ``order_month_yyyymm`` and ``payment_currency``.

``invoices.shopify_payout_transactions`` (INC only)
    Per-order Shopify fee (``SUM(fee)`` for ``type = 'charge'``).

``invoices.payment_fee_monthly_summary`` (INC only)
    Monthly fee total for the grand-total row; matches the P&G source.

Company → currency → region
---------------------------
  SL  → EUR, EU  (all countries except GB and US)
  LTD → GBP, UK  (country = GB)
  INC → USD, US  (country = US; grouped by shipping_state_code in Summary)

Public API
----------
``collect_gestoria_data(*, database_url, company_code, period_yyyymm)``
    Connects to the DB and returns a ``GestoriaReportData`` instance with
    all data needed to build the workbook.  Performs the fee queries for
    INC automatically.

``build_gestoria_workbook(data: GestoriaReportData) -> bytes``
    Pure function — takes the data bundle and returns the .xlsx as bytes.
    No DB or Drive access.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, UTC
from decimal import Decimal
from io import BytesIO
from itertools import groupby
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:  # pragma: no cover
    psycopg = None  # type: ignore[assignment]
    dict_row = None  # type: ignore[assignment]

# ---------------------------------------------------------------------------
# Styles (mirrors payment_reconciliation_workbook.py)
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")   # dark blue   – title row
SECTION_FILL = PatternFill("solid", fgColor="2E74B5")   # mid blue    – section / country header
COL_HDR_FILL = PatternFill("solid", fgColor="D6E4F0")   # light blue  – column headers
WARN_FILL    = PatternFill("solid", fgColor="FFF2CC")   # yellow      – diff row highlight
TOTAL_FILL   = PatternFill("solid", fgColor="BDD7EE")   # blue-grey   – totals row

WHITE_BOLD  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
BOLD        = Font(bold=True, name="Calibri", size=10)
NORMAL      = Font(name="Calibri", size=10)
ITALIC_GREY = Font(name="Calibri", size=9, italic=True, color="595959")
THIN_SIDE   = Side(style="thin", color="BFBFBF")
THIN        = Border(top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=False)
LEFT_WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
RIGHT       = Alignment(horizontal="right",  vertical="center", wrap_text=False)
MONEY_FMT   = '#,##0.00;[Red](#,##0.00);-'
ROW_H       = 14

MONTH_NAMES_EN = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# ---------------------------------------------------------------------------
# Company mappings
# ---------------------------------------------------------------------------

COMPANY_CURRENCY: dict[str, str] = {
    "SL":  "EUR",
    "LTD": "GBP",
    "INC": "USD",
}

COMPANY_REGION: dict[str, str] = {
    "SL":  "EU",
    "LTD": "UK",
    "INC": "US",
}

# ---------------------------------------------------------------------------
# Public data classes
# ---------------------------------------------------------------------------

@dataclass
class GestoriaReportData:
    period_yyyymm: str
    company_code: str
    currency: str
    region: str          # "EU", "UK", "US"
    resumen_rows: list[dict]   # raw rows from DB (filtered by currency + is_hannun_tag=0)
    detalle_rows: list[dict]   # raw rows from DB (filtered by currency)
    fees_by_order: dict[str, Decimal]   # order_name → Shopify payout fee (INC only, else {})
    monthly_total_fee: Decimal          # SUM(total_cost_amount) from payment_fee_monthly_summary (INC only, else 0)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def collect_gestoria_data(
    *,
    database_url: str,
    company_code: str,
    period_yyyymm: str,
) -> GestoriaReportData:
    """Query DB and return all data needed to build the workbook."""
    if psycopg is None:
        raise RuntimeError("psycopg is not installed.")

    if not period_yyyymm.isdigit() or len(period_yyyymm) != 6:
        raise ValueError(
            f"period_yyyymm must be exactly 6 digits (e.g. '202602'), got: {period_yyyymm!r}"
        )

    currency = COMPANY_CURRENCY.get(company_code.upper())
    if not currency:
        raise ValueError(
            f"Unknown company_code '{company_code}'. Expected one of: {list(COMPANY_CURRENCY)}"
        )
    region = COMPANY_REGION[company_code.upper()]

    resumen_table = f"finance.informe_vat_gestorias_resumen_{period_yyyymm}"

    with psycopg.connect(database_url, row_factory=dict_row) as conn:
        resumen_rows = conn.execute(
            f"""
            SELECT *
            FROM {resumen_table}
            WHERE payment_currency = %s
              AND is_hannun_tag = 0
            ORDER BY country, tax_rate_teorical
            """,
            (currency,),
        ).fetchall()

        detalle_rows = conn.execute(
            """
            SELECT *
            FROM finance.informe_vat_gestorias_detalle
            WHERE order_month_yyyymm = %s
              AND payment_currency   = %s
            ORDER BY shipping_country_code, order_date, order_name
            """,
            (period_yyyymm, currency),
        ).fetchall()

        fees_by_order: dict[str, Decimal] = {}
        monthly_total_fee: Decimal = Decimal("0")
        if region == "US" and detalle_rows:
            order_names = [r["order_name"] for r in detalle_rows if r.get("order_name")]
            if order_names:
                fee_rows = conn.execute(
                    """
                    SELECT order_name, SUM(fee) AS total_fee
                    FROM invoices.shopify_payout_transactions
                    WHERE company_code = %s
                      AND type         = 'charge'
                      AND order_name   = ANY(%s)
                    GROUP BY order_name
                    """,
                    (company_code.upper(), order_names),
                ).fetchall()
                fees_by_order = {
                    r["order_name"]: _dec(r["total_fee"])
                    for r in fee_rows
                    if r.get("order_name")
                }

            # Monthly total from payment_fee_monthly_summary — same source as PYG INC
            fee_summary = conn.execute(
                """
                SELECT COALESCE(SUM(total_cost_amount), 0) AS total
                FROM invoices.payment_fee_monthly_summary
                WHERE company_code = %s
                  AND period_yyyymm = %s
                """,
                (company_code.upper(), period_yyyymm),
            ).fetchone()
            if fee_summary:
                monthly_total_fee = _dec(fee_summary["total"])

    return GestoriaReportData(
        period_yyyymm=period_yyyymm,
        company_code=company_code.upper(),
        currency=currency,
        region=region,
        resumen_rows=[dict(r) for r in resumen_rows],
        detalle_rows=[dict(r) for r in detalle_rows],
        fees_by_order=fees_by_order,
        monthly_total_fee=monthly_total_fee,
    )


def build_gestoria_workbook(data: GestoriaReportData) -> bytes:
    """Return the xlsx as bytes."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    year        = int(data.period_yyyymm[:4])
    month       = int(data.period_yyyymm[4:])
    month_label = f"{MONTH_NAMES_EN[month - 1]} {year}"

    _add_summary_sheet(wb, data, month_label)
    _add_detail_sheet(wb, data, month_label)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _dec(value: Any) -> Decimal:
    """Convert a DB value to Decimal, returning 0 for None / empty."""
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


def _sum_dec(rows: list[dict], field: str) -> Decimal:
    return sum((_dec(r[field]) for r in rows), Decimal("0")).quantize(Decimal("0.01"))


def _vat_status(tax_rate_shopify: Any, tax_rate_teorical: Any) -> str:
    """Return '✓' or '⚠ diff'.

    Only flag as discrepancy when both rates are > 0 and differ by more than 0.001.
    If the Shopify rate is 0 while the theoretical is > 0 it is likely a free-shipping
    or exempt order — not a real mismatch.
    """
    s = float(tax_rate_shopify or 0)
    t = float(tax_rate_teorical or 0)
    if s > 0 and t > 0 and abs(s - t) > 0.001:
        return "⚠ diff"
    return "✓"


def _country_key(row: dict, region: str) -> str:
    """Return the grouping key for a resumen row (state code for US, country otherwise)."""
    if region == "US":
        return str(row.get("shipping_state_code") or row.get("country") or "")
    return str(row.get("country") or "")


def _make_cell_writer(ws: Any, last_col: str, ncols: int):
    """Return (_sh, _cell, _blank_row) helper closures bound to *ws*."""

    def _sh(r: int, h: int = ROW_H) -> None:
        ws.row_dimensions[r].height = h

    def _cell(
        r: int,
        col: int,
        value: Any,
        *,
        font: Font | None = None,
        fill: PatternFill | None = None,
        align: Alignment = LEFT,
        fmt: str | None = None,
        border: Border | None = THIN,
    ) -> None:
        c = ws.cell(row=r, column=col, value=value)
        c.font      = font  or NORMAL
        c.alignment = align
        if fill:   c.fill         = fill
        if border: c.border       = border
        if fmt:    c.number_format = fmt

    def _blank_row(r: int) -> None:
        _sh(r, 6)
        ws.merge_cells(f"A{r}:{last_col}{r}")
        ws.cell(row=r, column=1)

    return _sh, _cell, _blank_row


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _add_summary_sheet(wb: Workbook, data: GestoriaReportData, month_label: str) -> None:
    ws = wb.create_sheet("Summary")

    has_fees = data.region == "US"

    # Column widths  A  B   C   D   E   F    G    H    I     J(INC only)
    base_widths = [8, 13, 13, 13, 10, 15, 15, 15, 10]
    col_widths  = base_widths[:8] + [13, 10] if has_fees else base_widths
    #              A   B   C   D   E   F   G   H  fee  status
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    NCOLS    = len(col_widths)
    LAST_COL = get_column_letter(NCOLS)

    # column numbers (1-based)
    COL_FEE    = 9 if has_fees else None   # new fee column (INC only)
    COL_STATUS = 10 if has_fees else 9

    _sh, _cell, _blank_row = _make_cell_writer(ws, LAST_COL, NCOLS)

    generated_ts = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    period       = data.period_yyyymm

    # Pre-compute per-country-key fee totals from fees_by_order + detalle_rows
    from collections import defaultdict as _defaultdict
    fees_by_ck: dict[str, Decimal] = _defaultdict(Decimal)
    if has_fees:
        ck_for_order: dict[str, str] = {}
        for dr in data.detalle_rows:
            oname = dr.get("order_name")
            if oname:
                ck_for_order[str(oname)] = str(
                    dr.get("shipping_state_code") or dr.get("shipping_country_code") or ""
                )
        for oname, fee in data.fees_by_order.items():
            ck = ck_for_order.get(str(oname), "")
            fees_by_ck[ck] += fee

    row = 1

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    _sh(row, 22)
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    _cell(
        row, 1,
        f"VAT Report — {data.company_code} — {data.region} — {month_label}"
        f"  |  Generated: {generated_ts}",
        font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
        fill=HEADER_FILL,
        align=CENTER,
    )
    row += 1

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    _sh(row, 28)
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    _cell(
        row, 1,
        f"Breakdown by country and VAT rate. "
        f"Source: finance.informe_vat_gestorias_resumen_{period}. "
        f"Aggregated by country × VAT rate. "
        f"⚠ = Shopify rate differs from theoretical. Hannun orders excluded.",
        font=ITALIC_GREY,
        align=LEFT_WRAP,
        border=None,
    )
    row += 1

    # ── Row 3: blank ──────────────────────────────────────────────────────────
    _blank_row(row); row += 1

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    _sh(row)
    col_headers = [
        "Country / State",
        "VAT rate\ntheoretical",
        "VAT rate\nShopify",
        "VAT rate\ncalculated",
        "# Orders",
        "Gross",
        "VAT / Tax",
        "Net",
    ]
    if has_fees:
        col_headers.append("Shopify fee")
    col_headers.append("Status")
    for col, hdr in enumerate(col_headers, 1):
        _cell(row, col, hdr, font=BOLD, fill=COL_HDR_FILL, align=CENTER)
    row += 1

    # ── Data rows: aggregate resumen, group by country ─────────────────────────
    agg: dict[tuple, dict] = {}
    for r in data.resumen_rows:
        ck         = _country_key(r, data.region)
        teorical   = float(r.get("tax_rate_teorical")   or 0)
        shopify_r  = float(r.get("tax_rate_shopify")    or 0)
        calculated = float(r.get("tax_rate_calculated") or 0)
        key = (ck, teorical, shopify_r, calculated)
        if key not in agg:
            agg[key] = {
                "country_key":         ck,
                "tax_rate_teorical":   teorical,
                "tax_rate_shopify":    shopify_r,
                "tax_rate_calculated": calculated,
                "num_orders":          0,
                "imp_sales_gross":     Decimal("0"),
                "imp_sales_tax":       Decimal("0"),
                "imp_sales_net":       Decimal("0"),
            }
        agg[key]["num_orders"]      += int(r.get("num_orders") or 0)
        agg[key]["imp_sales_gross"] += _dec(r.get("imp_sales_gross"))
        agg[key]["imp_sales_tax"]   += _dec(r.get("imp_sales_tax"))
        agg[key]["imp_sales_net"]   += _dec(r.get("imp_sales_net"))

    agg_rows = sorted(agg.values(), key=lambda x: (x["country_key"], x["tax_rate_teorical"]))

    grand_orders = 0
    grand_gross  = Decimal("0")
    grand_tax    = Decimal("0")
    grand_net    = Decimal("0")
    grand_fee    = Decimal("0")

    for country_key, grp_iter in groupby(agg_rows, key=lambda x: x["country_key"]):
        grp      = list(grp_iter)
        n_groups = len(grp)

        # Country header row
        _sh(row, ROW_H + 2)
        ws.merge_cells(f"A{row}:{LAST_COL}{row}")
        _cell(
            row, 1,
            f"{country_key} — {n_groups} rate group{'s' if n_groups != 1 else ''}",
            font=WHITE_BOLD, fill=SECTION_FILL, align=LEFT,
        )
        row += 1

        c_orders = 0
        c_gross  = Decimal("0")
        c_tax    = Decimal("0")
        c_net    = Decimal("0")
        c_fee    = fees_by_ck.get(country_key, Decimal("0")).quantize(Decimal("0.01"))

        for ar in grp:
            status = _vat_status(ar["tax_rate_shopify"], ar["tax_rate_teorical"])
            fill   = WARN_FILL if status == "⚠ diff" else None

            gross = ar["imp_sales_gross"].quantize(Decimal("0.01"))
            tax   = ar["imp_sales_tax"].quantize(Decimal("0.01"))
            net   = ar["imp_sales_net"].quantize(Decimal("0.01"))

            _sh(row)
            _cell(row, 1, ar["country_key"],        font=NORMAL, fill=fill, align=LEFT)
            _cell(row, 2, ar["tax_rate_teorical"],   font=NORMAL, fill=fill, align=CENTER, fmt="0%")
            _cell(row, 3, ar["tax_rate_shopify"],    font=NORMAL, fill=fill, align=CENTER, fmt="0%")
            _cell(row, 4, ar["tax_rate_calculated"], font=NORMAL, fill=fill, align=CENTER, fmt="0%")
            _cell(row, 5, ar["num_orders"],          font=NORMAL, fill=fill, align=CENTER)
            _cell(row, 6, float(gross),              font=NORMAL, fill=fill, align=RIGHT, fmt=MONEY_FMT)
            _cell(row, 7, float(tax),                font=NORMAL, fill=fill, align=RIGHT, fmt=MONEY_FMT)
            _cell(row, 8, float(net),                font=NORMAL, fill=fill, align=RIGHT, fmt=MONEY_FMT)
            if has_fees:
                # Fee shown only on the first rate-group row for this state (others blank)
                _cell(row, COL_FEE, None, font=NORMAL, fill=fill, align=RIGHT, fmt=MONEY_FMT)
            _cell(row, COL_STATUS, status, font=NORMAL, fill=fill, align=CENTER)
            row += 1

            c_orders += ar["num_orders"]
            c_gross  += gross
            c_tax    += tax
            c_net    += net

        # Country subtotal row
        _sh(row, ROW_H + 1)
        ws.merge_cells(f"A{row}:D{row}")
        _cell(row, 1, f"  {country_key} subtotal", font=BOLD, fill=TOTAL_FILL, align=LEFT)
        _cell(row, 5, c_orders,
              font=BOLD, fill=TOTAL_FILL, align=CENTER)
        _cell(row, 6, float(c_gross.quantize(Decimal("0.01"))),
              font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 7, float(c_tax.quantize(Decimal("0.01"))),
              font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, 8, float(c_net.quantize(Decimal("0.01"))),
              font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
        if has_fees:
            _cell(row, COL_FEE, float(c_fee),
                  font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
        _cell(row, COL_STATUS, None, font=BOLD, fill=TOTAL_FILL, align=CENTER)
        row += 1

        grand_orders += c_orders
        grand_gross  += c_gross
        grand_tax    += c_tax
        grand_net    += c_net
        grand_fee    += c_fee

    # Blank row before grand total
    _blank_row(row); row += 1

    # Grand total row — fee uses monthly_total_fee (= PYG source) if available,
    # else falls back to sum of per-state fees
    grand_fee_total = (
        data.monthly_total_fee if data.monthly_total_fee > Decimal("0") else grand_fee
    ).quantize(Decimal("0.01"))

    _sh(row, ROW_H + 2)
    ws.merge_cells(f"A{row}:D{row}")
    _cell(row, 1, "GRAND TOTAL", font=WHITE_BOLD, fill=HEADER_FILL, align=LEFT)
    _cell(row, 5, grand_orders,
          font=WHITE_BOLD, fill=HEADER_FILL, align=CENTER)
    _cell(row, 6, float(grand_gross.quantize(Decimal("0.01"))),
          font=WHITE_BOLD, fill=HEADER_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row, 7, float(grand_tax.quantize(Decimal("0.01"))),
          font=WHITE_BOLD, fill=HEADER_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row, 8, float(grand_net.quantize(Decimal("0.01"))),
          font=WHITE_BOLD, fill=HEADER_FILL, align=RIGHT, fmt=MONEY_FMT)
    if has_fees:
        _cell(row, COL_FEE, float(grand_fee_total),
              font=WHITE_BOLD, fill=HEADER_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row, COL_STATUS, None, font=WHITE_BOLD, fill=HEADER_FILL, align=CENTER)


# ---------------------------------------------------------------------------
# Detail sheet
# ---------------------------------------------------------------------------

def _add_detail_sheet(wb: Workbook, data: GestoriaReportData, month_label: str) -> None:
    ws = wb.create_sheet("Detail")

    has_fees = data.region == "US"

    # Column widths — base 14 columns; add one fee column (12) for INC after Net
    base_widths = [12, 16, 7, 8, 10, 10, 10, 14, 14, 14, 12, 22, 8, 8]
    if has_fees:
        # Insert fee column (width 12) after Net (index 9, 0-based) → before Discrepancy
        col_widths = base_widths[:10] + [12] + base_widths[10:]
    else:
        col_widths = base_widths
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    LAST_COL = get_column_letter(len(col_widths))
    NCOLS    = len(col_widths)

    _sh, _cell, _blank_row = _make_cell_writer(ws, LAST_COL, NCOLS)

    generated_ts = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    period       = data.period_yyyymm
    rows         = data.detalle_rows

    # Pre-compute stats for subtitle row
    total_gross = _sum_dec(rows, "shown_gross_presentment")
    total_vat   = _sum_dec(rows, "shown_tax_presentment")
    total_net   = _sum_dec(rows, "shown_net_presentment")
    disc_rows   = [r for r in rows if abs(_dec(r.get("descuadre"))) > Decimal("0.01")]
    total_disc  = sum((_dec(r["descuadre"]) for r in disc_rows), Decimal("0")).quantize(Decimal("0.01"))
    n_orders    = len(rows)
    n_disc      = len(disc_rows)

    row = 1

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    _sh(row, 22)
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    _cell(
        row, 1,
        f"VAT Report — {data.company_code} — {data.region} — {month_label}"
        f"  |  Generated: {generated_ts}",
        font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
        fill=HEADER_FILL,
        align=CENTER,
    )
    row += 1

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    _sh(row, 22)
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    _cell(
        row, 1,
        f"Order detail for {period}. "
        f"Source: finance.informe_vat_gestorias_detalle. "
        f"Yellow rows: |descuadre| > 0.01. Hannun orders included.",
        font=ITALIC_GREY,
        align=LEFT_WRAP,
        border=None,
    )
    row += 1

    # ── Row 3: Quick stats ────────────────────────────────────────────────────
    _sh(row, 18)
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    _cell(
        row, 1,
        f"Total: {n_orders} orders  |  "
        f"Gross: {total_gross:,.2f}  |  "
        f"VAT: {total_vat:,.2f}  |  "
        f"Net: {total_net:,.2f}  |  "
        f"Discrepancies: {n_disc} orders ({total_disc:,.2f})",
        font=ITALIC_GREY,
        align=LEFT_WRAP,
        border=None,
    )
    row += 1

    # ── Row 4: blank ──────────────────────────────────────────────────────────
    _blank_row(row); row += 1

    # ── Row 5: Column headers ─────────────────────────────────────────────────
    _sh(row)
    col_headers = [
        "Date",
        "Order #",
        "Country",
        "State",
        "VAT rate\n(theor.)",
        "VAT rate\n(applied)",
        "VAT rate\n(calc.)",
        "Gross",
        "VAT / Tax",
        "Net",
    ]
    if has_fees:
        col_headers.append("Shopify fee")
    col_headers += [
        "Discrepancy",
        "Payment gateways",
        "Hannun",
        "Rever",
    ]
    for col, hdr in enumerate(col_headers, 1):
        _cell(row, col, hdr, font=BOLD, fill=COL_HDR_FILL, align=CENTER)
    row += 1

    # ── Rows 6+: Data ─────────────────────────────────────────────────────────
    # Sort: already ordered by shipping_country_code, order_date, order_name from DB
    # (but we rely on what DB returned; re-sort here for safety)
    sorted_rows = sorted(
        rows,
        key=lambda r: (
            str(r.get("shipping_country_code") or ""),
            str(r.get("order_date") or ""),
            str(r.get("order_name") or ""),
        ),
    )

    tot_gross = Decimal("0")
    tot_tax   = Decimal("0")
    tot_net   = Decimal("0")
    tot_fee   = Decimal("0")
    tot_disc  = Decimal("0")
    tot_count = 0

    # Column offsets: when has_fees, Discrepancy/Gateways/Hannun/Rever shift right by 1
    COL_DISC  = 12 if has_fees else 11
    COL_GW    = 13 if has_fees else 12
    COL_HAN   = 14 if has_fees else 13
    COL_REV   = 15 if has_fees else 14

    for r in sorted_rows:
        gross       = _dec(r.get("shown_gross_presentment")).quantize(Decimal("0.01"))
        tax         = _dec(r.get("shown_tax_presentment")).quantize(Decimal("0.01"))
        net         = _dec(r.get("shown_net_presentment")).quantize(Decimal("0.01"))
        descuadre   = _dec(r.get("descuadre")).quantize(Decimal("0.01"))
        is_warn     = abs(descuadre) > Decimal("0.01")
        fill        = WARN_FILL if is_warn else None

        # VAT rate calculated: shown_tax / shown_net when net != 0
        if net != Decimal("0"):
            rate_calc = float(tax / net)
        else:
            rate_calc = 0.0

        # Payment gateways: join jsonb array (stored as list or string) as CSV
        gw_raw = r.get("payment_gateway_names")
        if isinstance(gw_raw, list):
            gateways = ", ".join(str(g) for g in gw_raw)
        elif isinstance(gw_raw, str):
            import json as _json
            try:
                parsed = _json.loads(gw_raw)
                gateways = ", ".join(str(g) for g in parsed) if isinstance(parsed, list) else gw_raw
            except Exception:
                gateways = gw_raw
        else:
            gateways = ""

        state = r.get("shipping_state_code") or "—"
        hannun_flag = "✓" if int(r.get("is_hannun_tag") or 0) == 1 else ""
        rever_flag  = "✓" if int(r.get("is_rever_tag")  or 0) == 1 else ""

        _sh(row)
        _cell(row,  1, r.get("order_date"),                      font=NORMAL, fill=fill, align=CENTER)
        _cell(row,  2, r.get("order_name"),                      font=NORMAL, fill=fill, align=LEFT)
        _cell(row,  3, r.get("shipping_country_code"),           font=NORMAL, fill=fill, align=CENTER)
        _cell(row,  4, state,                                    font=NORMAL, fill=fill, align=CENTER)
        _cell(row,  5, float(r.get("standard_rate") or 0),      font=NORMAL, fill=fill, align=CENTER, fmt="0%")
        _cell(row,  6, float(r.get("tax_rate") or 0),            font=NORMAL, fill=fill, align=CENTER, fmt="0%")
        _cell(row,  7, rate_calc,                                font=NORMAL, fill=fill, align=CENTER, fmt="0%")
        _cell(row,  8, float(gross),                             font=NORMAL, fill=fill, align=RIGHT,  fmt=MONEY_FMT)
        _cell(row,  9, float(tax),                               font=NORMAL, fill=fill, align=RIGHT,  fmt=MONEY_FMT)
        _cell(row, 10, float(net),                               font=NORMAL, fill=fill, align=RIGHT,  fmt=MONEY_FMT)

        if has_fees:
            order_fee = data.fees_by_order.get(str(r.get("order_name") or ""), Decimal("0")).quantize(Decimal("0.01"))
            _cell(row, 11, float(order_fee), font=NORMAL, fill=fill, align=RIGHT, fmt=MONEY_FMT)
            tot_fee += order_fee

        # Discrepancy: 0.00 shown as "-" via MONEY_FMT ('-' in format string)
        disc_cell = ws.cell(row=row, column=COL_DISC, value=float(descuadre))
        disc_cell.font         = NORMAL
        disc_cell.alignment    = RIGHT
        disc_cell.border       = THIN
        disc_cell.number_format = MONEY_FMT
        if fill:
            disc_cell.fill = fill
        _cell(row, COL_GW,  gateways,    font=NORMAL, fill=fill, align=LEFT)
        _cell(row, COL_HAN, hannun_flag, font=NORMAL, fill=fill, align=CENTER)
        _cell(row, COL_REV, rever_flag,  font=NORMAL, fill=fill, align=CENTER)
        row += 1

        tot_gross += gross
        tot_tax   += tax
        tot_net   += net
        tot_disc  += descuadre
        tot_count += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    _sh(row, ROW_H + 1)
    ws.merge_cells(f"A{row}:G{row}")
    _cell(row, 1, f"TOTAL  ({tot_count} orders)", font=BOLD, fill=TOTAL_FILL, align=LEFT)
    _cell(row,  8, float(tot_gross.quantize(Decimal("0.01"))),
          font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row,  9, float(tot_tax.quantize(Decimal("0.01"))),
          font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row, 10, float(tot_net.quantize(Decimal("0.01"))),
          font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
    if has_fees:
        _cell(row, 11, float(tot_fee.quantize(Decimal("0.01"))),
              font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
    _cell(row, COL_DISC, float(tot_disc.quantize(Decimal("0.01"))),
          font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=MONEY_FMT)
    # Fill remaining cols
    for col in [COL_GW, COL_HAN, COL_REV]:
        _cell(row, col, None, font=BOLD, fill=TOTAL_FILL, align=CENTER)
