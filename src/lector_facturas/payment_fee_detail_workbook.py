from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from lector_facturas.api.store import ReviewStore
from lector_facturas.payment_fees import (
    PAYPAL_PLATFORM,
    SHOPIFY_PLATFORM,
    PaymentFeeSummaryRow,
    PaymentOrderTransaction,
    parse_datetime,
)


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_WHITE_BOLD = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY_FMT = '#,##0.00;[Red](#,##0.00);-'
_MADRID_TZ = ZoneInfo("Europe/Madrid")

_COMPANY_LABELS = {
    "SL": "Artesta Store, S.L",
    "LTD": "Artesta Stores (UK) Ltd",
    "INC": "Artesta Inc",
}


@dataclass(frozen=True)
class PaymentFeeDetailBundle:
    company_code: str
    period_yyyymm: str
    summaries: tuple[PaymentFeeSummaryRow, ...]
    transactions: tuple[PaymentOrderTransaction, ...]
    shopify_raw_rows: tuple[dict, ...]
    paypal_raw_rows: tuple[dict, ...]


def default_output_path(root: Path, company_code: str, period_yyyymm: str) -> Path:
    return root / "output" / "spreadsheet" / f"payment_fees_{company_code.lower()}_{period_yyyymm}.xlsx"


def collect_payment_fee_detail(*, company_code: str, period_yyyymm: str, database_url: str | None) -> PaymentFeeDetailBundle:
    store = ReviewStore(database_url=database_url) if database_url else ReviewStore()
    normalized_company = company_code.upper()
    transactions = tuple(
        tx for tx in store.list_payment_order_transactions(
            company_code=normalized_company,
            include_unpaid_shopify=True,
        )
        if _transaction_period(tx) == period_yyyymm
    )
    summaries = _summaries_from_transactions(transactions)
    shopify_raw_rows = tuple(
        row for row in store.list_shopify_payout_transactions()
        if str(row.get("company_code", "")).upper() == normalized_company
        and _transaction_month(str(row.get("transaction_date", "") or "")) == period_yyyymm
        and str(row.get("type", "")).lower() != "transfer"
    )
    paypal_raw_rows = tuple(
        row for row in store.list_paypal_transactions_raw()
        if str(row.get("company_code", "")).upper() == normalized_company
        and _transaction_month(str(row.get("transaction_date", "") or "")) == period_yyyymm
    )
    return PaymentFeeDetailBundle(
        company_code=normalized_company,
        period_yyyymm=period_yyyymm,
        summaries=summaries,
        transactions=transactions,
        shopify_raw_rows=shopify_raw_rows,
        paypal_raw_rows=paypal_raw_rows,
    )


def build_payment_fee_detail_workbook(bundle: PaymentFeeDetailBundle, output_path: Path | None = None) -> BytesIO:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary_sheet(ws_summary, bundle)
    _build_transaction_sheet(wb.create_sheet("Detail"), bundle)
    _build_shopify_raw_sheet(wb.create_sheet("Shopify Raw"), bundle.shopify_raw_rows)
    _build_paypal_raw_sheet(wb.create_sheet("PayPal Raw"), bundle.paypal_raw_rows)

    buf = BytesIO()
    wb.save(buf)
    if output_path is not None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(buf.getvalue())
    buf.seek(0)
    return buf


def _build_summary_sheet(ws, bundle: PaymentFeeDetailBundle) -> None:
    company_label = _COMPANY_LABELS.get(bundle.company_code, bundle.company_code)
    ws["A1"] = "Payment Fees Summary"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = "Company"
    ws["B2"] = company_label
    ws["A3"] = "Period"
    ws["B3"] = bundle.period_yyyymm

    headers = [
        "Platform",
        "Market",
        "Currency",
        "Transactions",
        "Payouts",
        "Fee amount",
        "Chargeback fee",
        "Chargeback principal",
        "PYG cost",
        "Net amount",
    ]
    start_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_BOLD
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

    for idx, summary in enumerate(bundle.summaries, start=start_row + 1):
        values = [
            summary.platform.upper(),
            summary.market_code,
            summary.currency_code,
            summary.transactions_count,
            summary.payout_count,
            summary.fee_amount,
            summary.chargeback_fee_amount,
            summary.chargeback_amount,
            summary.total_cost_amount,
            summary.net_amount,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.border = _BORDER
            if isinstance(value, Decimal):
                cell.number_format = _MONEY_FMT

    totals_row = start_row + 1 + len(bundle.summaries)
    ws.cell(row=totals_row, column=1, value="TOTAL").font = _BOLD
    ws.cell(row=totals_row, column=1).fill = _SECTION_FILL
    ws.cell(row=totals_row, column=1).border = _BORDER
    for col in range(2, len(headers) + 1):
        cell = ws.cell(row=totals_row, column=col)
        if col in {4, 5, 6, 7, 8, 9, 10}:
            col_letter = get_column_letter(col)
            cell.value = f"=SUM({col_letter}{start_row + 1}:{col_letter}{totals_row - 1})"
            if col >= 6:
                cell.number_format = _MONEY_FMT
        cell.fill = _SECTION_FILL
        cell.border = _BORDER
        if col in {4, 5}:
            cell.alignment = Alignment(horizontal="center")

    payout_start_row = totals_row + 3
    _build_shopify_payout_timing_section(ws, bundle, payout_start_row)
    _set_widths(ws, [18, 14, 12, 14, 12, 14, 16, 18, 14, 14])


def _build_transaction_sheet(ws, bundle: PaymentFeeDetailBundle) -> None:
    headers = [
        "Platform",
        "Market",
        "Order",
        "Type",
        "Status",
        "Transaction date",
        "Payout date",
        "Payout ID",
        "Currency",
        "Gross",
        "Fee",
        "Chargeback fee",
        "Chargeback principal",
        "PYG cost",
        "Net",
    ]
    _write_headers(ws, headers)
    for row_idx, tx in enumerate(bundle.transactions, start=2):
        values = [
            tx.platform.upper(),
            tx.market_code,
            tx.order_name,
            tx.transaction_type,
            tx.status,
            tx.transaction_date,
            tx.payout_date,
            tx.external_payout_id,
            tx.currency_code,
            tx.gross_amount,
            tx.fee_amount,
            tx.chargeback_fee_amount,
            tx.chargeback_amount,
            tx.fee_amount + tx.chargeback_fee_amount,
            tx.net_amount,
        ]
        _write_row(ws, row_idx, values, money_cols={10, 11, 12, 13, 14, 15})
    _set_widths(ws, [12, 12, 16, 14, 16, 22, 22, 18, 10, 12, 12, 16, 18, 12, 12])


def _build_shopify_raw_sheet(ws, rows: tuple[dict, ...]) -> None:
    headers = [
        "Transaction date",
        "Type",
        "Order",
        "Payout date",
        "Payout ID",
        "Amount",
        "Fee",
        "Net",
        "Method",
        "Currency",
        "Presentment amount",
        "Presentment currency",
    ]
    _write_headers(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        values = [
            row.get("transaction_date", ""),
            _display_shopify_type(row),
            row.get("order_name", ""),
            row.get("payout_date", ""),
            row.get("payout_id", ""),
            _to_decimal(row.get("amount")),
            _to_decimal(row.get("fee")),
            _to_decimal(row.get("net")),
            row.get("payment_method_name", ""),
            row.get("currency", ""),
            _to_decimal(row.get("presentment_amount")) if row.get("presentment_amount") not in ("", None) else None,
            row.get("presentment_currency", ""),
        ]
        _write_row(ws, row_idx, values, money_cols={6, 7, 8, 11})
    _set_widths(ws, [22, 12, 16, 22, 18, 12, 12, 12, 16, 10, 16, 18])


def _build_paypal_raw_sheet(ws, rows: tuple[dict, ...]) -> None:
    headers = [
        "Transaction date",
        "Order",
        "Type",
        "Status",
        "Currency",
        "Gross",
        "Fee",
        "Net",
        "Transaction ID",
        "Reference ID",
    ]
    _write_headers(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        values = [
            row.get("transaction_date", ""),
            row.get("shopify_order_name") or row.get("order_number") or row.get("invoice_number", ""),
            row.get("tipo", ""),
            row.get("estado", ""),
            row.get("divisa", ""),
            _to_decimal(row.get("bruto")),
            abs(_to_decimal(row.get("tarifa")) or Decimal("0.00")) if row.get("tarifa") not in ("", None) else None,
            _to_decimal(row.get("neto")),
            row.get("transaction_id", ""),
            row.get("reference_transaction_id", ""),
        ]
        _write_row(ws, row_idx, values, money_cols={6, 7, 8})
    _set_widths(ws, [22, 18, 14, 14, 10, 12, 12, 12, 20, 20])


def _write_headers(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_BOLD
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")


def _write_row(ws, row_idx: int, values: list, *, money_cols: set[int]) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = _BORDER
        if col in money_cols and value is not None:
            cell.number_format = _MONEY_FMT


def _set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _transaction_period(tx: PaymentOrderTransaction) -> str:
    if tx.platform == SHOPIFY_PLATFORM:
        return _transaction_month(tx.transaction_date)
    return tx.period_yyyymm


def _transaction_month(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    return parse_datetime(value).astimezone(_MADRID_TZ).strftime("%Y%m")


def _to_decimal(value) -> Decimal | None:
    if value in ("", None):
        return None
    return Decimal(str(value))


def _summaries_from_transactions(transactions: tuple[PaymentOrderTransaction, ...]) -> tuple[PaymentFeeSummaryRow, ...]:
    grouped: dict[tuple[str, str, str, str], list[PaymentOrderTransaction]] = defaultdict(list)
    for tx in transactions:
        grouped[(tx.company_code, tx.platform, tx.market_code, tx.currency_code)].append(tx)

    rows: list[PaymentFeeSummaryRow] = []
    for (company_code, platform, market_code, currency_code), items in sorted(grouped.items()):
        order_names = {tx.order_name for tx in items if tx.order_name}
        payout_ids = {tx.external_payout_id for tx in items if tx.external_payout_id}
        fee_amount = sum((tx.fee_amount for tx in items), Decimal("0.00"))
        chargeback_amount = sum((tx.chargeback_amount for tx in items), Decimal("0.00"))
        chargeback_fee_amount = sum((tx.chargeback_fee_amount for tx in items), Decimal("0.00"))
        gross_amount = sum((tx.gross_amount for tx in items), Decimal("0.00"))
        net_amount = sum((tx.net_amount for tx in items), Decimal("0.00"))
        rows.append(
            PaymentFeeSummaryRow(
                company_code=company_code,
                period_yyyymm=items[0].period_yyyymm,
                platform=platform,
                market_code=market_code,
                currency_code=currency_code,
                orders_count=len(order_names),
                transactions_count=len(items),
                gross_amount=gross_amount,
                fee_amount=fee_amount,
                chargeback_amount=chargeback_amount,
                chargeback_fee_amount=chargeback_fee_amount,
                total_cost_amount=fee_amount + chargeback_fee_amount,
                net_amount=net_amount,
                payout_count=len(payout_ids),
            )
        )
    return tuple(rows)


def _build_shopify_payout_timing_section(ws, bundle: PaymentFeeDetailBundle, start_row: int) -> None:
    ws.cell(row=start_row, column=1, value="Shopify Payout Timing").font = _BOLD
    ws.cell(row=start_row + 1, column=1, value="Only Shopify transactions with transaction date in the selected period.")

    headers = ["Payout month", "Transactions", "Fee amount", "Net amount"]
    header_row = start_row + 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_BOLD
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

    grouped: dict[str, dict[str, Decimal | int]] = defaultdict(
        lambda: {"transactions": 0, "fee_amount": Decimal("0.00"), "net_amount": Decimal("0.00")}
    )
    for row in bundle.shopify_raw_rows:
        payout_label = _payout_label(str(row.get("payout_date", "") or ""))
        grouped[payout_label]["transactions"] += 1
        grouped[payout_label]["fee_amount"] += _to_decimal(row.get("fee")) or Decimal("0.00")
        grouped[payout_label]["net_amount"] += _to_decimal(row.get("net")) or Decimal("0.00")

    ordered_labels = sorted(grouped.keys(), key=_payout_label_sort_key)
    for idx, label in enumerate(ordered_labels, start=header_row + 1):
        values = [
            label,
            grouped[label]["transactions"],
            grouped[label]["fee_amount"],
            grouped[label]["net_amount"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.border = _BORDER
            if col >= 3:
                cell.number_format = _MONEY_FMT

    totals_row = header_row + 1 + len(ordered_labels)
    ws.cell(row=totals_row, column=1, value="TOTAL").font = _BOLD
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=totals_row, column=col)
        cell.fill = _SECTION_FILL
        cell.border = _BORDER
    if ordered_labels:
        ws.cell(row=totals_row, column=2, value=f"=SUM(B{header_row + 1}:B{totals_row - 1})")
        ws.cell(row=totals_row, column=3, value=f"=SUM(C{header_row + 1}:C{totals_row - 1})").number_format = _MONEY_FMT
        ws.cell(row=totals_row, column=4, value=f"=SUM(D{header_row + 1}:D{totals_row - 1})").number_format = _MONEY_FMT


def _payout_label(value: str) -> str:
    if not value.strip():
        return "No payout"
    return _transaction_month(value)


def _payout_label_sort_key(label: str) -> tuple[int, str]:
    if label == "No payout":
        return (1, label)
    return (0, label)


def _display_shopify_type(row: dict) -> str:
    value = str(row.get("type", "")).lower()
    if value == "dispute_withdrawal":
        return "chargeback"
    return value
