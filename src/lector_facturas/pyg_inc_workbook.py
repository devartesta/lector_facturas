from __future__ import annotations

from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:  # pragma: no cover
    psycopg = None
    dict_row = None

from lector_facturas.fx_rates import EcbFxService, FxRateAuditRow


COMPANY_CODE = "INC"
COMPANY_NAME = "ARTESTA INC"
REPORTING_CURRENCY = "USD"
DISPLAY_TIMEZONE = ZoneInfo("Europe/Madrid")
MONTH_NAMES_ES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]
MONEY_FORMAT = '#,##0.00;[Red](#,##0.00);-'
PERCENT_FORMAT = '0.0%;[Red](0.0%);-'
ROW_HEIGHT = 12
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
KPI_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBSECTION_FILL = PatternFill("solid", fgColor="EAF3F8")
SUBTOTAL_FILL = PatternFill("solid", fgColor="F3F6F8")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
TITLE_FONT = Font(size=12, bold=True, color="FFFFFF")
THIN_TOP_BORDER = Border(top=Side(style="thin", color="A6A6A6"))
MEDIUM_TOP_BORDER = Border(top=Side(style="medium", color="7F7F7F"))

DEFAULT_SALES_MARKETS = ("US",)
DEFAULT_MANUFACTURING_LINES = ("JONDO", "TGI")
DEFAULT_LOGISTICS_LINES = ("TGI",)
DEFAULT_PAYMENT_FEE_LINES = ("SHOPIFY",)
DEFAULT_SHARED_SERVICE_LINES = ("SHAREDSERVICESSL",)
DEFAULT_ADMIN_LINES = ("CONTINUUM", "HUSHED", "IPOSTAL", "QUICKBOOKS", "REGUS")
DEFAULT_TECH_LINES = ("REVER",)


@dataclass(frozen=True)
class StageRow:
    yyyymm: str
    entity: str
    line_item: str
    detail: str
    amount_net: Decimal
    currency: str
    source: str


@dataclass(frozen=True)
class ExpenseRow:
    yyyymm: str
    entity: str
    category: str
    subcategory: str
    supplier_code: str
    detail: str
    amount_net: Decimal
    currency: str
    source: str
    invoice_number: str = ""
    drive_url: str = ""


@dataclass(frozen=True)
class PaymentFeeRow:
    yyyymm: str
    entity: str
    supplier_code: str
    amount_net: Decimal
    currency: str
    source: str


@dataclass(frozen=True)
class ProviderCatalogRow:
    supplier_code: str
    supplier_name: str
    current_folder: str
    destination_path: str
    notes: str


@dataclass(frozen=True)
class PygIncDataBundle:
    year: int
    generated_at: datetime
    sales_rows: tuple[StageRow, ...]
    expense_rows: tuple[ExpenseRow, ...]
    payment_fee_rows: tuple[PaymentFeeRow, ...]
    provider_catalog_rows: tuple[ProviderCatalogRow, ...]
    fx_rate_rows: tuple[FxRateAuditRow, ...] = field(default_factory=tuple)
    otros_ingresos_by_period: dict[str, Decimal] = field(default_factory=dict)
    diferencias_divisas_by_period: dict[str, Decimal] = field(default_factory=dict)
    frame_consumed_by_period: dict[str, Decimal] = field(default_factory=dict)
    frame_opening_by_period: dict[str, Decimal] = field(default_factory=dict)
    frame_closing_by_period: dict[str, Decimal] = field(default_factory=dict)


def default_output_path(root: Path, year: int) -> Path:
    return root / "output" / "spreadsheet" / f"pyg_inc_{year}.xlsx"


def month_keys(year: int) -> list[str]:
    return [f"{year}{month:02d}" for month in range(1, 13)]


def collect_pyg_inc_data(*, year: int, database_url: str | None) -> PygIncDataBundle:
    if not database_url or psycopg is None:
        return PygIncDataBundle(year=year, generated_at=datetime.now(UTC), sales_rows=(), expense_rows=(), payment_fee_rows=(), provider_catalog_rows=())

    with psycopg.connect(database_url, row_factory=dict_row) as conn:
        sales = conn.execute(
            """
            SELECT
                order_month_yyyymm,
                shipping_country_code,
                payment_currency,
                SUM(net) AS amount_net
            FROM finance.ventas_pyg
            WHERE order_month_yyyymm LIKE %(period)s
              AND payment_currency = 'USD'
              AND shipping_country_code = 'US'
              AND COALESCE(is_hannun_tag, 0) = 0
              AND COALESCE(is_rever_tag, 0) = 0
            GROUP BY order_month_yyyymm, shipping_country_code, payment_currency
            ORDER BY order_month_yyyymm
            """,
            {"period": f"{year}%"},
        ).fetchall()
        docs = conn.execute(
            """
            SELECT
                period_yyyymm,
                supplier_code,
                billed_company_name,
                division_invoice,
                document_type,
                currency_code,
                net_amount,
                invoice_number,
                drive_url
            FROM invoices.documents
            WHERE company_code = %(company)s
              AND status = 'classified'
              AND period_yyyymm LIKE %(period)s
            ORDER BY period_yyyymm, supplier_code, invoice_number
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        suppliers = conn.execute(
            """
            SELECT supplier_code, supplier_name, current_folder, destination_path, notes
            FROM invoices.suppliers
            WHERE company_code = %(company)s AND is_active = TRUE
            ORDER BY destination_path, supplier_code
            """,
            {"company": COMPANY_CODE},
        ).fetchall()
        payment_fees = conn.execute(
            """
            SELECT period_yyyymm, platform, currency_code, SUM(total_cost_amount) AS amount_net
            FROM invoices.payment_fee_monthly_summary
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s
            GROUP BY period_yyyymm, platform, currency_code
            ORDER BY period_yyyymm, platform
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        otros_gastos = conn.execute(
            """
            SELECT period_yyyymm, amount_eur, notes
            FROM invoices.otros_gastos
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s
            ORDER BY period_yyyymm
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        otros_ingresos = conn.execute(
            """
            SELECT period_yyyymm, amount_eur
            FROM invoices.otros_ingresos
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s
            ORDER BY period_yyyymm
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()
        diferencias_divisas = conn.execute(
            """
            SELECT period_yyyymm, amount_eur
            FROM invoices.diferencias_divisas
            WHERE company_code = %(company)s AND period_yyyymm LIKE %(period)s
            ORDER BY period_yyyymm
            """,
            {"company": COMPANY_CODE, "period": f"{year}%"},
        ).fetchall()

    sales_rows = tuple(
        StageRow(
            yyyymm=str(row["order_month_yyyymm"]),
            entity=COMPANY_CODE,
            line_item="US",
            detail=str(row["shipping_country_code"] or "US"),
            amount_net=_decimal(row["amount_net"]),
            currency=str(row["payment_currency"] or "USD"),
            source="finance.ventas_pyg",
        )
        for row in sales
    )

    supplier_map = {str(row["supplier_code"]): row for row in suppliers}
    expense_rows: list[ExpenseRow] = []
    for row in docs:
        supplier_code = str(row["supplier_code"] or "").upper()
        subcategory = _map_expense_subcategory(
            supplier_code=supplier_code,
            division_invoice=str(row["division_invoice"] or ""),
            supplier_meta=supplier_map.get(supplier_code),
        )
        if not subcategory:
            continue

        expense_rows.append(
            ExpenseRow(
                yyyymm=str(row["period_yyyymm"]),
                entity=COMPANY_CODE,
                category="opex" if subcategory in {"shared_services", "administration", "technology"} else "cogs",
                subcategory=subcategory,
                supplier_code=supplier_code,
                detail=str(row["division_invoice"] or row["document_type"] or "").lower(),
                amount_net=_decimal(row["net_amount"]),
                currency=str(row["currency_code"] or "USD"),
                source="documents",
                invoice_number=str(row["invoice_number"] or ""),
                drive_url=str(row["drive_url"] or ""),
            )
        )

    for row in otros_gastos:
        expense_rows.append(ExpenseRow(
            yyyymm=str(row["period_yyyymm"]),
            entity=COMPANY_CODE,
            category="opex",
            subcategory="otros_gastos",
            supplier_code="OTROSGASTOS",
            detail=str(row["notes"] or ""),
            amount_net=_decimal(row["amount_eur"]),
            currency="EUR",
            source="otros_gastos",
        ))

    payment_fee_rows = tuple(
        PaymentFeeRow(
            yyyymm=str(row["period_yyyymm"]),
            entity=COMPANY_CODE,
            supplier_code=str(row["platform"] or "").upper(),
            amount_net=_decimal(row["amount_net"]),
            currency=str(row["currency_code"] or "USD"),
            source="payment_fee_monthly_summary",
        )
        for row in payment_fees
    )

    provider_rows = tuple(
        ProviderCatalogRow(
            supplier_code=str(row["supplier_code"]),
            supplier_name=str(row["supplier_name"]),
            current_folder=str(row["current_folder"]),
            destination_path=str(row["destination_path"]),
            notes=str(row["notes"] or ""),
        )
        for row in suppliers
    )
    otros_ingresos_by_period = {str(row["period_yyyymm"]): _decimal(row["amount_eur"]) for row in otros_ingresos}
    diferencias_divisas_by_period = {str(row["period_yyyymm"]): _decimal(row["amount_eur"]) for row in diferencias_divisas}

    from lector_facturas.supply_stock import compute_frame_stock_by_year
    frame_stock = compute_frame_stock_by_year(fabricante="TGI", year=year, database_url=database_url)
    frame_consumed_by_period = {k: v.consumed_value for k, v in frame_stock.items()}
    frame_opening_by_period = {k: v.opening_value for k, v in frame_stock.items()}
    frame_closing_by_period = {k: v.closing_value for k, v in frame_stock.items()}

    return PygIncDataBundle(
        year=year,
        generated_at=datetime.now(UTC),
        sales_rows=sales_rows,
        expense_rows=tuple(expense_rows),
        payment_fee_rows=payment_fee_rows,
        provider_catalog_rows=provider_rows,
        otros_ingresos_by_period=otros_ingresos_by_period,
        diferencias_divisas_by_period=diferencias_divisas_by_period,
        frame_consumed_by_period=frame_consumed_by_period,
        frame_opening_by_period=frame_opening_by_period,
        frame_closing_by_period=frame_closing_by_period,
    )


def build_pyg_inc_workbook(bundle: PygIncDataBundle, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    fx_service = EcbFxService()
    sales_sheet_rows, sales_fx_rows = _stage_rows_with_fx(bundle.sales_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    expense_sheet_rows, expense_fx_rows = _expense_rows_with_fx(bundle.expense_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    payment_fee_sheet_rows, payment_fee_fx_rows = _payment_fee_rows_with_fx(bundle.payment_fee_rows, reporting_currency=REPORTING_CURRENCY, fx_service=fx_service)
    fx_rate_rows = _dedupe_fx_rows(bundle.fx_rate_rows + sales_fx_rows + expense_fx_rows + payment_fee_fx_rows)

    generated_at_utc = bundle.generated_at.astimezone(UTC) if bundle.generated_at.tzinfo else bundle.generated_at.replace(tzinfo=UTC)
    _sheet(wb, "params", ["key", "value"], [["entity", COMPANY_CODE], ["reporting_currency", REPORTING_CURRENCY], ["year", bundle.year], ["generated_at_utc", generated_at_utc.replace(microsecond=0).isoformat().replace("+00:00", "Z")]])
    _sheet(wb, "i-shopify-inc", ["yyyymm", "entity", "line_item", "detail", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source"], sales_sheet_rows)
    _sheet(wb, "g-expenses-inc", ["yyyymm", "entity", "category", "subcategory", "supplier_code", "detail", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source", "invoice_number", "drive_url"], expense_sheet_rows)
    _sheet(wb, "g-payment-fees-inc", ["yyyymm", "entity", "supplier_code", "amount_original", "currency_original", "reporting_currency", "fx_rate", "amount_reporting", "source"], payment_fee_sheet_rows)
    _sheet(wb, "fx-rates", ["yyyymm", "rate_date", "currency_original", "reporting_currency", "reference_rate", "fx_rate", "source"], [[r.yyyymm, r.rate_date, r.currency_original, r.reporting_currency, float(r.reference_rate), float(r.fx_rate), r.source] for r in fx_rate_rows])
    _sheet(wb, "catalog-inc", ["supplier_code", "supplier_name", "current_folder", "destination_path", "notes"], [[r.supplier_code, r.supplier_name, r.current_folder, r.destination_path, r.notes] for r in bundle.provider_catalog_rows])
    _main_sheet(wb, bundle)
    _count_sheet_inc(wb, bundle)
    _add_back_links(wb)
    wb.save(output_path)
    return output_path


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(header) + 2)
        if header in {"amount_original", "amount_reporting"}:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).number_format = MONEY_FORMAT
        elif header in {"fx_rate", "reference_rate"}:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).number_format = "0.00000000"
    ws.freeze_panes = "A2"


def _add_back_links(wb: Workbook) -> None:
    detail_sheets = ("i-shopify-inc", "g-expenses-inc", "g-payment-fees-inc", "fx-rates", "catalog-inc", "params")
    for title in detail_sheets:
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        ws.insert_rows(1)
        ws["A1"] = "<- Volver a P&G-INC"
        ws["A1"]._hyperlink = Hyperlink(ref="A1", location="'P&G-INC'!A1", display="<- Volver a P&G-INC")
        ws["A1"].font = Font(bold=True, color="1F1F1F")
        ws["A1"].fill = SUBTOTAL_FILL
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = ROW_HEIGHT
        _apply_invoice_links(ws)
        ws.freeze_panes = "A3"


def _apply_invoice_links(ws) -> None:
    headers = [cell.value for cell in ws[2]]
    if "invoice_number" not in headers or "drive_url" not in headers:
        return
    invoice_col = headers.index("invoice_number") + 1
    drive_url_col = headers.index("drive_url") + 1
    for row_idx in range(3, ws.max_row + 1):
        invoice_cell = ws.cell(row=row_idx, column=invoice_col)
        drive_url_cell = ws.cell(row=row_idx, column=drive_url_col)
        if invoice_cell.value and drive_url_cell.value:
            invoice_cell.hyperlink = str(drive_url_cell.value)
            invoice_cell.style = "Hyperlink"
    ws.column_dimensions[get_column_letter(drive_url_col)].hidden = True


def _stage_rows_with_fx(rows: tuple[StageRow, ...], *, reporting_currency: str, fx_service: EcbFxService) -> tuple[list[list[Any]], tuple[FxRateAuditRow, ...]]:
    rendered: list[list[Any]] = []
    fx_rows: list[FxRateAuditRow] = []
    for row in rows:
        converted, audit = fx_service.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency=reporting_currency, yyyymm=row.yyyymm)
        rendered.append([row.yyyymm, row.entity, row.line_item, row.detail, float(converted.amount_original), converted.currency_original, converted.reporting_currency, float(converted.fx_rate), float(converted.amount_reporting), row.source])
        fx_rows.append(audit)
    return rendered, tuple(fx_rows)


def _expense_rows_with_fx(rows: tuple[ExpenseRow, ...], *, reporting_currency: str, fx_service: EcbFxService) -> tuple[list[list[Any]], tuple[FxRateAuditRow, ...]]:
    rendered: list[list[Any]] = []
    fx_rows: list[FxRateAuditRow] = []
    for row in rows:
        converted, audit = fx_service.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency=reporting_currency, yyyymm=row.yyyymm)
        rendered.append([row.yyyymm, row.entity, row.category, row.subcategory, row.supplier_code, row.detail, float(converted.amount_original), converted.currency_original, converted.reporting_currency, float(converted.fx_rate), float(converted.amount_reporting), row.source, row.invoice_number, row.drive_url])
        fx_rows.append(audit)
    return rendered, tuple(fx_rows)


def _payment_fee_rows_with_fx(rows: tuple[PaymentFeeRow, ...], *, reporting_currency: str, fx_service: EcbFxService) -> tuple[list[list[Any]], tuple[FxRateAuditRow, ...]]:
    rendered: list[list[Any]] = []
    fx_rows: list[FxRateAuditRow] = []
    for row in rows:
        converted, audit = fx_service.convert(amount=row.amount_net, source_currency=row.currency, reporting_currency=reporting_currency, yyyymm=row.yyyymm)
        rendered.append([row.yyyymm, row.entity, row.supplier_code, float(converted.amount_original), converted.currency_original, converted.reporting_currency, float(converted.fx_rate), float(converted.amount_reporting), row.source])
        fx_rows.append(audit)
    return rendered, tuple(fx_rows)


def _dedupe_fx_rows(rows: tuple[FxRateAuditRow, ...]) -> tuple[FxRateAuditRow, ...]:
    unique: dict[tuple[str, str, str], FxRateAuditRow] = {}
    for row in rows:
        unique[(row.yyyymm, row.currency_original, row.reporting_currency)] = row
    return tuple(sorted(unique.values(), key=lambda item: (item.yyyymm, item.currency_original, item.reporting_currency)))


def _main_sheet(wb: Workbook, bundle: PygIncDataBundle) -> None:
    ws = wb.create_sheet("P&G-INC", 0)
    for idx, yyyymm in enumerate(month_keys(bundle.year), start=4):
        ws.cell(row=1, column=idx, value=yyyymm)
        ws.cell(row=2, column=idx, value=MONTH_NAMES_ES[idx - 4])
    ws["P1"] = "TOTAL"
    ws["P2"] = "Total"
    ws["Q1"] = ""
    ws["Q2"] = ""
    ws["A2"] = f"P&G {COMPANY_CODE} {bundle.year}"
    ws.merge_cells("A2:C2")
    generated_at_local = bundle.generated_at.astimezone(DISPLAY_TIMEZONE) if bundle.generated_at.tzinfo else bundle.generated_at.replace(tzinfo=UTC).astimezone(DISPLAY_TIMEZONE)
    ws["A3"] = f"Actualizado: {generated_at_local.strftime('%d/%m/%Y %H:%M')} h"
    ws.merge_cells("A3:C3")

    for row in (1, 2):
        for cell in ws[row]:
            cell.fill = HEADER_FILL
            cell.font = WHITE_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = TITLE_FONT
    ws["A3"].font = Font(size=9, italic=True, color="666666")
    ws.row_dimensions[1].hidden = True
    ws.row_dimensions[2].height = ROW_HEIGHT
    ws.row_dimensions[3].height = ROW_HEIGHT

    pos: dict[str, int] = {}
    row = 4
    pos["turnover"] = row; ws[f"A{row}"] = "Turnover"; ws[f"A{row}"].font = BOLD; row += 1
    pos["product_sales"] = row; ws[f"A{row}"] = "  Product sales"; ws[f"A{row}"].font = BOLD; row += 1
    pos["shopify"] = row; ws[f"A{row}"] = "    Shopify"; ws[f"A{row}"].font = BOLD; row += 1
    sales_rows = list(range(row, row + len(DEFAULT_SALES_MARKETS)))
    for idx, market in enumerate(DEFAULT_SALES_MARKETS):
        ws[f"A{row + idx}"] = f"      {market}"
    row += len(sales_rows)
    pos["otros_ingresos"] = row; ws[f"A{row}"] = "  Otros ingresos"; ws[f"A{row}"].font = BOLD; row += 2

    pos["expenses"] = row; ws[f"A{row}"] = "Expenses"; ws[f"A{row}"].font = BOLD; row += 1
    pos["cogs"] = row; ws[f"A{row}"] = "  COGS"; ws[f"A{row}"].font = BOLD; row += 1
    pos["manufacturing"] = row; ws[f"A{row}"] = "    Manufacturing"; ws[f"A{row}"].font = BOLD; row += 1
    manufacturing_rows = list(range(row, row + len(DEFAULT_MANUFACTURING_LINES)))
    for idx, supplier in enumerate(DEFAULT_MANUFACTURING_LINES):
        ws[f"A{row + idx}"] = f"      {supplier}"
    row += len(manufacturing_rows)
    pos["marcos_consumed"] = row; ws[f"A{row}"] = "      Consumo marcos"; row += 1
    pos["manufacturing_pct"] = row; ws[f"A{row}"] = "    % Manufacturing / sales"; ws[f"A{row}"].font = BOLD; row += 1
    pos["logistics"] = row; ws[f"A{row}"] = "    Logistics"; ws[f"A{row}"].font = BOLD; row += 1
    logistics_rows = list(range(row, row + len(DEFAULT_LOGISTICS_LINES)))
    for idx, supplier in enumerate(DEFAULT_LOGISTICS_LINES):
        ws[f"A{row + idx}"] = f"      {supplier}"
    row += len(logistics_rows)
    pos["logistics_pct"] = row; ws[f"A{row}"] = "    % Logistics / sales"; ws[f"A{row}"].font = BOLD; row += 1
    pos["payment_fees"] = row; ws[f"A{row}"] = "    Payment fees"; ws[f"A{row}"].font = BOLD; row += 1
    payment_fee_rows = list(range(row, row + len(DEFAULT_PAYMENT_FEE_LINES)))
    for idx, supplier in enumerate(DEFAULT_PAYMENT_FEE_LINES):
        ws[f"A{row + idx}"] = f"      {supplier}"
    row += len(payment_fee_rows) + 1
    pos["payment_fees_pct"] = row; ws[f"A{row}"] = "    % Payment fees / sales"; ws[f"A{row}"].font = BOLD; row += 2

    pos["gross_margin"] = row; ws[f"A{row}"] = "GROSS MARGIN (SALES-MANUFACTURING)"; ws[f"A{row}"].font = BOLD; row += 1
    pos["gross_margin_pct"] = row; ws[f"A{row}"] = "% GROSS MARGIN"; ws[f"A{row}"].font = BOLD; row += 1
    pos["contributive_margin"] = row; ws[f"A{row}"] = "CONTRIBUTIVE MARGIN (TURNOVER-COGS)"; ws[f"A{row}"].font = BOLD; row += 1
    pos["contributive_margin_pct"] = row; ws[f"A{row}"] = "% CONTRIBUTIVE MARGIN"; ws[f"A{row}"].font = BOLD; row += 2

    pos["opex"] = row; ws[f"A{row}"] = "  Opex"; ws[f"A{row}"].font = BOLD; row += 1
    pos["shared_services"] = row; ws[f"A{row}"] = "    Shared services"; ws[f"A{row}"].font = BOLD; row += 1
    shared_service_rows = list(range(row, row + len(DEFAULT_SHARED_SERVICE_LINES)))
    for idx, supplier in enumerate(DEFAULT_SHARED_SERVICE_LINES):
        ws[f"A{row + idx}"] = f"      {supplier}"
    row += len(shared_service_rows)
    pos["administration"] = row; ws[f"A{row}"] = "    Administration"; ws[f"A{row}"].font = BOLD; row += 1
    administration_rows = list(range(row, row + len(DEFAULT_ADMIN_LINES)))
    for idx, supplier in enumerate(DEFAULT_ADMIN_LINES):
        ws[f"A{row + idx}"] = f"      {supplier}"
    row += len(administration_rows)
    pos["technology"] = row; ws[f"A{row}"] = "    Technology"; ws[f"A{row}"].font = BOLD; row += 1
    technology_rows = list(range(row, row + len(DEFAULT_TECH_LINES)))
    for idx, supplier in enumerate(DEFAULT_TECH_LINES):
        ws[f"A{row + idx}"] = f"      {supplier}"
    row += len(technology_rows)
    pos["otros_gastos"] = row; ws[f"A{row}"] = "    Otros gastos"; ws[f"A{row}"].font = BOLD; row += 1
    pos["diferencias_divisas"] = row; ws[f"A{row}"] = "  Diferencias divisas"; ws[f"A{row}"].font = BOLD; row += 2
    pos["profit"] = row; ws[f"A{row}"] = "PROFIT"; ws[f"A{row}"].font = BOLD; row += 1
    pos["profit_pct"] = row; ws[f"A{row}"] = "% Profit / product sales"; ws[f"A{row}"].font = BOLD; row += 2
    pos["stock_inicial"] = row; ws[f"A{row}"] = "Stock inicial marcos"; ws[f"A{row}"].font = BOLD; row += 1
    pos["stock_final"] = row; ws[f"A{row}"] = "Stock final marcos"; ws[f"A{row}"].font = BOLD

    _fill_inc_formulas(
        ws,
        pos=pos,
        sales_rows=sales_rows,
        manufacturing_rows=manufacturing_rows,
        logistics_rows=logistics_rows,
        payment_fee_rows=payment_fee_rows,
        shared_service_rows=shared_service_rows,
        administration_rows=administration_rows,
        technology_rows=technology_rows,
    )
    for idx, yyyymm in enumerate(month_keys(bundle.year), start=4):
        amount = bundle.otros_ingresos_by_period.get(yyyymm)
        if amount is not None:
            ws.cell(row=pos["otros_ingresos"], column=idx).value = float(amount)
        amount_dd = bundle.diferencias_divisas_by_period.get(yyyymm)
        if amount_dd is not None:
            ws.cell(row=pos["diferencias_divisas"], column=idx).value = float(amount_dd)
        amount_mc = bundle.frame_consumed_by_period.get(yyyymm)
        if amount_mc is not None and amount_mc != 0:
            ws.cell(row=pos["marcos_consumed"], column=idx).value = float(amount_mc)
        amount_si = bundle.frame_opening_by_period.get(yyyymm)
        if amount_si is not None and amount_si != 0:
            ws.cell(row=pos["stock_inicial"], column=idx).value = float(amount_si)
        amount_sf = bundle.frame_closing_by_period.get(yyyymm)
        if amount_sf is not None and amount_sf != 0:
            ws.cell(row=pos["stock_final"], column=idx).value = float(amount_sf)
    _apply_inc_layout(
        ws,
        pos=pos,
        detail_rows=sales_rows + manufacturing_rows + logistics_rows + payment_fee_rows + shared_service_rows + administration_rows + technology_rows,
    )
    _add_navigation_links(
        ws,
        pos=pos,
        sales_rows=sales_rows,
        manufacturing_rows=manufacturing_rows,
        logistics_rows=logistics_rows,
        payment_fee_rows=payment_fee_rows,
        shared_service_rows=shared_service_rows,
        administration_rows=administration_rows,
        technology_rows=technology_rows,
    )
    percent_rows = {
        pos["manufacturing_pct"],
        pos["logistics_pct"],
        pos["payment_fees_pct"],
        pos["gross_margin_pct"],
        pos["contributive_margin_pct"],
        pos["profit_pct"],
    }
    last_formatted_row = pos.get("stock_final", pos["profit_pct"])
    for row_idx in range(3, last_formatted_row + 1):
        for col_idx in range(2, 18):
            ws.cell(row=row_idx, column=col_idx).number_format = PERCENT_FORMAT if row_idx in percent_rows else MONEY_FORMAT
    for col, width in (("A", 34), ("B", 2), ("C", 2), ("P", 14), ("Q", 3)):
        ws.column_dimensions[col].width = width
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["C"].hidden = True
    for idx in range(4, 16):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.freeze_panes = "D4"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False


def _fill_inc_formulas(
    ws,
    *,
    pos: dict[str, int],
    sales_rows: list[int],
    manufacturing_rows: list[int],
    logistics_rows: list[int],
    payment_fee_rows: list[int],
    shared_service_rows: list[int],
    administration_rows: list[int],
    technology_rows: list[int],
) -> None:
    for col in [get_column_letter(i) for i in range(4, 16)]:
        for row in sales_rows:
            ws[f"{col}{row}"] = f'=SUMIFS(\'i-shopify-inc\'!$I:$I,\'i-shopify-inc\'!$A:$A,{col}$1,\'i-shopify-inc\'!$C:$C,TRIM($A{row}))'
        ws[f"{col}{pos['shopify']}"] = f"=SUM({col}{sales_rows[0]}:{col}{sales_rows[-1]})"
        ws[f"{col}{pos['product_sales']}"] = f"={col}{pos['shopify']}"
        ws[f"{col}{pos['turnover']}"] = f"={col}{pos['product_sales']}+{col}{pos['otros_ingresos']}"

        for row in manufacturing_rows:
            ws[f"{col}{row}"] = f'=SUMIFS(\'g-expenses-inc\'!$K:$K,\'g-expenses-inc\'!$A:$A,{col}$1,\'g-expenses-inc\'!$D:$D,"manufacturing",\'g-expenses-inc\'!$E:$E,TRIM($A{row}))'
        for row in logistics_rows:
            ws[f"{col}{row}"] = f'=SUMIFS(\'g-expenses-inc\'!$K:$K,\'g-expenses-inc\'!$A:$A,{col}$1,\'g-expenses-inc\'!$D:$D,"logistics",\'g-expenses-inc\'!$E:$E,TRIM($A{row}))'
        for row in payment_fee_rows:
            ws[f"{col}{row}"] = f'=SUMIFS(\'g-payment-fees-inc\'!$H:$H,\'g-payment-fees-inc\'!$A:$A,{col}$1,\'g-payment-fees-inc\'!$C:$C,TRIM($A{row}))'
        for row in shared_service_rows:
            ws[f"{col}{row}"] = f'=SUMIFS(\'g-expenses-inc\'!$K:$K,\'g-expenses-inc\'!$A:$A,{col}$1,\'g-expenses-inc\'!$D:$D,"shared_services",\'g-expenses-inc\'!$E:$E,TRIM($A{row}))'
        for row in administration_rows:
            ws[f"{col}{row}"] = f'=SUMIFS(\'g-expenses-inc\'!$K:$K,\'g-expenses-inc\'!$A:$A,{col}$1,\'g-expenses-inc\'!$D:$D,"administration",\'g-expenses-inc\'!$E:$E,TRIM($A{row}))'
        for row in technology_rows:
            ws[f"{col}{row}"] = f'=SUMIFS(\'g-expenses-inc\'!$K:$K,\'g-expenses-inc\'!$A:$A,{col}$1,\'g-expenses-inc\'!$D:$D,"technology",\'g-expenses-inc\'!$E:$E,TRIM($A{row}))'

        ws[f"{col}{pos['manufacturing']}"] = f"=SUM({col}{manufacturing_rows[0]}:{col}{manufacturing_rows[-1]})+{col}{pos['marcos_consumed']}"
        ws[f"{col}{pos['manufacturing_pct']}"] = f'=IFERROR({col}{pos["manufacturing"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['logistics']}"] = f"=SUM({col}{logistics_rows[0]}:{col}{logistics_rows[-1]})"
        ws[f"{col}{pos['logistics_pct']}"] = f'=IFERROR({col}{pos["logistics"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['payment_fees']}"] = f"=SUM({col}{payment_fee_rows[0]}:{col}{payment_fee_rows[-1]})"
        ws[f"{col}{pos['payment_fees_pct']}"] = f'=IFERROR({col}{pos["payment_fees"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['cogs']}"] = f"={col}{pos['manufacturing']}+{col}{pos['logistics']}+{col}{pos['payment_fees']}"
        ws[f"{col}{pos['gross_margin']}"] = f"={col}{pos['product_sales']}-{col}{pos['manufacturing']}"
        ws[f"{col}{pos['gross_margin_pct']}"] = f'=IFERROR({col}{pos["gross_margin"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['contributive_margin']}"] = f"={col}{pos['turnover']}-{col}{pos['cogs']}"
        ws[f"{col}{pos['contributive_margin_pct']}"] = f'=IFERROR({col}{pos["contributive_margin"]}/{col}{pos["product_sales"]},0)'
        ws[f"{col}{pos['shared_services']}"] = f"=SUM({col}{shared_service_rows[0]}:{col}{shared_service_rows[-1]})"
        ws[f"{col}{pos['administration']}"] = f"=SUM({col}{administration_rows[0]}:{col}{administration_rows[-1]})"
        ws[f"{col}{pos['technology']}"] = f"=SUM({col}{technology_rows[0]}:{col}{technology_rows[-1]})"
        ws[f"{col}{pos['otros_gastos']}"] = f'=SUMIFS(\'g-expenses-inc\'!$K:$K,\'g-expenses-inc\'!$A:$A,{col}$1,\'g-expenses-inc\'!$D:$D,"otros_gastos")'
        ws[f"{col}{pos['opex']}"] = f"={col}{pos['shared_services']}+{col}{pos['administration']}+{col}{pos['technology']}+{col}{pos['otros_gastos']}"
        ws[f"{col}{pos['expenses']}"] = f"={col}{pos['cogs']}+{col}{pos['opex']}"
        ws[f"{col}{pos['profit']}"] = f"={col}{pos['turnover']}-{col}{pos['cogs']}-{col}{pos['opex']}-{col}{pos['diferencias_divisas']}"
        ws[f"{col}{pos['profit_pct']}"] = f'=IFERROR({col}{pos["profit"]}/{col}{pos["product_sales"]},0)'

    for row_idx in range(4, pos["profit_pct"] + 1):
        ws[f"P{row_idx}"] = f"=SUM(D{row_idx}:O{row_idx})"
    # Percentage rows: use ratio on totals, not sum of monthly percentages
    ws[f"P{pos['manufacturing_pct']}"]       = f'=IFERROR(P{pos["manufacturing"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['logistics_pct']}"]           = f'=IFERROR(P{pos["logistics"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['payment_fees_pct']}"]        = f'=IFERROR(P{pos["payment_fees"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['gross_margin_pct']}"]        = f'=IFERROR(P{pos["gross_margin"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['contributive_margin_pct']}"] = f'=IFERROR(P{pos["contributive_margin"]}/P{pos["product_sales"]},0)'
    ws[f"P{pos['profit_pct']}"]              = f'=IFERROR(P{pos["profit"]}/P{pos["product_sales"]},0)'
    # Stock rows (informational)
    for key in ("stock_inicial", "stock_final"):
        if key in pos:
            ws[f"P{pos[key]}"] = f"=SUM(D{pos[key]}:O{pos[key]})"


def _count_sheet_inc(wb: Workbook, bundle: PygIncDataBundle) -> None:
    """Mirror P&G-INC structure but show invoice counts instead of amounts."""
    from collections import defaultdict

    exp_sc: dict[tuple[str, str], dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in bundle.expense_rows:
        exp_sc[(r.subcategory, r.supplier_code)][r.yyyymm] += 1

    fee: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in bundle.payment_fee_rows:
        fee[r.supplier_code][r.yyyymm] += 1

    sales_by_market: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in bundle.sales_rows:
        sales_by_market[r.line_item][r.yyyymm] += 1

    all_yyyymm: set[str] = set()
    for d in list(exp_sc.values()) + list(fee.values()) + list(sales_by_market.values()):
        all_yyyymm.update(k for k, v in d.items() if v > 0)
    all_yyyymm.update(bundle.otros_ingresos_by_period.keys())
    all_yyyymm.update(bundle.diferencias_divisas_by_period.keys())
    valid_months = [m for m in all_yyyymm if m.startswith(str(bundle.year))]
    last_yyyymm = max(valid_months) if valid_months else None
    last_month_num = int(last_yyyymm[4:6]) if last_yyyymm else None
    prev_month_num = (last_month_num - 1) if last_month_num and last_month_num > 1 else None
    last_col_idx = (3 + last_month_num) if last_month_num else None
    prev_col_idx = (3 + prev_month_num) if prev_month_num else None

    ws = wb.create_sheet("Nº Facturas-INC")
    months = month_keys(bundle.year)
    for idx, yyyymm in enumerate(months, start=4):
        ws.cell(row=1, column=idx, value=yyyymm)
        ws.cell(row=2, column=idx, value=MONTH_NAMES_ES[idx - 4])
    ws["P1"] = "TOTAL"
    ws["P2"] = "Total"
    ws["A2"] = f"Nº Facturas {COMPANY_CODE} {bundle.year}"
    ws.merge_cells("A2:C2")
    for r in (1, 2):
        for cell in ws[r]:
            cell.fill = HEADER_FILL
            cell.font = WHITE_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = TITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 1
    ws.row_dimensions[2].height = ROW_HEIGHT

    pos: dict[str, int] = {}
    leaf_rows: list[int] = []

    def lbl(col: str, r: int, text: str, bold: bool = False) -> None:
        ws[f"{col}{r}"] = text
        if bold:
            ws[f"{col}{r}"].font = BOLD

    def write_counts(r: int, count_fn) -> None:
        total = 0
        for idx, yyyymm in enumerate(months, start=4):
            v = count_fn(yyyymm)
            ws.cell(row=r, column=idx, value=v)
            total += v
        ws[f"P{r}"] = total

    def write_sum_rows(r: int, child_rows: list[int]) -> None:
        for idx in range(4, 16):
            ws.cell(row=r, column=idx, value=sum(ws.cell(row=cr, column=idx).value or 0 for cr in child_rows))
        ws[f"P{r}"] = sum(ws[f"P{cr}"].value or 0 for cr in child_rows)

    row = 4
    pos["turnover"] = row; lbl("A", row, "Turnover", bold=True); row += 1
    pos["product_sales"] = row; lbl("B", row, "Product sales", bold=True); row += 1
    pos["shopify_header"] = row; lbl("C", row, "Shopify", bold=True); row += 1
    sales_rows_list = list(range(row, row + len(DEFAULT_SALES_MARKETS)))
    for idx, market in enumerate(DEFAULT_SALES_MARKETS):
        lbl("C", row + idx, market)
    row += len(DEFAULT_SALES_MARKETS)
    pos["otros_ingresos"] = row; lbl("B", row, "Otros ingresos", bold=True); row += 2

    pos["expenses"] = row; lbl("A", row, "Expenses", bold=True); row += 1
    pos["cogs"] = row; lbl("B", row, "COGS", bold=True); row += 1
    pos["manufacturing_header"] = row; lbl("C", row, "Manufacturing", bold=True); row += 1
    manufacturing_rows_list = list(range(row, row + len(DEFAULT_MANUFACTURING_LINES)))
    for idx, supplier in enumerate(DEFAULT_MANUFACTURING_LINES):
        lbl("C", row + idx, supplier)
    row += len(DEFAULT_MANUFACTURING_LINES)

    pos["logistics_header"] = row; lbl("C", row, "Logistics", bold=True); row += 1
    logistics_rows_list = list(range(row, row + len(DEFAULT_LOGISTICS_LINES)))
    for idx, supplier in enumerate(DEFAULT_LOGISTICS_LINES):
        lbl("C", row + idx, supplier)
    row += len(DEFAULT_LOGISTICS_LINES)

    pos["payment_fees_header"] = row; lbl("C", row, "Payment fees", bold=True); row += 1
    payment_fee_rows_list = list(range(row, row + len(DEFAULT_PAYMENT_FEE_LINES)))
    for idx, code in enumerate(DEFAULT_PAYMENT_FEE_LINES):
        lbl("C", row + idx, code)
    row += len(DEFAULT_PAYMENT_FEE_LINES)
    row += 1  # blank before opex

    pos["opex"] = row; lbl("B", row, "Opex", bold=True); row += 1
    pos["shared_services_header"] = row; lbl("C", row, "Shared services", bold=True); row += 1
    shared_service_rows_list = list(range(row, row + len(DEFAULT_SHARED_SERVICE_LINES)))
    for idx, supplier in enumerate(DEFAULT_SHARED_SERVICE_LINES):
        lbl("C", row + idx, supplier)
    row += len(DEFAULT_SHARED_SERVICE_LINES)

    pos["administration_header"] = row; lbl("C", row, "Administration", bold=True); row += 1
    administration_rows_list = list(range(row, row + len(DEFAULT_ADMIN_LINES)))
    for idx, supplier in enumerate(DEFAULT_ADMIN_LINES):
        lbl("C", row + idx, supplier)
    row += len(DEFAULT_ADMIN_LINES)

    pos["technology_header"] = row; lbl("C", row, "Technology", bold=True); row += 1
    technology_rows_list = list(range(row, row + len(DEFAULT_TECH_LINES)))
    for idx, supplier in enumerate(DEFAULT_TECH_LINES):
        lbl("C", row + idx, supplier)
    row += len(DEFAULT_TECH_LINES)
    pos["otros_gastos"] = row; lbl("C", row, "Otros gastos", bold=True); row += 1
    pos["diferencias_divisas"] = row; lbl("B", row, "Diferencias divisas", bold=True); row += 1
    max_data_row = row - 1

    # Write counts
    for r, market in zip(sales_rows_list, DEFAULT_SALES_MARKETS):
        write_counts(r, lambda yyyymm, m=market: sales_by_market[m][yyyymm])
        leaf_rows.append(r)

    write_counts(pos["otros_ingresos"], lambda yyyymm: 1 if yyyymm in bundle.otros_ingresos_by_period else 0)
    leaf_rows.append(pos["otros_ingresos"])

    for r, supplier in zip(manufacturing_rows_list, DEFAULT_MANUFACTURING_LINES):
        write_counts(r, lambda yyyymm, s=supplier: exp_sc[("manufacturing", s)][yyyymm])
        leaf_rows.append(r)

    for r, supplier in zip(logistics_rows_list, DEFAULT_LOGISTICS_LINES):
        write_counts(r, lambda yyyymm, s=supplier: exp_sc[("logistics", s)][yyyymm])
        leaf_rows.append(r)

    for r, code in zip(payment_fee_rows_list, DEFAULT_PAYMENT_FEE_LINES):
        write_counts(r, lambda yyyymm, c=code: fee[c][yyyymm])
        leaf_rows.append(r)

    for r, supplier in zip(shared_service_rows_list, DEFAULT_SHARED_SERVICE_LINES):
        write_counts(r, lambda yyyymm, s=supplier: exp_sc[("shared_services", s)][yyyymm])
        leaf_rows.append(r)

    for r, supplier in zip(administration_rows_list, DEFAULT_ADMIN_LINES):
        write_counts(r, lambda yyyymm, s=supplier: exp_sc[("administration", s)][yyyymm])
        leaf_rows.append(r)

    for r, supplier in zip(technology_rows_list, DEFAULT_TECH_LINES):
        write_counts(r, lambda yyyymm, s=supplier: exp_sc[("technology", s)][yyyymm])
        leaf_rows.append(r)

    otros_gastos_counts: dict[str, int] = defaultdict(int)
    for (sc, _), counts in exp_sc.items():
        if sc == "otros_gastos":
            for yyyymm, cnt in counts.items():
                otros_gastos_counts[yyyymm] += cnt
    write_counts(pos["otros_gastos"], lambda yyyymm: otros_gastos_counts[yyyymm])
    leaf_rows.append(pos["otros_gastos"])

    write_counts(pos["diferencias_divisas"], lambda yyyymm: 1 if yyyymm in bundle.diferencias_divisas_by_period else 0)
    leaf_rows.append(pos["diferencias_divisas"])

    # Aggregate rows
    write_sum_rows(pos["shopify_header"], sales_rows_list)
    write_sum_rows(pos["product_sales"], sales_rows_list)
    write_sum_rows(pos["turnover"], [pos["product_sales"], pos["otros_ingresos"]])
    write_sum_rows(pos["manufacturing_header"], manufacturing_rows_list)
    write_sum_rows(pos["logistics_header"], logistics_rows_list)
    write_sum_rows(pos["payment_fees_header"], payment_fee_rows_list)
    write_sum_rows(pos["cogs"], manufacturing_rows_list + logistics_rows_list + payment_fee_rows_list)
    write_sum_rows(pos["shared_services_header"], shared_service_rows_list)
    write_sum_rows(pos["administration_header"], administration_rows_list)
    write_sum_rows(pos["technology_header"], technology_rows_list)
    write_sum_rows(pos["opex"], [pos["shared_services_header"], pos["administration_header"], pos["technology_header"], pos["otros_gastos"]])
    write_sum_rows(pos["expenses"], [pos["cogs"], pos["opex"]])

    # Styling
    major_rows = {pos["turnover"], pos["expenses"]}
    subtotal_rows = {pos["product_sales"], pos["cogs"], pos["opex"]}
    section_rows = {
        pos["shopify_header"], pos["manufacturing_header"], pos["logistics_header"],
        pos["payment_fees_header"], pos["shared_services_header"], pos["administration_header"],
        pos["technology_header"], pos["otros_gastos"], pos["otros_ingresos"], pos["diferencias_divisas"],
    }
    for r in range(4, max_data_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT
        if r in major_rows:
            for c in range(1, 18):
                ws.cell(row=r, column=c).fill = SECTION_FILL
        elif r in section_rows:
            for c in range(1, 18):
                ws.cell(row=r, column=c).fill = SUBSECTION_FILL
        elif r in subtotal_rows:
            for c in range(1, 18):
                ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
    for c, w in (("A", 18), ("B", 20), ("C", 24)):
        ws.column_dimensions[c].width = w
    ws.column_dimensions["P"].width = 12
    for i in range(4, 17):
        ws.column_dimensions[get_column_letter(i)].width = 7
    ws.freeze_panes = "D4"
    ws.sheet_view.showGridLines = False

    if last_col_idx and prev_col_idx:
        red_fill = PatternFill("solid", fgColor="FFAAAA")
        for r in leaf_rows:
            last_val = ws.cell(row=r, column=last_col_idx).value or 0
            prev_val = ws.cell(row=r, column=prev_col_idx).value or 0
            if prev_val > 0 and (last_val == 0 or last_val < prev_val / 2):
                ws.cell(row=r, column=last_col_idx).fill = red_fill


def _apply_inc_layout(ws, *, pos: dict[str, int], detail_rows: list[int]) -> None:
    major_rows = {pos["turnover"], pos["expenses"], pos["gross_margin"], pos["contributive_margin"], pos["profit"]}
    subtotal_rows = {pos["product_sales"], pos["cogs"], pos["opex"]}
    section_rows = {pos["shopify"], pos["manufacturing"], pos["logistics"], pos["payment_fees"], pos["shared_services"], pos["administration"], pos["technology"], pos["otros_gastos"], pos["otros_ingresos"], pos["diferencias_divisas"]}
    percent_rows = {pos["manufacturing_pct"], pos["logistics_pct"], pos["payment_fees_pct"], pos["gross_margin_pct"], pos["contributive_margin_pct"], pos["profit_pct"]}

    for row_idx in range(2, pos["profit_pct"] + 1):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT
        for col_idx in range(1, 18):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx in major_rows:
                cell.fill = SECTION_FILL if row_idx in {pos["turnover"], pos["expenses"]} else KPI_FILL
                cell.border = MEDIUM_TOP_BORDER
                if col_idx == 1:
                    cell.font = BOLD
            elif row_idx in subtotal_rows:
                cell.fill = SUBSECTION_FILL
                cell.border = THIN_TOP_BORDER
                if col_idx == 1:
                    cell.font = BOLD
            elif row_idx in section_rows:
                cell.fill = SUBSECTION_FILL
                if col_idx == 1:
                    cell.font = BOLD
            elif row_idx in percent_rows:
                if col_idx == 1:
                    cell.font = Font(italic=True, color="666666")
        if row_idx in detail_rows:
            ws.row_dimensions[row_idx].outlineLevel = 3
            ws.row_dimensions[row_idx].hidden = True
        elif row_idx in section_rows:
            ws.row_dimensions[row_idx].outlineLevel = 2
            ws.row_dimensions[row_idx].hidden = True
        elif row_idx in subtotal_rows:
            ws.row_dimensions[row_idx].outlineLevel = 1
    for row_idx in percent_rows:
        ws.row_dimensions[row_idx].outlineLevel = 2 if row_idx in {pos["manufacturing_pct"], pos["logistics_pct"], pos["payment_fees_pct"]} else 1
        ws.row_dimensions[row_idx].hidden = True


def _add_navigation_links(
    ws,
    *,
    pos: dict[str, int],
    sales_rows: list[int],
    manufacturing_rows: list[int],
    logistics_rows: list[int],
    payment_fee_rows: list[int],
    shared_service_rows: list[int],
    administration_rows: list[int],
    technology_rows: list[int],
) -> None:
    links = {
        pos["shopify"]: "'i-shopify-inc'!A1",
        sales_rows[0]: "'i-shopify-inc'!A1",
        pos["manufacturing"]: "'g-expenses-inc'!A1",
        manufacturing_rows[0]: "'g-expenses-inc'!A1",
        pos["logistics"]: "'g-expenses-inc'!A1",
        logistics_rows[0]: "'g-expenses-inc'!A1",
        pos["payment_fees"]: "'g-payment-fees-inc'!A1",
        payment_fee_rows[0]: "'g-payment-fees-inc'!A1",
        pos["shared_services"]: "'g-expenses-inc'!A1",
        shared_service_rows[0]: "'g-expenses-inc'!A1",
        pos["administration"]: "'g-expenses-inc'!A1",
        administration_rows[0]: "'g-expenses-inc'!A1",
        pos["technology"]: "'g-expenses-inc'!A1",
        technology_rows[0]: "'g-expenses-inc'!A1",
        pos["otros_gastos"]: "'g-expenses-inc'!A1",
    }
    for row_idx, location in links.items():
        cell = ws[f"Q{row_idx}"]
        cell.value = "->"
        cell._hyperlink = Hyperlink(ref=cell.coordinate, location=location, display="->")


def _map_expense_subcategory(*, supplier_code: str, division_invoice: str, supplier_meta: dict[str, Any] | None) -> str:
    supplier = supplier_code.upper()
    division = division_invoice.strip().lower()
    if supplier == "TGI":
        if division in {"manufacturing", "logistics"}:
            return division
        return "manufacturing"
    if supplier == "CONTINUUM":
        return "administration"
    if supplier == "REGUS":
        return "administration"
    if supplier == "JONDO":
        return "manufacturing"
    if supplier == "PORTCLEARANCE":
        return "logistics"
    if supplier == "PROCO":
        if division in {"manufacturing", "logistics"}:
            return division
        return "manufacturing"
    if supplier == "SHAREDSERVICESSL":
        return "shared_services"
    if supplier == "YOURACCOUNTSTAXES":
        return "administration"
    if supplier == "REVER":
        return "technology"
    destination_path = str((supplier_meta or {}).get("destination_path") or "")
    if destination_path.startswith("expenses/"):
        parts = destination_path.split("/", 2)
        if len(parts) == 3:
            subcategory = parts[2].replace("-", "_")
            if subcategory == "manufacturing_logistics":
                return "logistics"
            return subcategory
    return ""


def _decimal(value: Any) -> Decimal:
    return value if isinstance(value, Decimal) else Decimal(str(value or 0))

