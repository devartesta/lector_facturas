from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import lector_facturas.fx_rates as fx_rates
from lector_facturas.pyg_sl_workbook import PygSlDataBundle, ProviderCatalogRow, build_pyg_sl_workbook


ECB_SAMPLE_XML = b"""<?xml version="1.0" encoding="UTF-8"?>
<gesmes:Envelope xmlns:gesmes="http://www.gesmes.org/xml/2002-08-01" xmlns="http://www.ecb.int/vocabulary/2002-08-01/eurofxref">
  <Cube>
    <Cube time="2026-01-30">
      <Cube currency="USD" rate="1.2000"/>
      <Cube currency="GBP" rate="0.8000"/>
    </Cube>
    <Cube time="2026-02-27">
      <Cube currency="USD" rate="1.2500"/>
      <Cube currency="GBP" rate="0.8200"/>
    </Cube>
    <Cube time="2026-03-31">
      <Cube currency="USD" rate="1.3000"/>
      <Cube currency="GBP" rate="0.8400"/>
    </Cube>
  </Cube>
</gesmes:Envelope>
"""


class _FakeResponse:
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False

    def read(self) -> bytes:
        return ECB_SAMPLE_XML


def _find_row(ws, label: str, column: str = "C") -> int:
    for row in range(1, ws.max_row + 1):
        value = ws[f"{column}{row}"].value
        if isinstance(value, str) and value.replace("\u00A0", " ").strip() == label.strip():
            return row
    raise AssertionError(f"Label not found: {label}")


def test_build_pyg_sl_workbook_creates_expected_sheets_and_formulas(tmp_path: Path) -> None:
    fx_rates.EcbFxService._daily_rates.cache_clear()
    bundle = PygSlDataBundle(
        year=2026,
        generated_at=datetime(2026, 3, 23, 10, 0, 0),
        shopify_rows=(),
        marketplace_rows=(),
        rappel_rows=(),
        supplies_rows=(),
        service_rows=(),
        expense_rows=(),
        payment_fee_rows=(),
        provider_catalog_rows=(
            ProviderCatalogRow("APPHOTOES", "APPHOTOES", "AP Photo", "expenses/cogs/manufacturing", ""),
            ProviderCatalogRow("GLS", "GLS", "GLS", "expenses/cogs/logistics", ""),
            ProviderCatalogRow("CLARIS", "CLARIS", "CLARIS", "expenses/opex/administration", ""),
            ProviderCatalogRow("ADOBE", "ADOBE", "ADOBE", "expenses/opex/technology", ""),
        ),
        shopify_markets=("ES", "FR"),
    )
    output_path = tmp_path / "pyg_sl_2026.xlsx"

    original_urlopen = fx_rates.urlopen
    fx_rates.urlopen = lambda *args, **kwargs: _FakeResponse()
    try:
        build_pyg_sl_workbook(bundle, output_path)
    finally:
        fx_rates.urlopen = original_urlopen
        fx_rates.EcbFxService._daily_rates.cache_clear()

    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames[0] == "P&G-SL"
    assert "i-marketplaces-sl" in workbook.sheetnames
    assert "g-expenses-sl" in workbook.sheetnames
    assert "fx-rates" in workbook.sheetnames
    ws = workbook["P&G-SL"]
    assert ws["D1"].value == "202601"
    assert ws["D2"].value == "Enero"
    assert ws["P2"].value == "Total"
    assert ws["A4"].value == "Turnover"
    assert ws["C7"].value == "ES"
    assert ws["D6"].value == "=SUM(D7:D8)"
    assert ws["C8"].value == "FR"
    assert ws["C9"].value == "Marketplaces"
    assert ws["D10"].value == "=SUMIFS('i-marketplaces-sl'!$I:$I,'i-marketplaces-sl'!$A:$A,D$1,'i-marketplaces-sl'!$C:$C,$C10)"
    product_sales_row = _find_row(ws, "  Product sales", column="A")
    manufacturing_row = _find_row(ws, "Manufacturing (% sales)", column="C")
    logistics_row = _find_row(ws, "Logistics (% sales)", column="C")
    royalties_eu_row = _find_row(ws, "eu", column="C")
    payment_fees_row = _find_row(ws, "Payment fees (% sales)", column="C")
    contributive_margin_row = _find_row(ws, "Contributive margin (product sales-COGS)", column="A")
    assert ws[f"D{contributive_margin_row}"].value == (
        f"=D{product_sales_row}-D{manufacturing_row}-D{logistics_row}-D{royalties_eu_row}-D{payment_fees_row}"
    )
    params_ws = workbook["params"]
    params = {
        params_ws[f"A{row}"].value: params_ws[f"B{row}"].value
        for row in range(2, params_ws.max_row + 1)
    }
    assert params["pct_staff_uk"] == 0.15
    assert params["pct_staff_us"] == 0.1
    assert params["pct_admin_uk"] == 0.15
    assert params["pct_admin_us"] == 0.1
    fx_ws = workbook["fx-rates"]
    fx_entries = {
        (fx_ws[f"A{row}"].value, fx_ws[f"C{row}"].value, fx_ws[f"D{row}"].value)
        for row in range(2, fx_ws.max_row + 1)
    }
    assert ("202601", "GBP", "EUR") in fx_entries
    assert ("202601", "USD", "EUR") in fx_entries
