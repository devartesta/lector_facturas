from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from lector_facturas.payment_fee_detail_workbook import PaymentFeeDetailBundle, build_payment_fee_detail_workbook
from lector_facturas.payment_fees import PaymentFeeSummaryRow, PaymentOrderTransaction


def test_build_payment_fee_detail_workbook_creates_summary_and_detail_sheets(tmp_path: Path) -> None:
    bundle = PaymentFeeDetailBundle(
        company_code="SL",
        period_yyyymm="202603",
        summaries=(
            PaymentFeeSummaryRow(
                company_code="SL",
                period_yyyymm="202603",
                platform="shopify",
                market_code="SL-EUR",
                currency_code="EUR",
                orders_count=2,
                transactions_count=3,
                gross_amount=Decimal("100.00"),
                fee_amount=Decimal("5.00"),
                chargeback_amount=Decimal("30.00"),
                chargeback_fee_amount=Decimal("15.00"),
                total_cost_amount=Decimal("20.00"),
                net_amount=Decimal("95.00"),
                payout_count=1,
            ),
        ),
        transactions=(
            PaymentOrderTransaction(
                id="tx-1",
                platform="shopify",
                company_code="SL",
                market_code="SL-EUR",
                currency_code="EUR",
                order_id="ord-1",
                order_name="AS-1001",
                external_transaction_id="shopify-1",
                external_payout_id="payout-1",
                transaction_date="2026-03-05T10:00:00Z",
                payout_date="2026-03-07T00:00:00Z",
                transaction_type="charge",
                status="paid",
                gross_amount=Decimal("100.00"),
                fee_amount=Decimal("5.00"),
                chargeback_fee_amount=Decimal("15.00"),
                chargeback_amount=Decimal("30.00"),
                net_amount=Decimal("95.00"),
            ),
        ),
        shopify_raw_rows=(
            {
                "transaction_date": "2026-03-05",
                "type": "charge",
                "order_name": "AS-1001",
                "payout_date": "2026-03-07",
                "payout_id": "payout-1",
                "amount": "100.00",
                "fee": "5.00",
                "net": "95.00",
                "payment_method_name": "visa",
                "currency": "EUR",
                "presentment_amount": "100.00",
                "presentment_currency": "EUR",
            },
        ),
        paypal_raw_rows=(
            {
                "transaction_date": "2026-03-08",
                "shopify_order_name": "AS-1002",
                "tipo": "Payment Received",
                "estado": "Completed",
                "divisa": "EUR",
                "bruto": "50.00",
                "tarifa": "-2.00",
                "neto": "48.00",
                "transaction_id": "pp-1",
                "reference_transaction_id": "pp-ref-1",
            },
        ),
    )
    output_path = tmp_path / "payment_fees_sl_202603.xlsx"

    build_payment_fee_detail_workbook(bundle, output_path)

    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames == ["Summary", "Detail", "Shopify Raw", "PayPal Raw"]
    ws_summary = workbook["Summary"]
    assert ws_summary["A1"].value == "Payment Fees Summary"
    assert ws_summary["B2"].value == "Artesta Store, S.L"
    assert ws_summary["B3"].value == "202603"
    assert ws_summary["A7"].value == "TOTAL"
    assert ws_summary["F6"].value == Decimal("5.00")
    assert ws_summary["I6"].value == Decimal("20.00")
    ws_detail = workbook["Detail"]
    assert ws_detail["A2"].value == "SHOPIFY"
    assert ws_detail["C2"].value == "AS-1001"
    assert ws_detail["N2"].value == Decimal("20.00")
