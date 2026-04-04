from __future__ import annotations

import unittest
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import load_workbook

from lector_facturas.fx_rates import EcbFxService
from lector_facturas.pyg_consolidated_workbook import ConsolidatedPygBundle, _aggregate_all, build_pyg_consolidated_workbook
from lector_facturas.pyg_inc_workbook import PygIncDataBundle, ProviderCatalogRow as IncProviderCatalogRow, _map_expense_subcategory as map_inc_expense_subcategory
from lector_facturas.pyg_ltd_workbook import PygLtdDataBundle
from lector_facturas.pyg_sl_workbook import ExpenseRow, ProviderCatalogRow, PygSlDataBundle, StageRow
from lector_facturas.pyg_snapshot import PygSnapshot, PygSnapshotRow, _build_consolidated_snapshot, _build_simple_company_snapshot, _build_sl_snapshot


def _snapshot_row(code: str, amount: str, *, label: str | None = None) -> PygSnapshotRow:
    value = Decimal(amount)
    return PygSnapshotRow(
        code=code,
        label=label or code,
        level=0,
        kind="detail",
        parent_code=None,
        style_key="detail",
        default_expanded=False,
        values_base=(value,),
        values_eur=(value,),
    )


class PygConsistencyTests(unittest.TestCase):
    def test_sl_snapshot_royalties_use_total_without_double_counting_scope_breakdown(self) -> None:
        bundle = PygSlDataBundle(
            year=2026,
            generated_at=datetime(2026, 4, 4, 12, 0, 0),
            shopify_rows=(),
            marketplace_rows=(),
            rappel_rows=(),
            supplies_rows=(),
            service_rows=(),
            expense_rows=(
                ExpenseRow("202601", "SL", "cogs", "royalties", "ROYALTIES", "", Decimal("10"), "EUR", "test"),
            ),
            payment_fee_rows=(),
            provider_catalog_rows=(),
            shopify_markets=("ES",),
            royalties_by_scope={
                "eu": {"202601": Decimal("6")},
                "uk": {"202601": Decimal("3")},
                "us": {"202601": Decimal("1")},
            },
        )

        with patch("lector_facturas.pyg_snapshot.collect_pyg_sl_data", return_value=bundle):
            snapshot = _build_sl_snapshot(months=["202601"], database_url="postgres://ignored", settings=None)

        rows = {row.code: row for row in snapshot.rows}
        self.assertEqual(rows["royalties_total"].values_eur[0], Decimal("10"))
        self.assertEqual(rows["royalties_eu"].values_eur[0], Decimal("6"))
        self.assertEqual(rows["royalties_uk"].values_eur[0], Decimal("3"))
        self.assertEqual(rows["royalties_us"].values_eur[0], Decimal("1"))
        self.assertEqual(rows["royalties"].values_eur[0], Decimal("10"))

    def test_sl_snapshot_administration_sums_all_first_level_suppliers_without_double_counting_nested_breakdown(self) -> None:
        bundle = PygSlDataBundle(
            year=2026,
            generated_at=datetime(2026, 4, 4, 12, 0, 0),
            shopify_rows=(),
            marketplace_rows=(),
            rappel_rows=(),
            supplies_rows=(),
            service_rows=(),
            expense_rows=(
                ExpenseRow("202601", "SL", "opex", "administration", "BBVACNC", "", Decimal("1644.62"), "EUR", "test"),
                ExpenseRow("202601", "SL", "opex", "administration", "CLARIS", "", Decimal("808.08"), "EUR", "test"),
                ExpenseRow("202601", "SL", "opex", "administration", "HANNUN", "office", Decimal("175.00"), "EUR", "test"),
                ExpenseRow("202601", "SL", "opex", "administration", "HANNUN", "services", Decimal("1504.92"), "EUR", "test"),
                ExpenseRow("202601", "SL", "opex", "administration", "NODA", "", Decimal("170.00"), "EUR", "test"),
                ExpenseRow("202601", "SL", "opex", "administration", "TORRAS", "", Decimal("15.00"), "EUR", "test"),
            ),
            payment_fee_rows=(),
            provider_catalog_rows=(
                ProviderCatalogRow("BBVACNC", "BBVACNC", "", "expenses/opex/administration", ""),
                ProviderCatalogRow("CLARIS", "CLARIS", "", "expenses/opex/administration", ""),
                ProviderCatalogRow("HANNUN", "HANNUN", "", "expenses/opex/administration", ""),
                ProviderCatalogRow("NODA", "NODA", "", "expenses/opex/administration", ""),
                ProviderCatalogRow("TORRAS", "TORRAS", "", "expenses/opex/administration", ""),
            ),
            shopify_markets=("ES",),
        )

        with patch("lector_facturas.pyg_snapshot.collect_pyg_sl_data", return_value=bundle):
            snapshot = _build_sl_snapshot(months=["202601"], database_url="postgres://ignored", settings=None)

        rows = {row.code: row for row in snapshot.rows}
        self.assertEqual(rows["administration_hannun"].values_eur[0], Decimal("1679.92"))
        self.assertEqual(rows["administration_bbvacnc"].values_eur[0], Decimal("1644.62"))
        self.assertEqual(rows["administration"].values_eur[0], Decimal("4317.62"))

    def test_consolidated_snapshot_uses_external_services_and_total_royalties(self) -> None:
        sl_snapshot = PygSnapshot(
            company="sl",
            base_currency="EUR",
            months=("202601",),
            generated_at=datetime(2026, 4, 4, 12, 0, 0),
            drive_file_name="sl.xlsx",
            drive_file_url="",
            fx_mode="monthly_historical",
            rows=(
                _snapshot_row("shopify", "100", label="Shopify"),
                _snapshot_row("marketplaces", "10", label="Marketplaces"),
                _snapshot_row("rappels", "-1", label="Rappels"),
                _snapshot_row("supplies", "-2", label="Supplies"),
                _snapshot_row("otros_ingresos_group", "3", label="Otros ingresos"),
                _snapshot_row("manufacturing", "20", label="Manufacturing"),
                _snapshot_row("manufacturing_bbvacnc", "0", label="BBVACNC"),
                _snapshot_row("logistics", "5", label="Logistics"),
                _snapshot_row("royalties", "20", label="Royalties doubled"),
                _snapshot_row("royalties_total", "10", label="Royalties total"),
                _snapshot_row("payment_fees", "2", label="Payment fees"),
                _snapshot_row("marketing", "4", label="Marketing"),
                _snapshot_row("staff", "6", label="Staff"),
                _snapshot_row("administration", "27", label="Administration"),
                _snapshot_row("administration_bbvacnc", "20", label="BBVACNC administration"),
                _snapshot_row("technology", "8", label="Technology"),
                _snapshot_row("otros_gastos_group", "9", label="Otros gastos"),
                _snapshot_row("diferencias_divisas_group", "1", label="Dif divisas"),
                _snapshot_row("service_hannun", "9", label="HANNUN"),
                _snapshot_row("services_external", "4", label="Services external"),
            ),
        )
        ltd_snapshot = PygSnapshot(
            company="ltd",
            base_currency="GBP",
            months=("202601",),
            generated_at=datetime(2026, 4, 4, 12, 0, 0),
            drive_file_name="ltd.xlsx",
            drive_file_url="",
            fx_mode="monthly_historical",
            rows=(
                _snapshot_row("product_sales", "30"),
                _snapshot_row("otros_ingresos_group", "0"),
                _snapshot_row("manufacturing", "1"),
                _snapshot_row("logistics", "2"),
                _snapshot_row("payment_fees", "3"),
                _snapshot_row("administration", "4"),
                _snapshot_row("technology", "5"),
                _snapshot_row("otros_gastos_group", "6"),
                _snapshot_row("diferencias_divisas_group", "7"),
            ),
        )
        inc_snapshot = PygSnapshot(
            company="inc",
            base_currency="USD",
            months=("202601",),
            generated_at=datetime(2026, 4, 4, 12, 0, 0),
            drive_file_name="inc.xlsx",
            drive_file_url="",
            fx_mode="monthly_historical",
            rows=(
                _snapshot_row("product_sales", "40"),
                _snapshot_row("otros_ingresos_group", "0"),
                _snapshot_row("manufacturing", "1"),
                _snapshot_row("logistics", "1"),
                _snapshot_row("payment_fees", "1"),
                _snapshot_row("administration", "1"),
                _snapshot_row("technology", "1"),
                _snapshot_row("otros_gastos_group", "1"),
                _snapshot_row("diferencias_divisas_group", "1"),
            ),
        )
        sl_bundle = PygSlDataBundle(
            year=2026,
            generated_at=datetime(2026, 4, 4, 12, 0, 0),
            shopify_rows=(),
            marketplace_rows=(),
            rappel_rows=(),
            supplies_rows=(),
            service_rows=(
                StageRow("202601", "SL", "HANNUN", "services", Decimal("4"), "EUR", "test"),
                StageRow("202601", "SL", "HANNUN", "renting_cnc", Decimal("5"), "EUR", "test"),
            ),
            expense_rows=(),
            payment_fee_rows=(),
            provider_catalog_rows=(),
            shopify_markets=("ES",),
        )

        with (
            patch("lector_facturas.pyg_snapshot._build_sl_snapshot", return_value=sl_snapshot),
            patch("lector_facturas.pyg_snapshot._build_ltd_snapshot", return_value=ltd_snapshot),
            patch("lector_facturas.pyg_snapshot._build_inc_snapshot", return_value=inc_snapshot),
            patch("lector_facturas.pyg_snapshot.collect_pyg_sl_data", return_value=sl_bundle),
        ):
            snapshot = _build_consolidated_snapshot(months=["202601"], database_url="postgres://ignored", settings=None)

        rows = {row.code: row for row in snapshot.rows}
        self.assertEqual(rows["services"].values_eur[0], Decimal("4"))
        self.assertEqual(rows["services"].level, 1)
        self.assertEqual(rows["otros_ingresos"].label, "Uncategorized income")
        self.assertEqual(rows["otros_gastos_group"].label, "Uncategorized Expenses")
        self.assertEqual(rows["diferencias_divisas_group"].label, "Currency Adjustment")
        self.assertEqual(rows["royalties"].values_eur[0], Decimal("10"))
        self.assertEqual(rows["administration"].values_eur[0], Decimal("12"))
        self.assertEqual(rows["shopify"].values_eur[0], Decimal("170"))
        self.assertEqual(rows["product_sales"].values_eur[0], Decimal("177"))
        self.assertEqual(rows["turnover"].values_eur[0], Decimal("184"))

    def test_simple_company_snapshot_keeps_income_hierarchy_with_services_bucket(self) -> None:
        bundle = PygIncDataBundle(
            year=2026,
            generated_at=datetime(2026, 4, 4, 12, 0, 0),
            sales_rows=(),
            expense_rows=(),
            payment_fee_rows=(),
            provider_catalog_rows=(),
            otros_ingresos_by_period={"202601": Decimal("11")},
        )

        snapshot = _build_simple_company_snapshot(
            company="inc",
            reporting_currency="USD",
            months=["202601"],
            database_url="postgres://ignored",
            settings=None,
            collect_bundle=lambda **_: bundle,
            expense_mapper=map_inc_expense_subcategory,
            sales_markets=("US",),
            manufacturing_lines=("JONDO",),
            logistics_lines=("TGI",),
            payment_fee_lines=("SHOPIFY",),
            shared_service_lines=("SHAREDSERVICESSL",),
            administration_lines=("CONTINUUM",),
            technology_lines=("REVER",),
            file_name="pyg_inc_2026.xlsx",
        )

        rows = {row.code: row for row in snapshot.rows}
        self.assertEqual(rows["services"].label, "Services")
        self.assertEqual(rows["services"].level, 1)
        self.assertEqual(rows["services"].values_eur[0], Decimal("0"))
        self.assertEqual(rows["otros_ingresos_group"].label, "Uncategorized income")
        self.assertEqual(rows["otros_ingresos"].label, "Uncategorized income")
        self.assertEqual(rows["turnover"].values_eur[0], Decimal("11"))

    def test_consolidated_workbook_turnover_formula_does_not_double_count_product_sales_components(self) -> None:
        bundle = ConsolidatedPygBundle(year=2026, generated_at=datetime(2026, 4, 4, 12, 0, 0))

        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "pyg_consolidado_2026.xlsx"
            build_pyg_consolidated_workbook(bundle, output_path)
            workbook = load_workbook(output_path, data_only=False)

        ws = workbook["P&G-CONSOLIDADO"]
        self.assertEqual(ws["D4"].value, "=D5+D8+D11")

    def test_consolidated_workbook_aggregate_includes_frame_consumption_in_manufacturing(self) -> None:
        bundle = ConsolidatedPygBundle(
            year=2026,
            generated_at=datetime(2026, 4, 4, 12, 0, 0),
            ltd_bundle=PygLtdDataBundle(
                year=2026,
                generated_at=datetime(2026, 4, 4, 12, 0, 0),
                sales_rows=(),
                expense_rows=(),
                payment_fee_rows=(),
                provider_catalog_rows=(),
                frame_consumed_by_period={"202601": Decimal("12.5")},
            ),
            inc_bundle=PygIncDataBundle(
                year=2026,
                generated_at=datetime(2026, 4, 4, 12, 0, 0),
                sales_rows=(),
                expense_rows=(),
                payment_fee_rows=(),
                provider_catalog_rows=(),
                frame_consumed_by_period={"202601": Decimal("7.25")},
            ),
        )

        _sl_rows, ltd_rows, inc_rows, _fx_rows = _aggregate_all(bundle, fx=EcbFxService())

        self.assertIn(["202601", "manufacturing", 12.5], ltd_rows)
        self.assertIn(["202601", "manufacturing", 7.25], inc_rows)

    def test_inc_snapshot_includes_delaware_in_administration(self) -> None:
        bundle = PygIncDataBundle(
            year=2026,
            generated_at=datetime(2026, 4, 4, 12, 0, 0),
            sales_rows=(),
            expense_rows=(
                ExpenseRow("202603", "INC", "opex", "administration", "CONTINUUM", "", Decimal("1000"), "USD", "test"),
                ExpenseRow("202603", "INC", "opex", "administration", "DELAWARE", "", Decimal("99"), "USD", "test"),
            ),
            payment_fee_rows=(),
            provider_catalog_rows=(
                IncProviderCatalogRow("CONTINUUM", "CONTINUUM", "", "expenses/opex/administration", ""),
                IncProviderCatalogRow("DELAWARE", "DELAWARE", "", "expenses/opex/administration", ""),
            ),
        )

        with patch("lector_facturas.pyg_snapshot.collect_pyg_inc_data", return_value=bundle):
            snapshot = _build_simple_company_snapshot(
                company="inc",
                reporting_currency="USD",
                months=["202603"],
                database_url="postgres://ignored",
                settings=None,
                collect_bundle=lambda **_: bundle,
                expense_mapper=map_inc_expense_subcategory,
                sales_markets=("US",),
                manufacturing_lines=("JONDO", "TGI"),
                logistics_lines=("TGI",),
                payment_fee_lines=("SHOPIFY",),
                shared_service_lines=("SHAREDSERVICESSL",),
                administration_lines=("CONTINUUM", "DELAWARE", "HUSHED", "IPOSTAL", "QUICKBOOKS", "REGUS"),
                technology_lines=("REVER",),
                file_name="pyg_inc_2026.xlsx",
            )

        rows = {row.code: row for row in snapshot.rows}
        self.assertEqual(rows["administration_delaware"].values_base[0], Decimal("99"))
        self.assertEqual(
            rows["administration"].values_eur[0],
            rows["administration_continuum"].values_eur[0] + rows["administration_delaware"].values_eur[0],
        )


if __name__ == "__main__":
    unittest.main()
