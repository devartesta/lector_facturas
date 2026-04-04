from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from lector_facturas.pyg_ltd_workbook import PygLtdDataBundle, ProviderCatalogRow, build_pyg_ltd_workbook


def _find_row(ws, label: str) -> int:
    for row in range(1, ws.max_row + 1):
        value = ws[f"A{row}"].value
        if isinstance(value, str) and value.replace("\u00A0", " ").strip() == label.strip():
            return row
    raise AssertionError(f"Label not found: {label}")


def test_build_pyg_ltd_workbook_creates_expected_sheets_and_formulas(tmp_path: Path) -> None:
    bundle = PygLtdDataBundle(
        year=2026,
        generated_at=datetime(2026, 3, 23, 10, 0, 0),
        sales_rows=(),
        expense_rows=(),
        payment_fee_rows=(),
        royalties_by_period={"202601": 12.34},
        provider_catalog_rows=(
            ProviderCatalogRow("PROCO", "PROCO", "PROCO", "expenses/cogs/manufacturing-logistics", ""),
            ProviderCatalogRow("JONDO", "JONDO", "JONDO", "expenses/cogs/manufacturing-logistics", ""),
            ProviderCatalogRow("SHAREDSERVICESSL", "SHAREDSERVICESSL", "SHAREDSERVICESSL", "expenses/opex/shared-services", ""),
            ProviderCatalogRow("REVER", "REVER", "REVER", "expenses/opex/technology", ""),
        ),
    )
    output_path = tmp_path / "pyg_ltd_2026.xlsx"

    build_pyg_ltd_workbook(bundle, output_path)

    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames[0] == "P&G-LTD"
    assert "i-shopify-ltd" in workbook.sheetnames
    assert "g-expenses-ltd" in workbook.sheetnames
    assert "fx-rates" in workbook.sheetnames
    ws = workbook["P&G-LTD"]
    assert ws["D1"].value == "202601"
    assert ws["D2"].value == "Enero"
    assert ws["P2"].value == "Total"
    assert ws["A4"].value == "Turnover"
    assert ws["A7"].value.strip() == "GB"
    assert ws["D6"].value == "=SUM(D7:D7)"
    assert ws["A14"].value.strip() == "ARTLINK"
    assert ws["A15"].value.strip() == "JONDO"
    assert ws["D14"].value == "=SUMIFS('g-expenses-ltd'!$K:$K,'g-expenses-ltd'!$A:$A,D$1,'g-expenses-ltd'!$D:$D,\"manufacturing\",'g-expenses-ltd'!$E:$E,TRIM($A14))"
    assert "% Manufacturing / sales" in str(ws["A18"].value)
    assert ws["D18"].value == '=IFERROR(D13/D5,0)'
    royalties_row = _find_row(ws, "    Royalties")
    cogs_row = _find_row(ws, "  COGS")
    manufacturing_row = _find_row(ws, "    Manufacturing")
    logistics_row = _find_row(ws, "    Logistics")
    payment_fees_row = _find_row(ws, "    Payment fees")
    contributive_margin_row = _find_row(ws, "Contributive margin (product sales-COGS)")
    assert ws[f"D{royalties_row}"].value == 12.34
    assert ws[f"D{cogs_row}"].value == f"=D{manufacturing_row}+D{logistics_row}+D{payment_fees_row}"
    assert ws[f"D{contributive_margin_row}"].value == f"=D5-D{cogs_row}-D{royalties_row}"
