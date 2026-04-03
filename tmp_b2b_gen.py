import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Existing notes from original file (col18)
NOTES = {
    'AS-97386': 'x', 'AS-97387': 'x', 'AS-97422': 'x',
    'AS-97680': 'Falta', 'AS-98317': 'Quiron', 'AS-98341': 'x',
    'AS-98379': 'x', 'AS-98540': 'x', 'AS-98882': 'x',
    'AS-98965': 'Quiron', 'AS-98971': 'x', 'AS-99006': 'x',
    'AS-99054': 'x', 'AS-99071': 'x', 'AS-99183': 'x',
    'AS-99224': 'x', 'AS-99271': 'x', 'AS-99303': 'x',
    'AS-99612': 'Falta Almond', 'AS-99613': 'Falta Almond', 'AS-99752': 'Falta Almond',
}

# Data from DB (order, date, country, category, vat_pct, base, vat, total)
ROWS = [
    ('AS-97386', '02/02/2026', 'ES', 'b2b',      0.21, 369.22,  77.54,  446.76),
    ('AS-97387', '02/02/2026', 'ES', 'b2b',      0.21, 452.08,  94.94,  547.02),
    ('AS-97422', '02/02/2026', 'ES', 'b2b',      0.21, 143.21,  30.07,  173.28),
    ('AS-97680', '05/02/2026', 'ES', 'Shooting', 0.21, 204.52,  42.94,  247.46),
    ('AS-98317', '11/02/2026', 'ES', 'b2b',      0.21,  77.21,  16.21,   93.42),
    ('AS-98341', '11/02/2026', 'ES', 'b2b',      0.21,3120.69, 655.34, 3776.03),
    ('AS-98379', '11/02/2026', 'PT', 'b2b',      0.23,1281.50, 294.74, 1576.24),
    ('AS-98540', '13/02/2026', 'ES', 'b2b',      0.21, 379.10,  79.61,  458.71),
    ('AS-98882', '16/02/2026', 'ES', 'b2b',      0.21, 701.32, 147.28,  848.60),
    ('AS-98965', '17/02/2026', 'ES', 'b2b',      0.21,  56.45,  11.85,   68.30),
    ('AS-98971', '17/02/2026', 'ES', 'b2b',      0.21,  31.58,   6.63,   38.21),
    ('AS-99006', '17/02/2026', 'ES', 'b2b',      0.21, 261.15,  54.84,  315.99),
    ('AS-99054', '18/02/2026', 'ES', 'b2b',      0.21,2737.81, 574.94, 3312.75),
    ('AS-99071', '18/02/2026', 'ES', 'b2b',      0.21, 409.23,  85.94,  495.17),
    ('AS-99183', '19/02/2026', 'ES', 'b2b',      0.21, 286.41,  60.14,  346.55),
    ('AS-99224', '19/02/2026', 'ES', 'b2b',      0.21, 254.09,  53.36,  307.45),
    ('AS-99271', '20/02/2026', 'ES', 'b2b',      0.21, 259.78,  54.55,  314.33),
    ('AS-99303', '20/02/2026', 'ES', 'b2b',      0.21,  95.47,  20.05,  115.52),
    ('AS-99612', '23/02/2026', 'ES', 'b2b',      0.21, 190.93,  40.10,  231.03),
    ('AS-99613', '23/02/2026', 'ES', 'b2b',      0.21, 497.04, 104.38,  601.42),
    ('AS-99752', '24/02/2026', 'ES', 'b2b',      0.21, 686.60, 144.19,  830.79),
]

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='1F4E78')  # dark blue  – title
COL_FILL  = PatternFill('solid', fgColor='D6E4F0')  # light blue – col headers
TOT_FILL  = PatternFill('solid', fgColor='BDD7EE')  # blue-grey  – totals
YES_FILL  = PatternFill('solid', fgColor='E2EFDA')  # green      – paid
NO_FILL   = PatternFill('solid', fgColor='FCE4D6')  # orange     – not paid / pending
PEND_FILL = PatternFill('solid', fgColor='FFF2CC')  # yellow     – pending/unclear
LEG_FILL  = PatternFill('solid', fgColor='F2F2F2')  # grey       – summary

THIN_S = Side(style='thin', color='BFBFBF')
THIN   = Border(top=THIN_S, bottom=THIN_S, left=THIN_S, right=THIN_S)
C      = Alignment(horizontal='center', vertical='center')
L      = Alignment(horizontal='left',   vertical='center')
R      = Alignment(horizontal='right',  vertical='center')
LW     = Alignment(horizontal='left',   vertical='top', wrap_text=True)
MFMT   = '#,##0.00'
PCFMT  = '0%'

def bold(color='000000', size=10):
    return Font(bold=True, name='Calibri', size=size, color=color)

NORMAL = Font(name='Calibri', size=10)
ITALIC_G = Font(name='Calibri', size=9, italic=True, color='595959')

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Cobros B2B'

HEADERS = [
    'Nº Pedido', 'Fecha pedido', 'País', 'Categoría',
    'IVA %', 'Base imponible', 'IVA', 'Total (con IVA)',
    'Notas', 'Pagado', 'Fecha cobro', 'Comentarios',
]
WIDTHS = [14, 12, 6, 12, 6, 16, 12, 16, 24, 12, 14, 32]
NCOLS = len(HEADERS)
LAST  = get_column_letter(NCOLS)

for i, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = 1

# Title -----------------------------------------------------------------------
ws.row_dimensions[row].height = 22
ws.merge_cells(f'A{row}:{LAST}{row}')
c = ws.cell(row=row, column=1,
    value='Seguimiento cobros B2B — Artesta Store, S.L — Febrero 2026')
c.font = bold('FFFFFF', 12)
c.fill = HDR_FILL; c.alignment = C
row += 1

# Subtitle --------------------------------------------------------------------
ws.row_dimensions[row].height = 22
ws.merge_cells(f'A{row}:{LAST}{row}')
c = ws.cell(row=row, column=1,
    value=('Pedidos con pago por transferencia bancaria (gateway manual). '
           'Actualizar las columnas "Pagado" y "Fecha cobro" a medida que lleguen los pagos. '
           'Verde = cobrado · Amarillo = pendiente / incidencia.'))
c.font = ITALIC_G; c.alignment = LW
row += 1

# Column headers --------------------------------------------------------------
ws.row_dimensions[row].height = 15
for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=row, column=col, value=h)
    c.font = bold(); c.fill = COL_FILL; c.alignment = C; c.border = THIN
row += 1
first_data = row

# Data rows -------------------------------------------------------------------
for order, fecha, pais, cat, iva_pct, base, iva, total in ROWS:
    nota_raw = NOTES.get(order, '')
    pagado   = 'Sí' if nota_raw == 'x' else 'Pendiente'
    nota_vis = '' if nota_raw == 'x' else nota_raw
    row_fill = YES_FILL if pagado == 'Sí' else PEND_FILL

    ws.row_dimensions[row].height = 14
    vals = [
        (order,     L,    None,  NORMAL),
        (fecha,     C,    None,  NORMAL),
        (pais,      C,    None,  NORMAL),
        (cat,       C,    None,  NORMAL),
        (iva_pct,   C,    PCFMT, NORMAL),
        (base,      R,    MFMT,  NORMAL),
        (iva,       R,    MFMT,  NORMAL),
        (total,     R,    MFMT,  bold()),
        (nota_vis,  L,    None,  ITALIC_G if nota_vis else NORMAL),
        (pagado,    C,    None,  bold()),   # Pagado
        (None,      C,    None,  NORMAL),   # Fecha cobro
        ('',        L,    None,  NORMAL),   # Comentarios
    ]
    for col, (val, align, fmt, font) in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font; c.alignment = align; c.border = THIN
        if fmt: c.number_format = fmt
        # Row background: full row tinted based on status
        c.fill = row_fill
    row += 1

last_data = row - 1

# Totals row ------------------------------------------------------------------
ws.row_dimensions[row].height = 15
tot_base  = sum(r[5] for r in ROWS)
tot_iva   = sum(r[6] for r in ROWS)
tot_total = sum(r[7] for r in ROWS)
tot_cob   = sum(r[7] for r in ROWS if NOTES.get(r[0]) == 'x')
tot_pend  = tot_total - tot_cob

tot_row = [
    (f'TOTAL  ({len(ROWS)} pedidos)', L, None),
    (None, C, None), (None, C, None), (None, C, None), (None, C, None),
    (tot_base,  R, MFMT),
    (tot_iva,   R, MFMT),
    (tot_total, R, MFMT),
    (None,L,None),(None,C,None),(None,C,None),(None,L,None),
]
for col, (val, align, fmt) in enumerate(tot_row, 1):
    c = ws.cell(row=row, column=col, value=val)
    c.font = bold(); c.fill = TOT_FILL; c.alignment = align; c.border = THIN
    if fmt: c.number_format = fmt
row += 2

# Summary mini-table ----------------------------------------------------------
for label, amount, fill in [
    ('Cobrado',            tot_cob,  YES_FILL),
    ('Pendiente de cobro', tot_pend, PEND_FILL),
]:
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f'A{row}:E{row}')
    c = ws.cell(row=row, column=1, value=label)
    c.font = bold(); c.fill = fill; c.alignment = L; c.border = THIN
    c2 = ws.cell(row=row, column=6, value=amount)
    c2.font = bold(); c2.fill = fill; c2.alignment = R
    c2.border = THIN; c2.number_format = MFMT
    # pct of total
    pct = amount / tot_total if tot_total else 0
    c3 = ws.cell(row=row, column=7, value=pct)
    c3.font = bold(); c3.fill = fill; c3.alignment = C
    c3.border = THIN; c3.number_format = '0.0%'
    for col in range(8, NCOLS + 1):
        cx = ws.cell(row=row, column=col)
        cx.fill = fill; cx.border = THIN
    row += 1

# Data validation for Pagado (col 10) ----------------------------------------
dv = DataValidation(
    type='list',
    formula1='"Si,No,Pendiente"',
    allow_blank=False,
    showDropDown=False,
)
dv.sqref = f'J{first_data}:J{last_data}'
ws.add_data_validation(dv)

ws.freeze_panes = 'A4'

# Save ------------------------------------------------------------------------
out = r'C:\Users\AdriàSebastià\OneDrive - Artesta\ARTESTA - 6. Finances\Artesta Store, S.L\2026\1Q\202602\Ingresos\VENTAS\202602_Cobros_B2B.xlsx'
wb.save(out)
print(f'Guardado: {out}')
print(f'Total: {tot_total:,.2f} EUR  |  Cobrado: {tot_cob:,.2f}  |  Pendiente: {tot_pend:,.2f}')
